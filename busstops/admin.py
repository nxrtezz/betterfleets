import csv
import re
from datetime import timedelta
from functools import lru_cache
from io import BytesIO, StringIO
from itertools import pairwise
from pathlib import Path

import requests
from django import forms
from django.conf import settings
from django.forms import ModelForm, Textarea
from django.contrib import admin, messages
from django.contrib.gis.geos import LineString
from django.contrib.admin.helpers import ACTION_CHECKBOX_NAME
from django.contrib.gis.admin import GISModelAdmin
from django.contrib.postgres.search import SearchQuery, SearchRank
from django.core.exceptions import PermissionDenied
from django.http import HttpResponse
from django.db import connection, transaction
from django.db.models import CharField, Count, Exists, F, OuterRef, Prefetch, Q, Value
from django.db.models.aggregates import StringAgg
from django.db.models.functions import Cast
from django.template.response import TemplateResponse
from django.urls import path, reverse
from django.utils import timezone
from django.utils.html import format_html, format_html_join
from django.utils.safestring import mark_safe
from django.utils.text import slugify
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from shapely.geometry import LineString as ShapelyLineString, Point as ShapelyPoint
from sql_util.utils import SubqueryCount
from bustimes.admin import log_change
from bustimes.models import Calendar, Garage, Route, RouteLink, StopTime, Trip, VehicleType as TimetableVehicleType
from fleet.parsers.pdf_fleet_parser import TARGET_COLUMNS, parse_pdf
from vehicles.models import Livery, Vehicle, VehicleFeature, VehicleType, vehicle_slug

from . import models
from . import views as busstops_views
from .data_changes import apply_pending_change, reject_pending_change
from .fleet_imports import (
    create_garage_for_operator as shared_create_garage_for_operator,
    rows_text_from_upload as shared_rows_text_from_upload,
)


@lru_cache(maxsize=1)
def vehicle_db_columns():
    try:
        with connection.cursor() as cursor:
            return {
                column.name
                for column in connection.introspection.get_table_description(
                    cursor, Vehicle._meta.db_table
                )
            }
    except Exception:
        return {
            field.column
            for field in Vehicle._meta.fields
            if getattr(field, "column", None)
        }


def current_vehicle_filters(**filters):
    columns = vehicle_db_columns()
    unsupported = {
        "withdrawn": "withdrawn",
        "preserved": "preserved",
        "operator": "operator_id",
        "operator_id": "operator_id",
    }
    for key, column in unsupported.items():
        if key in filters and column not in columns:
            filters.pop(key)
    if "historical_fleet_id" in columns:
        filters["historical_fleet__isnull"] = True
    return filters


@admin.register(models.AdminArea)
class AdminAreaAdmin(admin.ModelAdmin):
    list_display = ("name", "id", "atco_code", "region_id")
    list_filter = ("region_id",)
    search_fields = ("atco_code",)


class StopCodeInline(admin.TabularInline):
    model = models.StopCode
    raw_id_fields = ["source"]


@admin.register(models.StopPoint)
class StopPointAdmin(GISModelAdmin):
    list_display = [
        "atco_code",
        "naptan_code",
        "crs_code",
        "locality",
        "admin_area",
        "common_name",
        "modified_at",
        "created_at",
    ]
    list_select_related = ["locality", "admin_area"]
    list_filter = [
        ("source", admin.RelatedOnlyFieldListFilter),
        "modified_at",
        "created_at",
        "active",
        "stop_type",
        "service__region",
        "admin_area",
    ]
    raw_id_fields = ["source", "parents", "stop_area", "locality", "admin_area"]
    search_fields = ["atco_code", "crs_code", "common_name"]
    ordering = ["atco_code"]
    inlines = [StopCodeInline]
    show_full_result_count = False
    readonly_fields = ["search_vector"]
    filter_horizontal = ["features"]

    def get_search_results(self, request, queryset, search_term):
        if not search_term:
            return super().get_search_results(request, queryset, search_term)

        query = SearchQuery(search_term, search_type="websearch", config="english")
        rank = SearchRank(F("locality__search_vector"), query)
        query = Q(locality__search_vector=query)
        if " " not in search_term:
            query |= Q(atco_code=search_term)
        queryset = queryset.annotate(rank=rank).filter(query).order_by("-rank")
        return queryset, False


@admin.register(models.StopFeature)
class StopFeatureAdmin(admin.ModelAdmin):
    list_display = ("name", "category")
    list_filter = ("category",)
    search_fields = ("name",)


@admin.register(models.StopCode)
class StopCodeAdmin(admin.ModelAdmin):
    list_display = ["stop", "code", "source"]
    raw_id_fields = ["stop"]


@admin.register(models.StopArea)
class StopAreaAdmin(GISModelAdmin):
    raw_id_fields = ["admin_area", "parent"]
    list_filter = ["stop_area_type"]


class StopGroupStopInline(admin.TabularInline):
    model = models.StopGroupStop
    extra = 1
    autocomplete_fields = ["stop"]


class StopGroupAdminForm(forms.ModelForm):
    stops_selection = forms.CharField(required=False, widget=forms.HiddenInput())

    class Meta:
        model = models.StopGroup
        fields = "__all__"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        if self.instance.pk:
            selected = self.instance.stopgroupstop_set.order_by(
                "order", "stop__common_name", "stop__indicator"
            ).values_list("stop_id", flat=True)
            self.fields["stops_selection"].initial = ",".join(selected)

    def clean_stops_selection(self):
        raw_value = self.cleaned_data.get("stops_selection", "")
        stop_ids = []
        seen = set()
        for stop_id in re.split(r"[\s,]+", raw_value.strip()):
            if not stop_id or stop_id in seen:
                continue
            seen.add(stop_id)
            stop_ids.append(stop_id)

        if not stop_ids:
            return []

        existing = set(
            models.StopPoint.objects.filter(atco_code__in=stop_ids).values_list(
                "atco_code", flat=True
            )
        )
        missing = [stop_id for stop_id in stop_ids if stop_id not in existing]
        if missing:
            raise forms.ValidationError(
                f"Unknown stop code(s): {', '.join(missing[:10])}"
            )

        return stop_ids


@admin.register(models.StopGroup)
class StopGroupAdmin(GISModelAdmin):
    form = StopGroupAdminForm
    change_form_template = "admin/busstops/stopgroup/change_form.html"
    list_display = ("name", "slug", "active", "stop_count", "modified_at")
    list_filter = ("active",)
    search_fields = ("name", "slug", "stops__atco_code", "stops__common_name")
    prepopulated_fields = {"slug": ("name",)}
    readonly_fields = ("stops_map",)
    fieldsets = (
        (None, {"fields": ("name", "slug", "location", "active", "stops_map")}),
    )

    @admin.display(description="Stops")
    def stop_count(self, obj):
        return obj.stops.count()

    @admin.display(description="Stop selector")
    def stops_map(self, obj):
        selected_rows = []
        location = ""
        if obj and obj.pk:
            selected_rows = list(
                obj.stopgroupstop_set.select_related("stop", "stop__locality")
                .order_by("order", "stop__common_name", "stop__indicator")
            )
            if obj.location:
                lng, lat = obj.location.coords
                location = format_html(
                    ' data-location-lng="{}" data-location-lat="{}"',
                    lng,
                    lat,
                )

        selected_list = "".join(
            format_html(
                '<li data-stop-id="{}" data-stop-name="{}"><button type="button" class="stop-group-selector-remove" data-stop-id="{}">Remove</button> {} <span class="quiet">{}</span></li>',
                row.stop_id,
                row.stop.get_qualified_name(),
                row.stop_id,
                row.stop.get_qualified_name(),
                row.stop_id,
            )
            for row in selected_rows
        )
        if not selected_list:
            selected_list = '<li class="empty">No stops selected yet.</li>'

        return mark_safe(
            (
                '<div class="stop-group-selector"'
                ' data-stops-url="/stops.json"'
                ' data-default-style="alidade_smooth"'
                f' data-selected-stop-ids="{",".join(row.stop_id for row in selected_rows)}"'
                f"{location}>"
                "<p>Select stops from the map below. Pan and zoom to load nearby stops, then click a stop to add or remove it from the group.</p>"
                '<div class="stop-group-selector-status">Zoom in to load stops.</div>'
                '<div class="stop-group-selector-map" id="stop-group-selector-map"></div>'
                '<h3>Selected stops</h3>'
                f'<ol class="stop-group-selector-list">{selected_list}</ol>'
                "</div>"
            )
        )

    def save_related(self, request, form, formsets, change):
        super().save_related(request, form, formsets, change)
        selected_stop_ids = form.cleaned_data.get("stops_selection", [])
        group = form.instance

        if not selected_stop_ids:
            group.stopgroupstop_set.all().delete()
            return

        existing = {
            row.stop_id: row
            for row in group.stopgroupstop_set.filter(stop_id__in=selected_stop_ids)
        }
        group.stopgroupstop_set.exclude(stop_id__in=selected_stop_ids).delete()

        to_create = []
        to_update = []
        for order, stop_id in enumerate(selected_stop_ids):
            row = existing.get(stop_id)
            if row is None:
                to_create.append(
                    models.StopGroupStop(group=group, stop_id=stop_id, order=order)
                )
            elif row.order != order:
                row.order = order
                to_update.append(row)

        if to_create:
            models.StopGroupStop.objects.bulk_create(to_create)
        if to_update:
            models.StopGroupStop.objects.bulk_update(to_update, ["order"])


@admin.register(models.BustimesSyncState)
class BustimesSyncStateAdmin(admin.ModelAdmin):
    list_display = ("object_type", "external_id", "local_model", "local_pk", "last_synced_at")
    list_filter = ("object_type", "local_model", "last_synced_at")
    search_fields = ("external_id", "local_pk", "local_model")
    readonly_fields = (
        "object_type",
        "external_id",
        "local_model",
        "local_pk",
        "last_fields",
        "last_payload",
        "protected_fields",
        "last_synced_at",
    )


@admin.register(models.DataChangeLog)
class DataChangeLogAdmin(admin.ModelAdmin):
    list_display = (
        "created_at",
        "source",
        "target_model",
        "target_pk",
        "operation",
        "status",
        "approved_by",
        "applied_at",
    )
    list_filter = ("status", "source", "target_model", "operation", "created_at")
    search_fields = ("source", "target_model", "target_pk", "target_repr", "reason")
    readonly_fields = (
        "source",
        "target_model",
        "target_pk",
        "target_repr",
        "operation",
        "changes",
        "payload",
        "status",
        "reason",
        "created_at",
        "applied_at",
        "approved_by",
        "change_summary",
    )
    actions = ("approve_pending_changes", "reject_pending_changes")

    @admin.display(description="Change summary")
    def change_summary(self, obj):
        rows = format_html_join(
            "",
            "<tr><th>{}</th><td>{}</td><td>{}</td></tr>",
            (
                (
                    field,
                    change.get("from", ""),
                    change.get("to", ""),
                )
                for field, change in (obj.changes or {}).items()
            ),
        )
        if not rows:
            return ""
        return format_html(
            "<table><thead><tr><th>Field</th><th>Current/manual</th><th>Imported</th></tr></thead><tbody>{}</tbody></table>",
            rows,
        )

    @admin.action(description="Approve and apply selected pending data changes")
    def approve_pending_changes(self, request, queryset):
        applied = 0
        for log in queryset.filter(status=models.DataChangeLog.STATUS_PENDING):
            apply_pending_change(log, user=request.user)
            applied += 1
        self.message_user(request, f"Applied {applied} pending data change(s).")

    @admin.action(description="Reject selected pending data changes")
    def reject_pending_changes(self, request, queryset):
        rejected = 0
        for log in queryset.filter(status=models.DataChangeLog.STATUS_PENDING):
            reject_pending_change(log, user=request.user, reason="Rejected in admin.")
            rejected += 1
        self.message_user(request, f"Rejected {rejected} pending data change(s).")

    def has_add_permission(self, request):
        return False


class OperatorCodeInline(admin.TabularInline):
    model = models.OperatorCode


class OperatorVehicleColumnInline(admin.TabularInline):
    model = models.OperatorVehicleColumn
    extra = 0
    fields = ("name", "slug", "help_text", "display_order")


class OperatorGroupDepotInline(admin.TabularInline):
    model = models.OperatorGroupDepot
    extra = 0
    fields = ("name", "address", "notes")
    show_change_link = True


class ChangeNOCForm(forms.Form):
    new_noc = forms.CharField(
        max_length=10,
        label="New NOC",
        help_text="Enter the new NOC (National Operator Code) for this operator"
    )


class OperatorAdminForm(ModelForm):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        if self.instance and self.instance.pk:
            self.fields['noc'].disabled = True

    class Meta:
        widgets = {
            "address": Textarea,
            "twitter": Textarea,
        }


class OperatorVehicleInline(admin.TabularInline):
    model = Vehicle
    fk_name = "operator"
    extra = 1
    fields = (
        "code",
        "fleet_number",
        "fleet_code",
        "reg",
        "prev_registration",
        "vehicle_type",
        "livery",
        "garage",
        "branding",
        "name",
        "notes",
        "withdrawn",
    )
    autocomplete_fields = ("vehicle_type", "livery", "garage")
    show_change_link = True

    def get_queryset(self, request):
        return super().get_queryset(request).filter(**current_vehicle_filters())


@admin.register(models.BlogTag)
class BlogTagAdmin(admin.ModelAdmin):
    list_display = ("name", "slug")
    prepopulated_fields = {"slug": ("name",)}
    search_fields = ("name",)


@admin.register(models.BlogPost)
class BlogPostAdmin(admin.ModelAdmin):
    list_display = ("title", "published", "published_at", "updated_at")
    list_filter = ("published", "tags")
    prepopulated_fields = {"slug": ("title",)}
    search_fields = ("title", "excerpt", "body")
    autocomplete_fields = ("tags",)

    def save_model(self, request, obj, form, change):
        if obj.published and not obj.published_at:
            obj.published_at = timezone.now()
        super().save_model(request, obj, form, change)


class MassAddBusesForm(forms.Form):
    rows_text = forms.CharField(
        required=False,
        widget=forms.Textarea(attrs={"rows": 18, "cols": 120}),
        help_text=(
            "Paste CSV or TSV with headers such as code, fleet_num, registration, "
            "prev_registration, operator_code, vehicle_type, livery, garage, name, notes, withdrawn, features."
        ),
    )
    workbook = forms.FileField(
        required=False,
        widget=forms.ClearableFileInput(
            attrs={"accept": ".xlsx,.csv,text/csv,application/vnd.ms-excel"}
        ),
        help_text="Upload a completed .xlsx or .csv file instead of pasting rows.",
    )



class MassEditBusesForm(MassAddBusesForm):
    pass


# HistoricalFleet model was removed - this form is no longer functional
# class NewHistoricalFleetForm(MassAddBusesForm):
#     pass


class MassAddRoutesForm(forms.Form):
    rows_text = forms.CharField(
        required=False,
        widget=forms.Textarea(attrs={"rows": 18, "cols": 120}),
        help_text=(
            "Paste CSV or TSV with headers such as service_code, line_name, "
            "description, line_brand, mode, current, public_use, and colour."
        ),
    )
    workbook = forms.FileField(
        required=False,
        widget=forms.ClearableFileInput(
            attrs={"accept": ".xlsx,.csv,text/csv,application/vnd.ms-excel"}
        ),
        help_text="Upload a completed .xlsx or .csv file instead of pasting rows.",
    )


class MassEditTimetableForm(forms.Form):
    rows_text = forms.CharField(
        required=False,
        widget=forms.Textarea(attrs={"rows": 18, "cols": 140}),
        help_text=(
            "Paste CSV or TSV with headers such as import_key, trip_id, route_id, "
            "calendar_id, inbound, sequence, stop_atco_code, arrival, departure, "
            "pick_up, set_down."
        ),
    )
    workbook = forms.FileField(
        required=False,
        widget=forms.ClearableFileInput(
            attrs={"accept": ".pdf,.xlsx,.csv,text/csv,application/vnd.ms-excel,application/pdf"}
        ),
        help_text="Upload a fleet PDF, completed .xlsx, or .csv file instead of pasting rows.",
    )


class OperatorBulkAssignOrganisationForm(forms.Form):
    organisation = forms.ModelChoiceField(
        queryset=models.Organisation.objects.order_by("name"),
        required=True,
        label="Major Operator",
        help_text="Choose the major operator to apply to all selected operators.",
    )


class OperatorBulkAssignGroupForm(forms.Form):
    group = forms.ModelChoiceField(
        queryset=models.OperatorGroup.objects.order_by("name"),
        required=True,
        label="Division",
        help_text="Choose the division to apply to all selected operators.",
    )


class OperatorBulkAssignGovernmentAuthorityForm(forms.Form):
    government_authority = forms.ModelChoiceField(
        queryset=models.GovernmentAuthority.objects.order_by("name"),
        required=True,
        label="Government Authority",
        help_text="Choose the government authority to apply to all selected operators.",
    )


class ServiceBulkAssignOperatorForm(forms.Form):
    operator = forms.ModelChoiceField(
        queryset=models.Operator.objects.order_by("name"),
        required=True,
        help_text="Choose the operator to add to all selected services.",
    )


class DuplicateOperatorFilter(admin.SimpleListFilter):
    title = "duplicate"
    parameter_name = "duplicate"

    def lookups(self, request, model_admin):
        return ((1, "Yes"),)

    def queryset(self, request, queryset):
        if self.value():
            exists = Exists(
                models.Operator.objects.filter(
                    ~Q(pk=OuterRef("pk")),
                    name=OuterRef("name"),
                )
            )
            queryset = queryset.filter(exists)

        return queryset


@admin.register(models.Organisation)
class OrganisationAdmin(admin.ModelAdmin):
    list_display = ("name", "slug", "short_name", "website")
    search_fields = ("name", "slug")
    prepopulated_fields = {"slug": ("name",)}
    verbose_name = "Major Operator"
    readonly_fields = ("garages_note",)
    fieldsets = (
        (
            None,
            {
                "fields": (
                    "name",
                    "slug",
                    "short_name",
                    "legal_name",
                    "slogan",
                    "description",
                    "about",
                )
            },
        ),
        (
            "Branding",
            {
                "fields": (
                    "logo",
                    "banner",
                    "header_background",
                    "header_foreground",
                    "accent_colour",
                    "card_background",
                    "button_background",
                    "button_foreground",
                    "custom_css",
                )
            },
        ),
        (
            "Contact",
            {
                "fields": (
                    "website",
                    "email",
                    "phone",
                )
            },
        ),
        (
            "Garages",
            {
                "fields": ("garages_note",),
            },
        ),
        (
            "Social",
            {
                "fields": (
                    "social_x",
                    "social_fb",
                    "social_instagram",
                    "social_linkedin",
                    "social_youtube",
                    "social_tiktok",
                    "social_threads",
                    "social_bluesky",
                    "social_mastodon",
                    "social_other",
                )
            },
        ),
    )

    @admin.display(description="Garages")
    def garages_note(self, obj):
        url = reverse("admin:bustimes_garage_changelist")
        return format_html(
            'Garages hold depot/location data. Manage them from <a href="{}">Garage admin</a>.',
            url,
        )

    def _build_mass_add_template_workbook(self, rows=None):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Vehicles"
        worksheet.append(
            (
                "operator_code",
                "external_id",
                "code",
                "fleet_number",
                "fleet_code",
                "registration",
                "prev_registration",
                "vehicle_type",
                "livery",
                "colours",
                "garage",
                "name",
                "branding",
                "notes",
                "withdrawn",
                "preserved",
                "fleet_support_vehicle",
                "vor",
                "awaiting_delivery",
                "trainer_vehicle",
                "demonstrator",
                "features",
                "slug",
            )
        )
        for row in rows or ():
            worksheet.append(row)
        worksheet.freeze_panes = "A2"

        instructions = workbook.create_sheet("Instructions")
        instructions.append(["Field", "Notes"])
        instructions.append(
            [
                "operator_code",
                "Optional. Operator NOC/slug/operator code. Leave blank to use the first operator in this organisation.",
            ]
        )
        instructions.append(["external_id", "Optional external id for matching or creating vehicles"])
        instructions.append(["code", "Required unless fleet_number or registration is supplied"])
        instructions.append(["fleet_number", "Integer fleet number"])
        instructions.append(["fleet_code", "Optional displayed fleet code"])
        instructions.append(["registration", "Vehicle registration"])
        instructions.append(["prev_registration", "Previous registration"])
        instructions.append(["vehicle_type", "Vehicle type id, external id, or exact name"])
        instructions.append(["livery", "Livery id, external id, or exact name"])
        instructions.append(["colours", "Optional space-separated colour values"])
        instructions.append(["garage", "Garage id, external id, or exact code"])
        instructions.append(["name", "Vehicle name"])
        instructions.append(["branding", "Branding text"])
        instructions.append(["notes", "Free-text notes"])
        instructions.append(["withdrawn", "Boolean: true/false, yes/no, 1/0"])
        instructions.append(["preserved", "Boolean: true/false, yes/no, 1/0"])
        instructions.append(["fleet_support_vehicle", "Boolean: true/false, yes/no, 1/0"])
        instructions.append(["vor", "Boolean: true/false, yes/no, 1/0"])
        instructions.append(["awaiting_delivery", "Boolean: true/false, yes/no, 1/0"])
        instructions.append(["trainer_vehicle", "Boolean: true/false, yes/no, 1/0"])
        instructions.append(["demonstrator", "Boolean: true/false, yes/no, 1/0"])
        instructions.append(["features", "Comma-separated feature names or ids"])
        return workbook

    def _rows_text_from_workbook(self, uploaded_file):
        if not uploaded_file:
            return ""

        filename = (uploaded_file.name or "").lower()
        if filename.endswith(".csv"):
            try:
                content = uploaded_file.read().decode("utf-8-sig")
            except UnicodeDecodeError as exc:
                raise ValueError("CSV upload must be UTF-8 encoded") from exc
            return content.strip()
        if not filename.endswith(".xlsx"):
            raise ValueError("Upload must be a .xlsx or .csv file")

        workbook = load_workbook(uploaded_file, data_only=True)
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return ""

        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        if not any(headers):
            return ""

        output = StringIO()
        writer = csv.writer(output, delimiter="\t", lineterminator="\n")
        writer.writerow(headers)
        for row in rows[1:]:
            values = ["" if value is None else str(value).strip() for value in row[: len(headers)]]
            if any(values):
                writer.writerow(values)
        return output.getvalue()

    def _mass_rows_from_organisation_fleet(self, organisation):
        vehicles = (
            Vehicle.objects.filter(operator__organisation=organisation)
            .select_related("vehicle_type", "livery", "garage", "operator")
            .prefetch_related("features")
            .filter(**current_vehicle_filters(withdrawn=False))
            .order_by("operator__name", "fleet_number", "fleet_code", "reg", "code")
        )
        rows = []
        for vehicle in vehicles:
            rows.append(
                (
                    vehicle.operator.noc,
                    vehicle.external_id or "",
                    vehicle.code or "",
                    vehicle.fleet_number if vehicle.fleet_number is not None else "",
                    vehicle.fleet_code or "",
                    vehicle.reg or "",
                    vehicle.prev_registration or "",
                    vehicle.vehicle_type.name if vehicle.vehicle_type else "",
                    vehicle.livery.name if vehicle.livery else "",
                    vehicle.colours or "",
vehicle.garage.name if vehicle.garage else "",
                    vehicle.name or "",
                    vehicle.branding or "",
                    vehicle.notes or "",
                    "true" if vehicle.withdrawn else "false",
                    "true" if vehicle.preserved else "false",
                    "true" if vehicle.fleet_support_vehicle else "false",
                    "true" if vehicle.vor else "false",
                    "true" if vehicle.awaiting_delivery else "false",
                    "true" if vehicle.trainer_vehicle else "false",
                    "true" if vehicle.demonstrator else "false",
                    ", ".join(feature.name for feature in vehicle.features.all()),
                    vehicle.slug or "",
                )
            )
        return rows

    def mass_add_buses_template_view(self, request, object_id):
        if not request.user.is_superuser:
            raise PermissionDenied

        organisation = self.get_object(request, object_id)
        if organisation is None:
            raise PermissionDenied

        workbook = self._build_mass_add_template_workbook()
        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)

        response = HttpResponse(
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{organisation.slug}-mass-add-template.xlsx"'
        )
        return response

    def mass_add_buses_current_fleet_view(self, request, object_id):
        if not request.user.is_superuser:
            raise PermissionDenied

        organisation = self.get_object(request, object_id)
        if organisation is None:
            raise PermissionDenied

        workbook = self._build_mass_add_template_workbook(
            rows=self._mass_rows_from_organisation_fleet(organisation)
        )
        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)

        response = HttpResponse(
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{organisation.slug}-current-fleet.xlsx"'
        )
        return response

    def _normalise_header(self, header):
        key = header.strip().lower().replace(" ", "_")
        header_aliases = {
            "fleet_num": "fleet_number",
            "fleet": "fleet_number",
            "registration": "reg",
            "prev_reg": "prev_registration",
            "previous_reg": "prev_registration",
            "vehicle_id": "vehicle_type",
            "vehicle_type_id": "vehicle_type",
            "livery_id": "livery",
            "color": "colours",
            "colors": "colours",
            "garage_id": "garage",
            "operator_noc": "operator_code",
            "operator": "operator_code",
        }
        return header_aliases.get(key, key)

    @staticmethod
    def _coerce_bool(value):
        value = value.strip().lower()
        if value in {"", "none", "null"}:
            return None
        if value in {"1", "true", "yes", "y", "on"}:
            return True
        if value in {"0", "false", "no", "n", "off"}:
            return False
        raise ValueError(f"Invalid boolean value '{value}'")

    def _resolve_reference(self, model, value, label):
        if value == "":
            return None
        if value.isdigit():
            try:
                return model.objects.get(pk=int(value))
            except model.DoesNotExist as exc:
                raise ValueError(f"Unknown {label} id '{value}'") from exc

        try:
            return model.objects.get(external_id=value)
        except (model.DoesNotExist, model.MultipleObjectsReturned):
            pass

        if hasattr(model, "name"):
            item = model.objects.filter(name__iexact=value).first()
            if item:
                return item

        if hasattr(model, "code"):
            item = model.objects.filter(code__iexact=value).first()
            if item:
                return item

        raise ValueError(f"Unknown {label} '{value}'")

    def _resolve_operator(self, organisation, value):
        if not value:
            # Default to first operator in the organisation
            first_operator = models.Operator.objects.filter(organisation=organisation).first()
            if first_operator:
                return first_operator
            raise ValueError("No operators in this organisation and no operator_code provided")

        text = str(value).strip()
        if not text:
            first_operator = models.Operator.objects.filter(organisation=organisation).first()
            if first_operator:
                return first_operator
            raise ValueError("No operators in this organisation and no operator_code provided")

        operator = models.Operator.objects.filter(
            Q(noc__iexact=text)
            | Q(slug__iexact=text)
            | Q(operatorcode__code__iexact=text)
        ).first()
        if operator:
            return operator
        raise ValueError(f"Unknown operator_code '{text}'")

    def _parse_mass_rows(self, organisation, rows_text):
        rows = []

        if not rows_text.strip():
            return rows

        delimiter = "\t" if "\t" in rows_text.splitlines()[0] else ","
        reader = csv.DictReader(StringIO(rows_text), delimiter=delimiter)
        if not reader.fieldnames:
            return rows

        for index, original_row in enumerate(reader, start=2):
            mapped = {}
            for key, value in original_row.items():
                if not key:
                    continue
                mapped[self._normalise_header(key)] = (value or "").strip()

            if not any(mapped.values()):
                continue

            row = {
                "row_number": index,
                "raw": mapped,
                "errors": [],
                "action": "skip",
                "operator": None,
                "vehicle": None,
                "values": {},
            }

            try:
                row["operator"] = self._resolve_operator(organisation, mapped.get("operator_code"))
            except ValueError as exc:
                row["errors"].append(str(exc))

            if not row["operator"]:
                row["errors"].append("Could not determine operator")
                rows.append(row)
                continue

            provided_code = mapped.get("code", "")
            provided_fleet_code = mapped.get("fleet_code", "")
            code = provided_code or provided_fleet_code
            fleet_number = None
            if mapped.get("fleet_number"):
                try:
                    fleet_number = int(mapped["fleet_number"])
                except ValueError:
                    row["errors"].append("fleet_number must be an integer")
            elif mapped.get("fleet_number") == "":
                fleet_number = None

            reg = mapped.get("reg", "").upper().replace(" ", "")
            prev_registration = mapped.get("prev_registration", "").upper().replace(" ", "")

            if not code:
                if fleet_number is not None:
                    code = str(fleet_number)
                elif reg:
                    code = reg

            external_id = mapped.get("external_id") or None
            vehicle = None
            if external_id:
                vehicle = Vehicle.objects.filter(external_id=external_id).first()
            if not vehicle and code:
                vehicle = row["operator"].vehicle_set.filter(
                    **current_vehicle_filters(preserved=False)
                ).filter(code__iexact=code).first()
            if not vehicle and reg:
                vehicle = row["operator"].vehicle_set.filter(
                    **current_vehicle_filters(preserved=False)
                ).filter(reg__iexact=reg).first()

            if not vehicle and not code:
                row["errors"].append(
                    "Could not determine vehicle identifier (code/fleet_num/registration)"
                )

            row["vehicle"] = vehicle
            row["action"] = "update" if vehicle else "create"

            if row["action"] == "create":
                if code:
                    row["values"]["code"] = code
            elif provided_code:
                row["values"]["code"] = provided_code
            if provided_fleet_code:
                row["values"]["fleet_code"] = provided_fleet_code
            if reg:
                row["values"]["reg"] = reg
            if prev_registration:
                row["values"]["prev_registration"] = prev_registration
            if mapped.get("name"):
                row["values"]["name"] = mapped["name"]
            if mapped.get("notes"):
                row["values"]["notes"] = mapped["notes"]
            if mapped.get("branding"):
                row["values"]["branding"] = mapped["branding"]
            if mapped.get("colours"):
                row["values"]["colours"] = mapped["colours"]
            if mapped.get("fleet_number"):
                row["values"]["fleet_number"] = fleet_number
            if external_id:
                row["values"]["external_id"] = external_id

            for field, model, label in (
                ("vehicle_type", VehicleType, "vehicle_type"),
                ("livery", Livery, "livery"),
                ("garage", Garage, "garage"),
            ):
                if mapped.get(field):
                    try:
                        row["values"][field] = self._resolve_reference(
                            model, mapped.get(field, ""), label
                        )
                    except ValueError as exc:
                        row["errors"].append(str(exc))

            for bool_field in (
                "withdrawn",
                "preserved",
                "fleet_support_vehicle",
                "vor",
                "awaiting_delivery",
                "trainer_vehicle",
                "demonstrator",
            ):
                if bool_field in mapped and mapped[bool_field]:
                    try:
                        row["values"][bool_field] = self._coerce_bool(mapped[bool_field])
                    except ValueError as exc:
                        row["errors"].append(str(exc))

            if mapped.get("features"):
                feature_names = [f.strip() for f in mapped["features"].split(",")]
                features = []
                for feature_name in feature_names:
                    if not feature_name:
                        continue
                    try:
                        feature = self._resolve_reference(VehicleFeature, feature_name, "feature")
                        features.append(feature)
                    except ValueError:
                        row["errors"].append(f"Unknown feature '{feature_name}'")
                if features:
                    row["values"]["features"] = features

            rows.append(row)

        return rows

    def _commit_mass_rows(self, rows):
        created = 0
        updated = 0
        errors = 0

        for row in rows:
            if row["errors"]:
                errors += 1
                continue

            try:
                with transaction.atomic():
                    vehicle = row["vehicle"] or Vehicle()
                    vehicle.operator = row["operator"]
                    
                    for field, value in row["values"].items():
                        if field == "features":
                            vehicle.features.set(value)
                        else:
                            setattr(vehicle, field, value)

                    vehicle.is_manual = True
                    vehicle.manual_updated_at = timezone.now()
                    vehicle.save()

                    if row["action"] == "create":
                        created += 1
                    else:
                        updated += 1
            except Exception as exc:
                row["errors"].append(str(exc))
                errors += 1

        return created, updated, errors

    def mass_add_buses_view(self, request, object_id):
        if not request.user.is_superuser:
            raise PermissionDenied

        organisation = self.get_object(request, object_id)
        if organisation is None:
            raise PermissionDenied

        rows = []
        created = 0
        updated = 0
        errors = 0

        if request.method == "POST":
            form = MassAddBusesForm(request.POST, request.FILES)
            if form.is_valid():
                rows_text = form.cleaned_data.get("rows_text") or ""
                workbook = form.cleaned_data.get("workbook")
                try:
                    if workbook:
                        rows_text = self._rows_text_from_workbook(workbook)
                except ValueError as exc:
                    form.add_error("workbook", str(exc))

                if not form.errors and not rows_text.strip():
                    form.add_error(None, "Paste rows or upload a completed workbook.")

                if not form.errors:
                    rows = self._parse_mass_rows(organisation, rows_text)
                    form = MassAddBusesForm(initial={"rows_text": rows_text})

                    if request.POST.get("action") == "commit":
                        created, updated, errors = self._commit_mass_rows(rows)
                        if created or updated:
                            self.message_user(
                                request,
                                f"Mass add complete: created {created}, updated {updated}, errors {errors}",
                            )
                        elif errors:
                            self.message_user(
                                request,
                                f"No rows imported. {errors} row(s) had errors.",
                                level=messages.WARNING,
                            )
                    else:
                        self.message_user(
                            request,
                            "Preview generated. Review rows and click Commit import when ready.",
                        )
        else:
            form = MassAddBusesForm()

        context = {
            **self.admin_site.each_context(request),
            "opts": self.model._meta,
            "original": organisation,
            "organisation": organisation,
            "title": f"Mass add buses for {organisation}",
            "form": form,
            "rows": rows,
            "can_commit": any(not row["errors"] for row in rows),
            "created": created,
            "updated": updated,
            "errors": errors,
            "template_download_url": reverse(
                "admin:busstops_organisation_mass_add_buses_template", args=(organisation.pk,)
            ),
            "export_download_url": reverse(
                "admin:busstops_organisation_mass_add_buses_current_fleet", args=(organisation.pk,)
            ),
            "allow_create": True,
            "commit_label": "Commit import",
        }

        return TemplateResponse(
            request,
            "admin/busstops/organisation/mass_buses.html",
            context,
        )

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "<path:object_id>/mass-add-buses/",
                self.admin_site.admin_view(self.mass_add_buses_view),
                name="busstops_organisation_mass_add_buses",
            ),
            path(
                "<path:object_id>/mass-add-buses/template.xlsx",
                self.admin_site.admin_view(self.mass_add_buses_template_view),
                name="busstops_organisation_mass_add_buses_template",
            ),
            path(
                "<path:object_id>/mass-add-buses/current-fleet.xlsx",
                self.admin_site.admin_view(self.mass_add_buses_current_fleet_view),
                name="busstops_organisation_mass_add_buses_current_fleet",
            ),
        ]
        return custom_urls + urls

    @admin.display(description="Mass add buses")
    def mass_add_buses_link(self, obj):
        url = reverse("admin:busstops_organisation_mass_add_buses", args=(obj.pk,))
        return format_html('<a class="button" href="{}">Mass add buses</a>', url)

    def get_readonly_fields(self, request, obj=None):
        fields = list(super().get_readonly_fields(request, obj))
        if request.user.is_superuser and obj:
            fields.append("mass_add_buses_link")
        return fields


@admin.register(models.GovernmentAuthority)
class GovernmentAuthorityAdmin(admin.ModelAdmin):
    list_display = ("name", "slug", "short_name", "website")
    search_fields = ("name", "slug")
    prepopulated_fields = {"slug": ("name",)}
    fieldsets = (
        (
            None,
            {
                "fields": (
                    "name",
                    "slug",
                    "short_name",
                    "legal_name",
                    "slogan",
                    "description",
                    "about",
                )
            },
        ),
        (
            "Branding",
            {
                "fields": (
                    "logo",
                    "banner",
                    "header_background",
                    "header_foreground",
                    "accent_colour",
                    "card_background",
                    "button_background",
                    "button_foreground",
                    "custom_css",
                )
            },
        ),
        (
            "Contact",
            {
                "fields": (
                    "website",
                    "email",
                    "phone",
                )
            },
        ),
        (
            "Social",
            {
                "fields": (
                    "social_x",
                    "social_fb",
                    "social_instagram",
                    "social_linkedin",
                    "social_youtube",
                    "social_tiktok",
                    "social_threads",
                    "social_bluesky",
                    "social_mastodon",
                    "social_other",
                )
            },
        ),
    )

    def _build_mass_add_template_workbook(self, rows=None):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Vehicles"
        worksheet.append(
            (
                "operator_code",
                "external_id",
                "code",
                "fleet_number",
                "fleet_code",
                "registration",
                "prev_registration",
                "vehicle_type",
                "livery",
                "colours",
                "garage",
                "name",
                "branding",
                "notes",
                "withdrawn",
                "preserved",
                "fleet_support_vehicle",
                "vor",
                "awaiting_delivery",
                "trainer_vehicle",
                "demonstrator",
                "features",
                "slug",
            )
        )
        for row in rows or ():
            worksheet.append(row)
        worksheet.freeze_panes = "A2"

        instructions = workbook.create_sheet("Instructions")
        instructions.append(["Field", "Notes"])
        instructions.append(
            [
                "operator_code",
                "Optional. Operator NOC/slug/operator code. Leave blank to use the first operator in this organisation.",
            ]
        )
        instructions.append(["external_id", "Optional external id for matching or creating vehicles"])
        instructions.append(["code", "Required unless fleet_number or registration is supplied"])
        instructions.append(["fleet_number", "Integer fleet number"])
        instructions.append(["fleet_code", "Optional displayed fleet code"])
        instructions.append(["registration", "Vehicle registration"])
        instructions.append(["prev_registration", "Previous registration"])
        instructions.append(["vehicle_type", "Vehicle type id, external id, or exact name"])
        instructions.append(["livery", "Livery id, external id, or exact name"])
        instructions.append(["colours", "Optional space-separated colour values"])
        instructions.append(["garage", "Garage id, external id, or exact code"])
        instructions.append(["name", "Vehicle name"])
        instructions.append(["branding", "Branding text"])
        instructions.append(["notes", "Free-text notes"])
        instructions.append(["withdrawn", "Boolean: true/false, yes/no, 1/0"])
        instructions.append(["preserved", "Boolean: true/false, yes/no, 1/0"])
        instructions.append(["fleet_support_vehicle", "Boolean: true/false, yes/no, 1/0"])
        instructions.append(["vor", "Boolean: true/false, yes/no, 1/0"])
        instructions.append(["awaiting_delivery", "Boolean: true/false, yes/no, 1/0"])
        instructions.append(["trainer_vehicle", "Boolean: true/false, yes/no, 1/0"])
        instructions.append(["demonstrator", "Boolean: true/false, yes/no, 1/0"])
        instructions.append(["features", "Comma-separated feature names or ids"])
        return workbook

    def _rows_text_from_workbook(self, uploaded_file):
        if not uploaded_file:
            return ""

        filename = (uploaded_file.name or "").lower()
        if filename.endswith(".csv"):
            try:
                content = uploaded_file.read().decode("utf-8-sig")
            except UnicodeDecodeError as exc:
                raise ValueError("CSV upload must be UTF-8 encoded") from exc
            return content.strip()
        if not filename.endswith(".xlsx"):
            raise ValueError("Upload must be a .xlsx or .csv file")

        workbook = load_workbook(uploaded_file, data_only=True)
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return ""

        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        if not any(headers):
            return ""

        output = StringIO()
        writer = csv.writer(output, delimiter="\t", lineterminator="\n")
        writer.writerow(headers)
        for row in rows[1:]:
            values = ["" if value is None else str(value).strip() for value in row[: len(headers)]]
            if any(values):
                writer.writerow(values)
        return output.getvalue()

    def _mass_rows_from_organisation_fleet(self, organisation):
        vehicles = (
            Vehicle.objects.filter(operator__organisation=organisation)
            .select_related("vehicle_type", "livery", "garage", "operator")
            .prefetch_related("features")
            .filter(**current_vehicle_filters(withdrawn=False))
            .order_by("operator__name", "fleet_number", "fleet_code", "reg", "code")
        )
        rows = []
        for vehicle in vehicles:
            rows.append(
                (
                    vehicle.operator.noc,
                    vehicle.external_id or "",
                    vehicle.code or "",
                    vehicle.fleet_number if vehicle.fleet_number is not None else "",
                    vehicle.fleet_code or "",
                    vehicle.reg or "",
                    vehicle.prev_registration or "",
                    vehicle.vehicle_type.name if vehicle.vehicle_type else "",
                    vehicle.livery.name if vehicle.livery else "",
                    vehicle.colours or "",
vehicle.garage.name if vehicle.garage else "",
                    vehicle.name or "",
                    vehicle.branding or "",
                    vehicle.notes or "",
                    "true" if vehicle.withdrawn else "false",
                    "true" if vehicle.preserved else "false",
                    "true" if vehicle.fleet_support_vehicle else "false",
                    "true" if vehicle.vor else "false",
                    "true" if vehicle.awaiting_delivery else "false",
                    "true" if vehicle.trainer_vehicle else "false",
                    "true" if vehicle.demonstrator else "false",
                    ", ".join(feature.name for feature in vehicle.features.all()),
                    vehicle.slug or "",
                )
            )
        return rows

    def mass_add_buses_template_view(self, request, object_id):
        if not request.user.is_superuser:
            raise PermissionDenied

        organisation = self.get_object(request, object_id)
        if organisation is None:
            raise PermissionDenied

        workbook = self._build_mass_add_template_workbook()
        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)

        response = HttpResponse(
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{organisation.slug}-mass-add-template.xlsx"'
        )
        return response

    def mass_add_buses_current_fleet_view(self, request, object_id):
        if not request.user.is_superuser:
            raise PermissionDenied

        organisation = self.get_object(request, object_id)
        if organisation is None:
            raise PermissionDenied

        workbook = self._build_mass_add_template_workbook(
            rows=self._mass_rows_from_organisation_fleet(organisation)
        )
        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)

        response = HttpResponse(
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{organisation.slug}-current-fleet.xlsx"'
        )
        return response

    def _normalise_header(self, header):
        key = header.strip().lower().replace(" ", "_")
        header_aliases = {
            "fleet_num": "fleet_number",
            "fleet": "fleet_number",
            "registration": "reg",
            "prev_reg": "prev_registration",
            "previous_reg": "prev_registration",
            "vehicle_id": "vehicle_type",
            "vehicle_type_id": "vehicle_type",
            "livery_id": "livery",
            "color": "colours",
            "colors": "colours",
            "garage_id": "garage",
            "operator_noc": "operator_code",
            "operator": "operator_code",
        }
        return header_aliases.get(key, key)

    @staticmethod
    def _coerce_bool(value):
        value = value.strip().lower()
        if value in {"", "none", "null"}:
            return None
        if value in {"1", "true", "yes", "y", "on"}:
            return True
        if value in {"0", "false", "no", "n", "off"}:
            return False
        raise ValueError(f"Invalid boolean value '{value}'")

    def _resolve_reference(self, model, value, label):
        if value == "":
            return None
        if value.isdigit():
            try:
                return model.objects.get(pk=int(value))
            except model.DoesNotExist as exc:
                raise ValueError(f"Unknown {label} id '{value}'") from exc

        try:
            return model.objects.get(external_id=value)
        except (model.DoesNotExist, model.MultipleObjectsReturned):
            pass

        if hasattr(model, "name"):
            item = model.objects.filter(name__iexact=value).first()
            if item:
                return item

        if hasattr(model, "code"):
            item = model.objects.filter(code__iexact=value).first()
            if item:
                return item

        raise ValueError(f"Unknown {label} '{value}'")

    def _resolve_operator(self, organisation, value):
        if not value:
            # Default to first operator in the organisation
            first_operator = models.Operator.objects.filter(organisation=organisation).first()
            if first_operator:
                return first_operator
            raise ValueError("No operators in this organisation and no operator_code provided")

        text = str(value).strip()
        if not text:
            first_operator = models.Operator.objects.filter(organisation=organisation).first()
            if first_operator:
                return first_operator
            raise ValueError("No operators in this organisation and no operator_code provided")

        operator = models.Operator.objects.filter(
            Q(noc__iexact=text)
            | Q(slug__iexact=text)
            | Q(operatorcode__code__iexact=text)
        ).first()
        if operator:
            return operator
        raise ValueError(f"Unknown operator_code '{text}'")

    def _parse_mass_rows(self, organisation, rows_text):
        rows = []

        if not rows_text.strip():
            return rows

        delimiter = "\t" if "\t" in rows_text.splitlines()[0] else ","
        reader = csv.DictReader(StringIO(rows_text), delimiter=delimiter)
        if not reader.fieldnames:
            return rows

        for index, original_row in enumerate(reader, start=2):
            mapped = {}
            for key, value in original_row.items():
                if not key:
                    continue
                mapped[self._normalise_header(key)] = (value or "").strip()

            if not any(mapped.values()):
                continue

            row = {
                "row_number": index,
                "raw": mapped,
                "errors": [],
                "action": "skip",
                "operator": None,
                "vehicle": None,
                "values": {},
            }

            try:
                row["operator"] = self._resolve_operator(organisation, mapped.get("operator_code"))
            except ValueError as exc:
                row["errors"].append(str(exc))

            if not row["operator"]:
                row["errors"].append("Could not determine operator")
                rows.append(row)
                continue

            provided_code = mapped.get("code", "")
            provided_fleet_code = mapped.get("fleet_code", "")
            code = provided_code or provided_fleet_code
            fleet_number = None
            if mapped.get("fleet_number"):
                try:
                    fleet_number = int(mapped["fleet_number"])
                except ValueError:
                    row["errors"].append("fleet_number must be an integer")
            elif mapped.get("fleet_number") == "":
                fleet_number = None

            reg = mapped.get("reg", "").upper().replace(" ", "")
            prev_registration = mapped.get("prev_registration", "").upper().replace(" ", "")

            if not code:
                if fleet_number is not None:
                    code = str(fleet_number)
                elif reg:
                    code = reg

            external_id = mapped.get("external_id") or None
            vehicle = None
            if external_id:
                vehicle = Vehicle.objects.filter(external_id=external_id).first()
            if not vehicle and code:
                vehicle = row["operator"].vehicle_set.filter(
                    **current_vehicle_filters(preserved=False)
                ).filter(code__iexact=code).first()
            if not vehicle and reg:
                vehicle = row["operator"].vehicle_set.filter(
                    **current_vehicle_filters(preserved=False)
                ).filter(reg__iexact=reg).first()

            if not vehicle and not code:
                row["errors"].append(
                    "Could not determine vehicle identifier (code/fleet_num/registration)"
                )

            row["vehicle"] = vehicle
            row["action"] = "update" if vehicle else "create"

            if row["action"] == "create":
                if code:
                    row["values"]["code"] = code
            elif provided_code:
                row["values"]["code"] = provided_code
            if provided_fleet_code:
                row["values"]["fleet_code"] = provided_fleet_code
            if reg:
                row["values"]["reg"] = reg
            if prev_registration:
                row["values"]["prev_registration"] = prev_registration
            if mapped.get("name"):
                row["values"]["name"] = mapped["name"]
            if mapped.get("notes"):
                row["values"]["notes"] = mapped["notes"]
            if mapped.get("branding"):
                row["values"]["branding"] = mapped["branding"]
            if mapped.get("colours"):
                row["values"]["colours"] = mapped["colours"]
            if mapped.get("fleet_number"):
                row["values"]["fleet_number"] = fleet_number
            if external_id:
                row["values"]["external_id"] = external_id

            for field, model, label in (
                ("vehicle_type", VehicleType, "vehicle_type"),
                ("livery", Livery, "livery"),
                ("garage", Garage, "garage"),
            ):
                if mapped.get(field):
                    try:
                        row["values"][field] = self._resolve_reference(
                            model, mapped.get(field, ""), label
                        )
                    except ValueError as exc:
                        row["errors"].append(str(exc))

            for bool_field in (
                "withdrawn",
                "preserved",
                "fleet_support_vehicle",
                "vor",
                "awaiting_delivery",
                "trainer_vehicle",
                "demonstrator",
            ):
                if bool_field in mapped and mapped[bool_field]:
                    try:
                        row["values"][bool_field] = self._coerce_bool(mapped[bool_field])
                    except ValueError as exc:
                        row["errors"].append(str(exc))

            if mapped.get("features"):
                feature_names = [f.strip() for f in mapped["features"].split(",")]
                features = []
                for feature_name in feature_names:
                    if not feature_name:
                        continue
                    try:
                        feature = self._resolve_reference(VehicleFeature, feature_name, "feature")
                        features.append(feature)
                    except ValueError:
                        row["errors"].append(f"Unknown feature '{feature_name}'")
                if features:
                    row["values"]["features"] = features

            rows.append(row)

        return rows

    def _commit_mass_rows(self, rows):
        created = 0
        updated = 0
        errors = 0

        for row in rows:
            if row["errors"]:
                errors += 1
                continue

            try:
                with transaction.atomic():
                    vehicle = row["vehicle"] or Vehicle()
                    for field, value in row["values"].items():
                        if field == "features":
                            vehicle.features.set(value)
                        else:
                            setattr(vehicle, field, value)

                    vehicle.is_manual = True
                    vehicle.manual_updated_at = timezone.now()
                    vehicle.save()

                    if row["action"] == "create":
                        created += 1
                    else:
                        updated += 1
            except Exception as exc:
                row["errors"].append(str(exc))
                errors += 1

        return created, updated, errors

    def mass_add_buses_view(self, request, object_id):
        if not request.user.is_superuser:
            raise PermissionDenied

        organisation = self.get_object(request, object_id)
        if organisation is None:
            raise PermissionDenied

        rows = []
        created = 0
        updated = 0
        errors = 0

        if request.method == "POST":
            form = MassAddBusesForm(request.POST, request.FILES)
            if form.is_valid():
                rows_text = form.cleaned_data.get("rows_text") or ""
                workbook = form.cleaned_data.get("workbook")
                try:
                    if workbook:
                        rows_text = self._rows_text_from_workbook(workbook)
                except ValueError as exc:
                    form.add_error("workbook", str(exc))

                if not form.errors and not rows_text.strip():
                    form.add_error(None, "Paste rows or upload a completed workbook.")

                if not form.errors:
                    rows = self._parse_mass_rows(organisation, rows_text)
                    form = MassAddBusesForm(initial={"rows_text": rows_text})

                    if request.POST.get("action") == "commit":
                        created, updated, errors = self._commit_mass_rows(rows)
                        if created or updated:
                            self.message_user(
                                request,
                                f"Mass add complete: created {created}, updated {updated}, errors {errors}",
                            )
                        elif errors:
                            self.message_user(
                                request,
                                f"No rows imported. {errors} row(s) had errors.",
                                level=messages.WARNING,
                            )
                    else:
                        self.message_user(
                            request,
                            "Preview generated. Review rows and click Commit import when ready.",
                        )
        else:
            form = MassAddBusesForm()

        context = {
            **self.admin_site.each_context(request),
            "opts": self.model._meta,
            "original": organisation,
            "organisation": organisation,
            "title": f"Mass add buses for {organisation}",
            "form": form,
            "rows": rows,
            "can_commit": any(not row["errors"] for row in rows),
            "created": created,
            "updated": updated,
            "errors": errors,
            "template_download_url": reverse(
                "admin:busstops_organisation_mass_add_buses_template", args=(organisation.pk,)
            ),
            "export_download_url": reverse(
                "admin:busstops_organisation_mass_add_buses_current_fleet", args=(organisation.pk,)
            ),
            "allow_create": True,
            "commit_label": "Commit import",
        }

        return TemplateResponse(
            request,
            "admin/busstops/organisation/mass_buses.html",
            context,
        )


@admin.register(models.PreservationGroup)
class PreservationGroupAdmin(admin.ModelAdmin):
    list_display = ("name", "slug", "website", "founded_date", "vehicle_count")
    list_filter = ("founded_date",)
    search_fields = ("name", "slug", "description")
    prepopulated_fields = {"slug": ("name",)}
    readonly_fields = ["mass_add_buses_link"]
    fieldsets = (
        (None, {"fields": ("name", "slug", "description", "founded_date", "mass_add_buses_link")}),
        ("Branding", {"fields": ("logo", "banner")}),
        (
            "Links",
            {
                "fields": (
                    "website",
                    "social_x",
                    "social_fb",
                    "social_instagram",
                    "social_linkedin",
                    "social_youtube",
                    "social_tiktok",
                    "social_threads",
                    "social_bluesky",
                    "social_mastodon",
                    "social_other",
                )
            },
        ),
    )

    def get_queryset(self, request):
        return super().get_queryset(request).annotate(
            preserved_vehicle_count=Count("preserved_vehicles")
        )

    @admin.display(description="Vehicles", ordering="preserved_vehicle_count")
    def vehicle_count(self, obj):
        return obj.preserved_vehicle_count

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "<path:object_id>/mass-add-buses/",
                self.admin_site.admin_view(self.mass_add_buses_view),
                name="busstops_preservationgroup_mass_add_buses",
            ),
            path(
                "<path:object_id>/mass-add-buses/template.xlsx",
                self.admin_site.admin_view(self.mass_add_buses_template_view),
                name="busstops_preservationgroup_mass_add_buses_template",
            ),
            path(
                "<path:object_id>/mass-add-buses/current-fleet.xlsx",
                self.admin_site.admin_view(self.mass_add_buses_current_fleet_view),
                name="busstops_preservationgroup_mass_add_buses_current_fleet",
            ),
        ]
        return custom_urls + urls

    @admin.display(description="Mass add buses")
    def mass_add_buses_link(self, obj):
        url = reverse("admin:busstops_preservationgroup_mass_add_buses", args=(obj.pk,))
        return format_html('<a class="button" href="{}">Mass add buses</a>', url)

    def _build_mass_add_template_workbook(self, rows=None):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Vehicles"
        worksheet.append(
            (
                "code",
                "fleet_number",
                "fleet_code",
                "registration",
                "prev_registration",
                "vehicle_type",
                "livery",
                "colours",
                "name",
                "branding",
                "notes",
                "slug",
            )
        )
        for row in rows or ():
            worksheet.append(row)
        worksheet.freeze_panes = "A2"

        instructions = workbook.create_sheet("Instructions")
        instructions.append(["Field", "Notes"])
        instructions.append(["code", "Required unless fleet_number or registration is supplied"])
        instructions.append(["fleet_number", "Integer fleet number"])
        instructions.append(["fleet_code", "Optional displayed fleet code"])
        instructions.append(["registration", "Vehicle registration"])
        instructions.append(["prev_registration", "Previous registration"])
        instructions.append(["vehicle_type", "Vehicle type id, external id, or exact name"])
        instructions.append(["livery", "Livery id, external id, or exact name"])
        instructions.append(["colours", "Optional space-separated colour values"])
        instructions.append(["name", "Vehicle name"])
        instructions.append(["branding", "Branding text"])
        instructions.append(["notes", "Free-text notes"])
        instructions.append(["slug", "URL slug for the vehicle"])
        return workbook

    def _rows_text_from_workbook(self, uploaded_file):
        if not uploaded_file:
            return ""

        filename = (uploaded_file.name or "").lower()
        if filename.endswith(".csv"):
            try:
                content = uploaded_file.read().decode("utf-8-sig")
            except UnicodeDecodeError as exc:
                raise ValueError("CSV upload must be UTF-8 encoded") from exc
            return content.strip()
        if not filename.endswith(".xlsx"):
            raise ValueError("Upload must be a .xlsx or .csv file")

        workbook = load_workbook(uploaded_file, data_only=True)
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return ""

        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        if not any(headers):
            return ""

        output = StringIO()
        writer = csv.writer(output, delimiter="\t", lineterminator="\n")
        writer.writerow(headers)
        for row in rows[1:]:
            values = ["" if value is None else str(value).strip() for value in row[: len(headers)]]
            if any(values):
                writer.writerow(values)
        return output.getvalue()

    def _mass_rows_from_preservation_group_fleet(self, preservation_group):
        vehicles = (
            preservation_group.preserved_vehicles.select_related("vehicle_type", "livery")
            .order_by("fleet_number", "fleet_code", "reg", "code")
        )
        rows = []
        for vehicle in vehicles:
            rows.append(
                (
                    vehicle.code or "",
                    vehicle.fleet_number if vehicle.fleet_number is not None else "",
                    vehicle.fleet_code or "",
                    vehicle.reg or "",
                    vehicle.prev_registration or "",
                    vehicle.vehicle_type.name if vehicle.vehicle_type else "",
                    vehicle.livery.name if vehicle.livery else "",
                    vehicle.colours or "",
                    vehicle.name or "",
                    vehicle.branding or "",
                    vehicle.notes or "",
                    vehicle.slug or "",
                )
            )
        return rows

    def mass_add_buses_template_view(self, request, object_id):
        if not request.user.is_superuser:
            raise PermissionDenied

        preservation_group = self.get_object(request, object_id)
        if preservation_group is None:
            raise PermissionDenied

        workbook = self._build_mass_add_template_workbook()
        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)

        response = HttpResponse(
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{preservation_group.slug}-mass-add-template.xlsx"'
        )
        return response

    def mass_add_buses_current_fleet_view(self, request, object_id):
        if not request.user.is_superuser:
            raise PermissionDenied

        preservation_group = self.get_object(request, object_id)
        if preservation_group is None:
            raise PermissionDenied

        workbook = self._build_mass_add_template_workbook(
            rows=self._mass_rows_from_preservation_group_fleet(preservation_group)
        )
        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)

        response = HttpResponse(
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{preservation_group.slug}-current-fleet.xlsx"'
        )
        return response

    def _normalise_header(self, header):
        key = header.strip().lower().replace(" ", "_")
        header_aliases = {
            "fleet_num": "fleet_number",
            "fleet": "fleet_number",
            "registration": "reg",
            "prev_reg": "prev_registration",
            "previous_reg": "prev_registration",
            "vehicle_id": "vehicle_type",
            "vehicle_type_id": "vehicle_type",
            "livery_id": "livery",
            "color": "colours",
            "colors": "colours",
        }
        return header_aliases.get(key, key)

    @staticmethod
    def _coerce_bool(value):
        value = value.strip().lower()
        if value in {"", "none", "null"}:
            return None
        if value in {"1", "true", "yes", "y", "on"}:
            return True
        if value in {"0", "false", "no", "n", "off"}:
            return False
        raise ValueError(f"Invalid boolean value '{value}'")

    def _resolve_reference(self, model, value, label):
        if value == "":
            return None
        if value.isdigit():
            try:
                return model.objects.get(pk=int(value))
            except model.DoesNotExist as exc:
                raise ValueError(f"Unknown {label} id '{value}'") from exc

        try:
            return model.objects.get(external_id=value)
        except (model.DoesNotExist, model.MultipleObjectsReturned):
            pass

        if hasattr(model, "name"):
            item = model.objects.filter(name__iexact=value).first()
            if item:
                return item

        if hasattr(model, "code"):
            item = model.objects.filter(code__iexact=value).first()
            if item:
                return item

        raise ValueError(f"Unknown {label} '{value}'")

    def _parse_mass_rows(self, preservation_group, rows_text):
        rows = []

        if not rows_text.strip():
            return rows

        delimiter = "\t" if "\t" in rows_text.splitlines()[0] else ","
        reader = csv.DictReader(StringIO(rows_text), delimiter=delimiter)
        if not reader.fieldnames:
            return rows

        for index, original_row in enumerate(reader, start=2):
            mapped = {}
            for key, value in original_row.items():
                if not key:
                    continue
                mapped[self._normalise_header(key)] = (value or "").strip()

            if not any(mapped.values()):
                continue

            row = {
                "row_number": index,
                "raw": mapped,
                "errors": [],
                "action": "skip",
                "vehicle": None,
                "values": {},
            }

            provided_code = mapped.get("code", "")
            provided_fleet_code = mapped.get("fleet_code", "")
            code = provided_code or provided_fleet_code
            fleet_number = None
            if mapped.get("fleet_number"):
                try:
                    fleet_number = int(mapped["fleet_number"])
                except ValueError:
                    row["errors"].append("fleet_number must be an integer")
            elif mapped.get("fleet_number") == "":
                fleet_number = None

            reg = mapped.get("registration", "").upper().replace(" ", "")
            prev_registration = mapped.get("prev_registration", "").upper().replace(" ", "")

            if not code:
                if fleet_number is not None:
                    code = str(fleet_number)
                elif reg:
                    code = reg

            vehicle = None
            if code:
                vehicle = Vehicle.objects.filter(code__iexact=code).first()
            if not vehicle and reg:
                vehicle = Vehicle.objects.filter(reg__iexact=reg).first()

            if not vehicle and not code:
                row["errors"].append(
                    "Could not determine vehicle identifier (code/fleet_num/registration)"
                )

            row["vehicle"] = vehicle
            row["action"] = "update" if vehicle else "create"

            if row["action"] == "create":
                if code:
                    row["values"]["code"] = code
            elif provided_code:
                row["values"]["code"] = provided_code
            if provided_fleet_code:
                row["values"]["fleet_code"] = provided_fleet_code
            if reg:
                row["values"]["reg"] = reg
            if prev_registration:
                row["values"]["prev_registration"] = prev_registration
            if mapped.get("name"):
                row["values"]["name"] = mapped["name"]
            if mapped.get("notes"):
                row["values"]["notes"] = mapped["notes"]
            if mapped.get("branding"):
                row["values"]["branding"] = mapped["branding"]
            if mapped.get("colours"):
                row["values"]["colours"] = mapped["colours"]
            if mapped.get("fleet_number"):
                row["values"]["fleet_number"] = fleet_number

            for field, model, label in (
                ("vehicle_type", VehicleType, "vehicle_type"),
                ("livery", Livery, "livery"),
            ):
                if mapped.get(field):
                    try:
                        row["values"][field] = self._resolve_reference(
                            model, mapped.get(field, ""), label
                        )
                    except ValueError as exc:
                        row["errors"].append(str(exc))

            rows.append(row)

        return rows

    def _commit_mass_rows(self, preservation_group, rows):
        created = 0
        updated = 0
        errors = 0

        for row in rows:
            if row["errors"]:
                errors += 1
                continue

            try:
                with transaction.atomic():
                    vehicle = row["vehicle"] or Vehicle()
                    for field, value in row["values"].items():
                        setattr(vehicle, field, value)

                    vehicle.preservation_group = preservation_group
                    vehicle.preserved = True
                    vehicle.is_manual = True
                    vehicle.manual_updated_at = timezone.now()
                    vehicle.save()

                    if row["action"] == "create":
                        created += 1
                    else:
                        updated += 1
            except Exception as exc:
                row["errors"].append(str(exc))
                errors += 1

        return created, updated, errors

    def mass_add_buses_view(self, request, object_id):
        if not request.user.is_superuser:
            raise PermissionDenied

        preservation_group = self.get_object(request, object_id)
        if preservation_group is None:
            raise PermissionDenied

        rows = []
        created = 0
        updated = 0
        errors = 0

        if request.method == "POST":
            form = MassAddBusesForm(request.POST, request.FILES)
            if form.is_valid():
                rows_text = form.cleaned_data.get("rows_text") or ""
                workbook = form.cleaned_data.get("workbook")
                try:
                    if workbook:
                        rows_text = self._rows_text_from_workbook(workbook)
                except ValueError as exc:
                    form.add_error("workbook", str(exc))

                if not form.errors and not rows_text.strip():
                    form.add_error(None, "Paste rows or upload a completed workbook.")

                if not form.errors:
                    rows = self._parse_mass_rows(preservation_group, rows_text)
                    form = MassAddBusesForm(initial={"rows_text": rows_text})

                    if request.POST.get("action") == "commit":
                        created, updated, errors = self._commit_mass_rows(preservation_group, rows)
                        if created or updated:
                            self.message_user(
                                request,
                                f"Mass add complete: created {created}, updated {updated}, errors {errors}",
                            )
                        elif errors:
                            self.message_user(
                                request,
                                f"No rows imported. {errors} row(s) had errors.",
                                level=messages.WARNING,
                            )
                    else:
                        self.message_user(
                            request,
                            "Preview generated. Review rows and click Commit import when ready.",
                        )
        else:
            form = MassAddBusesForm()

        context = {
            **self.admin_site.each_context(request),
            "opts": self.model._meta,
            "original": preservation_group,
            "preservation_group": preservation_group,
            "title": f"Mass add buses for {preservation_group}",
            "form": form,
            "rows": rows,
            "can_commit": any(not row["errors"] for row in rows),
            "created": created,
            "updated": updated,
            "errors": errors,
            "template_download_url": reverse(
                "admin:busstops_preservationgroup_mass_add_buses_template", args=(preservation_group.pk,)
            ),
            "export_download_url": reverse(
                "admin:busstops_preservationgroup_mass_add_buses_current_fleet", args=(preservation_group.pk,)
            ),
            "allow_create": True,
            "commit_label": "Commit import",
        }

        return TemplateResponse(
            request,
            "admin/busstops/preservationgroup/mass_buses.html",
            context,
        )


class ManufacturerSiteInline(admin.TabularInline):
    model = models.ManufacturerSite
    extra = 0


@admin.register(models.Manufacturer)
class ManufacturerAdmin(admin.ModelAdmin):
    list_display = ("name", "slug", "short_name", "website")
    search_fields = ("name", "slug")
    prepopulated_fields = {"slug": ("name",)}
    inlines = [ManufacturerSiteInline]
    fieldsets = (
        (
            None,
            {
                "fields": (
                    "name",
                    "slug",
                    "short_name",
                    "legal_name",
                    "slogan",
                    "description",
                )
            },
        ),
        (
            "Branding",
            {
                "fields": (
                    "logo",
                    "banner",
                    "header_background",
                    "header_foreground",
                    "accent_colour",
                    "card_background",
                    "button_background",
                    "button_foreground",
                    "custom_css",
                )
            },
        ),
        (
            "Contact",
            {
                "fields": (
                    "website",
                    "email",
                    "phone",
                )
            },
        ),
        (
            "Social",
            {
                "fields": (
                    "social_x",
                    "social_fb",
                    "social_instagram",
                    "social_linkedin",
                    "social_youtube",
                    "social_tiktok",
                    "social_threads",
                    "social_bluesky",
                    "social_mastodon",
                    "social_other",
                )
            },
        ),
    )


@admin.register(models.ManufacturerSite)
class ManufacturerSiteAdmin(GISModelAdmin):
    list_display = ("name", "manufacturer", "site_type", "address")
    list_filter = ("site_type", "manufacturer")
    search_fields = ("name", "manufacturer__name", "address", "notes")
    autocomplete_fields = ("manufacturer",)


@admin.register(models.OperatorGroup)
class OperatorGroupAdmin(admin.ModelAdmin):
    list_display = ("name", "slug", "organisation", "group_fleet_numbering", "garages")
    list_filter = ("organisation", "group_fleet_numbering", "allow_transfers")
    search_fields = ("name", "slug")
    autocomplete_fields = ("organisation",)
    prepopulated_fields = {"slug": ("name",)}
    readonly_fields = ("garages_note", "mass_add_buses_link")
    fieldsets = (
        (None, {"fields": ("name", "slug", "organisation", "description", "mass_add_buses_link")}),
        (
            "Branding",
            {
                "fields": (
                    "logo",
                    "banner",
                    "header_background",
                    "header_foreground",
                    "accent_colour",
                    "custom_css",
                )
            },
        ),
        (
            "Links",
            {
                "fields": (
                    "website",
                    "social_x",
                    "social_fb",
                    "social_instagram",
                    "social_linkedin",
                    "social_youtube",
                    "social_tiktok",
                    "social_threads",
                    "social_bluesky",
                    "social_mastodon",
                    "social_other",
                )
            },
        ),
        (
            "Operations",
            {
                "fields": (
                    "group_fleet_numbering",
                    "allow_transfers",
                    "garages_note",
                )
            },
        ),
    )

    @admin.display(description="Garages")
    def garages(self, obj):
        return Garage.objects.filter(operators__group=obj).count()

    @admin.display(description="Garages")
    def garages_note(self, obj):
        url = reverse("admin:bustimes_garage_changelist")
        return format_html(
            'Garages hold depot/location data. Manage this group from <a href="{}?operators__group__id__exact={}">Garage admin</a>.',
            url,
            obj.pk,
        )

    def get_readonly_fields(self, request, obj=None):
        fields = list(super().get_readonly_fields(request, obj))
        if request.user.is_superuser and obj:
            fields.append("mass_add_buses_link")
        return fields

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "<path:object_id>/mass-add-buses/",
                self.admin_site.admin_view(self.mass_add_buses_view),
                name="busstops_operatorgroup_mass_add_buses",
            ),
            path(
                "<path:object_id>/mass-add-buses/template.xlsx",
                self.admin_site.admin_view(self.mass_add_buses_template_view),
                name="busstops_operatorgroup_mass_add_buses_template",
            ),
            path(
                "<path:object_id>/mass-add-buses/current-fleet.xlsx",
                self.admin_site.admin_view(self.mass_add_buses_current_fleet_view),
                name="busstops_operatorgroup_mass_add_buses_current_fleet",
            ),
        ]
        return custom_urls + urls

    @admin.display(description="Mass add buses")
    def mass_add_buses_link(self, obj):
        url = reverse("admin:busstops_operatorgroup_mass_add_buses", args=(obj.pk,))
        return format_html('<a class="button" href="{}">Mass add buses</a>', url)

    def _build_mass_add_template_workbook(self, rows=None):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Vehicles"
        worksheet.append(
            (
                "operator_code",
                "external_id",
                "code",
                "fleet_number",
                "fleet_code",
                "registration",
                "prev_registration",
                "vehicle_type",
                "livery",
                "colours",
                "garage",
                "name",
                "branding",
                "notes",
                "withdrawn",
                "preserved",
                "fleet_support_vehicle",
                "vor",
                "awaiting_delivery",
                "trainer_vehicle",
                "demonstrator",
                "features",
                "slug",
            )
        )
        for row in rows or ():
            worksheet.append(row)
        worksheet.freeze_panes = "A2"

        instructions = workbook.create_sheet("Instructions")
        instructions.append(["Field", "Notes"])
        instructions.append(
            [
                "operator_code",
                "Optional. Operator NOC/slug/operator code. Leave blank to use the first operator in this group.",
            ]
        )
        instructions.append(["external_id", "Optional external id for matching or creating vehicles"])
        instructions.append(["code", "Required unless fleet_number or registration is supplied"])
        instructions.append(["fleet_number", "Integer fleet number"])
        instructions.append(["fleet_code", "Optional displayed fleet code"])
        instructions.append(["registration", "Vehicle registration"])
        instructions.append(["prev_registration", "Previous registration"])
        instructions.append(["vehicle_type", "Vehicle type id, external id, or exact name"])
        instructions.append(["livery", "Livery id, external id, or exact name"])
        instructions.append(["colours", "Optional space-separated colour values"])
        instructions.append(["garage", "Garage id, external id, or exact code"])
        instructions.append(["name", "Vehicle name"])
        instructions.append(["branding", "Branding text"])
        instructions.append(["notes", "Free-text notes"])
        instructions.append(["withdrawn", "Boolean: true/false, yes/no, 1/0"])
        instructions.append(["preserved", "Boolean: true/false, yes/no, 1/0"])
        instructions.append(["fleet_support_vehicle", "Boolean: true/false, yes/no, 1/0"])
        instructions.append(["vor", "Boolean: true/false, yes/no, 1/0"])
        instructions.append(["awaiting_delivery", "Boolean: true/false, yes/no, 1/0"])
        instructions.append(["trainer_vehicle", "Boolean: true/false, yes/no, 1/0"])
        instructions.append(["demonstrator", "Boolean: true/false, yes/no, 1/0"])
        instructions.append(["features", "Comma-separated feature names or ids"])
        return workbook

    def _rows_text_from_workbook(self, uploaded_file):
        if not uploaded_file:
            return ""

        filename = (uploaded_file.name or "").lower()
        if filename.endswith(".csv"):
            try:
                content = uploaded_file.read().decode("utf-8-sig")
            except UnicodeDecodeError as exc:
                raise ValueError("CSV upload must be UTF-8 encoded") from exc
            return content.strip()
        if not filename.endswith(".xlsx"):
            raise ValueError("Upload must be a .xlsx or .csv file")

        workbook = load_workbook(uploaded_file, data_only=True)
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return ""

        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        if not any(headers):
            return ""

        output = StringIO()
        writer = csv.writer(output, delimiter="\t", lineterminator="\n")
        writer.writerow(headers)
        for row in rows[1:]:
            values = ["" if value is None else str(value).strip() for value in row[: len(headers)]]
            if any(values):
                writer.writerow(values)
        return output.getvalue()

    def _mass_rows_from_operator_group_fleet(self, operator_group):
        vehicles = (
            Vehicle.objects.filter(operator__group=operator_group)
            .select_related("vehicle_type", "livery", "garage", "operator")
            .prefetch_related("features")
            .filter(**current_vehicle_filters(withdrawn=False))
            .order_by("operator__name", "fleet_number", "fleet_code", "reg", "code")
        )
        rows = []
        for vehicle in vehicles:
            rows.append(
                (
                    vehicle.operator.noc,
                    vehicle.external_id or "",
                    vehicle.code or "",
                    vehicle.fleet_number if vehicle.fleet_number is not None else "",
                    vehicle.fleet_code or "",
                    vehicle.reg or "",
                    vehicle.prev_registration or "",
                    vehicle.vehicle_type.name if vehicle.vehicle_type else "",
                    vehicle.livery.name if vehicle.livery else "",
                    vehicle.colours or "",
vehicle.garage.name if vehicle.garage else "",
                    vehicle.name or "",
                    vehicle.branding or "",
                    vehicle.notes or "",
                    "true" if vehicle.withdrawn else "false",
                    "true" if vehicle.preserved else "false",
                    "true" if vehicle.fleet_support_vehicle else "false",
                    "true" if vehicle.vor else "false",
                    "true" if vehicle.awaiting_delivery else "false",
                    "true" if vehicle.trainer_vehicle else "false",
                    "true" if vehicle.demonstrator else "false",
                    ", ".join(feature.name for feature in vehicle.features.all()),
                    vehicle.slug or "",
                )
            )
        return rows

    def mass_add_buses_template_view(self, request, object_id):
        if not request.user.is_superuser:
            raise PermissionDenied

        operator_group = self.get_object(request, object_id)
        if operator_group is None:
            raise PermissionDenied

        workbook = self._build_mass_add_template_workbook()
        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)

        response = HttpResponse(
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{operator_group.slug}-mass-add-template.xlsx"'
        )
        return response

    def mass_add_buses_current_fleet_view(self, request, object_id):
        if not request.user.is_superuser:
            raise PermissionDenied

        operator_group = self.get_object(request, object_id)
        if operator_group is None:
            raise PermissionDenied

        workbook = self._build_mass_add_template_workbook(
            rows=self._mass_rows_from_operator_group_fleet(operator_group)
        )
        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)

        response = HttpResponse(
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{operator_group.slug}-current-fleet.xlsx"'
        )
        return response

    def _normalise_header(self, header):
        key = header.strip().lower().replace(" ", "_")
        header_aliases = {
            "fleet_num": "fleet_number",
            "fleet": "fleet_number",
            "registration": "reg",
            "prev_reg": "prev_registration",
            "previous_reg": "prev_registration",
            "vehicle_id": "vehicle_type",
            "vehicle_type_id": "vehicle_type",
            "livery_id": "livery",
            "color": "colours",
            "colors": "colours",
            "garage_id": "garage",
            "operator_noc": "operator_code",
            "operator": "operator_code",
        }
        return header_aliases.get(key, key)

    @staticmethod
    def _coerce_bool(value):
        value = value.strip().lower()
        if value in {"", "none", "null"}:
            return None
        if value in {"1", "true", "yes", "y", "on"}:
            return True
        if value in {"0", "false", "no", "n", "off"}:
            return False
        raise ValueError(f"Invalid boolean value '{value}'")

    def _resolve_reference(self, model, value, label):
        if value == "":
            return None
        if value.isdigit():
            try:
                return model.objects.get(pk=int(value))
            except model.DoesNotExist as exc:
                raise ValueError(f"Unknown {label} id '{value}'") from exc

        try:
            return model.objects.get(external_id=value)
        except (model.DoesNotExist, model.MultipleObjectsReturned):
            pass

        if hasattr(model, "name"):
            item = model.objects.filter(name__iexact=value).first()
            if item:
                return item

        if hasattr(model, "code"):
            item = model.objects.filter(code__iexact=value).first()
            if item:
                return item

        raise ValueError(f"Unknown {label} '{value}'")

    def _resolve_operator(self, operator_group, value):
        if not value:
            # Default to first operator in the group
            first_operator = models.Operator.objects.filter(group=operator_group).first()
            if first_operator:
                return first_operator
            raise ValueError("No operators in this group and no operator_code provided")

        text = str(value).strip()
        if not text:
            first_operator = models.Operator.objects.filter(group=operator_group).first()
            if first_operator:
                return first_operator
            raise ValueError("No operators in this group and no operator_code provided")

        operator = models.Operator.objects.filter(
            Q(noc__iexact=text)
            | Q(slug__iexact=text)
            | Q(operatorcode__code__iexact=text)
        ).first()
        if operator:
            return operator
        raise ValueError(f"Unknown operator_code '{text}'")

    def _parse_mass_rows(self, operator_group, rows_text):
        rows = []

        if not rows_text.strip():
            return rows

        delimiter = "\t" if "\t" in rows_text.splitlines()[0] else ","
        reader = csv.DictReader(StringIO(rows_text), delimiter=delimiter)
        if not reader.fieldnames:
            return rows

        for index, original_row in enumerate(reader, start=2):
            mapped = {}
            for key, value in original_row.items():
                if not key:
                    continue
                mapped[self._normalise_header(key)] = (value or "").strip()

            if not any(mapped.values()):
                continue

            row = {
                "row_number": index,
                "raw": mapped,
                "errors": [],
                "action": "skip",
                "operator": None,
                "vehicle": None,
                "values": {},
            }

            try:
                row["operator"] = self._resolve_operator(operator_group, mapped.get("operator_code"))
            except ValueError as exc:
                row["errors"].append(str(exc))

            if not row["operator"]:
                row["errors"].append("Could not determine operator")
                rows.append(row)
                continue

            provided_code = mapped.get("code", "")
            provided_fleet_code = mapped.get("fleet_code", "")
            code = provided_code or provided_fleet_code
            fleet_number = None
            if mapped.get("fleet_number"):
                try:
                    fleet_number = int(mapped["fleet_number"])
                except ValueError:
                    row["errors"].append("fleet_number must be an integer")
            elif mapped.get("fleet_number") == "":
                fleet_number = None

            reg = mapped.get("reg", "").upper().replace(" ", "")
            prev_registration = mapped.get("prev_registration", "").upper().replace(" ", "")

            if not code:
                if fleet_number is not None:
                    code = str(fleet_number)
                elif reg:
                    code = reg

            external_id = mapped.get("external_id") or None
            vehicle = None
            if external_id:
                vehicle = Vehicle.objects.filter(external_id=external_id).first()
            if not vehicle and code:
                vehicle = row["operator"].vehicle_set.filter(
                    **current_vehicle_filters(preserved=False)
                ).filter(code__iexact=code).first()
            if not vehicle and reg:
                vehicle = row["operator"].vehicle_set.filter(
                    **current_vehicle_filters(preserved=False)
                ).filter(reg__iexact=reg).first()

            if not vehicle and not code:
                row["errors"].append(
                    "Could not determine vehicle identifier (code/fleet_num/registration)"
                )

            row["vehicle"] = vehicle
            row["action"] = "update" if vehicle else "create"

            if row["action"] == "create":
                if code:
                    row["values"]["code"] = code
            elif provided_code:
                row["values"]["code"] = provided_code
            if provided_fleet_code:
                row["values"]["fleet_code"] = provided_fleet_code
            if reg:
                row["values"]["reg"] = reg
            if prev_registration:
                row["values"]["prev_registration"] = prev_registration
            if mapped.get("name"):
                row["values"]["name"] = mapped["name"]
            if mapped.get("notes"):
                row["values"]["notes"] = mapped["notes"]
            if mapped.get("branding"):
                row["values"]["branding"] = mapped["branding"]
            if mapped.get("colours"):
                row["values"]["colours"] = mapped["colours"]
            if mapped.get("fleet_number"):
                row["values"]["fleet_number"] = fleet_number
            if external_id:
                row["values"]["external_id"] = external_id

            for field, model, label in (
                ("vehicle_type", VehicleType, "vehicle_type"),
                ("livery", Livery, "livery"),
                ("garage", Garage, "garage"),
            ):
                if mapped.get(field):
                    try:
                        row["values"][field] = self._resolve_reference(
                            model, mapped.get(field, ""), label
                        )
                    except ValueError as exc:
                        row["errors"].append(str(exc))

            for bool_field in (
                "withdrawn",
                "preserved",
                "fleet_support_vehicle",
                "vor",
                "awaiting_delivery",
                "trainer_vehicle",
                "demonstrator",
            ):
                if bool_field in mapped and mapped[bool_field]:
                    try:
                        row["values"][bool_field] = self._coerce_bool(mapped[bool_field])
                    except ValueError as exc:
                        row["errors"].append(str(exc))

            if mapped.get("features"):
                feature_names = [f.strip() for f in mapped["features"].split(",")]
                features = []
                for feature_name in feature_names:
                    if not feature_name:
                        continue
                    try:
                        feature = self._resolve_reference(VehicleFeature, feature_name, "feature")
                        features.append(feature)
                    except ValueError:
                        row["errors"].append(f"Unknown feature '{feature_name}'")
                if features:
                    row["values"]["features"] = features

            rows.append(row)

        return rows

    def _commit_mass_rows(self, rows):
        created = 0
        updated = 0
        errors = 0

        for row in rows:
            if row["errors"]:
                errors += 1
                continue

            try:
                with transaction.atomic():
                    vehicle = row["vehicle"] or Vehicle()
                    for field, value in row["values"].items():
                        if field == "features":
                            vehicle.features.set(value)
                        else:
                            setattr(vehicle, field, value)

                    vehicle.is_manual = True
                    vehicle.manual_updated_at = timezone.now()
                    vehicle.save()

                    if row["action"] == "create":
                        created += 1
                    else:
                        updated += 1
            except Exception as exc:
                row["errors"].append(str(exc))
                errors += 1

        return created, updated, errors

    def mass_add_buses_view(self, request, object_id):
        if not request.user.is_superuser:
            raise PermissionDenied

        operator_group = self.get_object(request, object_id)
        if operator_group is None:
            raise PermissionDenied

        rows = []
        created = 0
        updated = 0
        errors = 0

        if request.method == "POST":
            form = MassAddBusesForm(request.POST, request.FILES)
            if form.is_valid():
                rows_text = form.cleaned_data.get("rows_text") or ""
                workbook = form.cleaned_data.get("workbook")
                try:
                    if workbook:
                        rows_text = self._rows_text_from_workbook(workbook)
                except ValueError as exc:
                    form.add_error("workbook", str(exc))

                if not form.errors and not rows_text.strip():
                    form.add_error(None, "Paste rows or upload a completed workbook.")

                if not form.errors:
                    rows = self._parse_mass_rows(operator_group, rows_text)
                    form = MassAddBusesForm(initial={"rows_text": rows_text})

                    if request.POST.get("action") == "commit":
                        created, updated, errors = self._commit_mass_rows(rows)
                        if created or updated:
                            self.message_user(
                                request,
                                f"Mass add complete: created {created}, updated {updated}, errors {errors}",
                            )
                        elif errors:
                            self.message_user(
                                request,
                                f"No rows imported. {errors} row(s) had errors.",
                                level=messages.WARNING,
                            )
                    else:
                        self.message_user(
                            request,
                            "Preview generated. Review rows and click Commit import when ready.",
                        )
        else:
            form = MassAddBusesForm()

        context = {
            **self.admin_site.each_context(request),
            "opts": self.model._meta,
            "original": operator_group,
            "operator_group": operator_group,
            "title": f"Mass add buses for {operator_group}",
            "form": form,
            "rows": rows,
            "can_commit": any(not row["errors"] for row in rows),
            "created": created,
            "updated": updated,
            "errors": errors,
            "template_download_url": reverse(
                "admin:busstops_operatorgroup_mass_add_buses_template", args=(operator_group.pk,)
            ),
            "export_download_url": reverse(
                "admin:busstops_operatorgroup_mass_add_buses_current_fleet", args=(operator_group.pk,)
            ),
            "allow_create": True,
            "commit_label": "Commit import",
        }

        return TemplateResponse(
            request,
            "admin/busstops/operatorgroup/mass_buses.html",
            context,
        )


@admin.register(models.Operator)
class OperatorAdmin(admin.ModelAdmin):
    form = OperatorAdminForm
    change_form_template = "admin/busstops/operator/change_form.html"
    vehicle_inline_limit = 50
    list_display = [
        "name",
        "preserved",
        "ceased_operations_on",
        "slug",
        "operator_codes",
        "noc",
        "external_id",
        "vehicle_mode",
        "region_id",
        "services",
        "vehicles",
        "twitter",
    ]
    list_filter = (
        "modified_at",
        DuplicateOperatorFilter,
        "region",
        "vehicle_mode",
        "payment_methods",
        "organisation",
        "government_authority",
        "group",
        "preserved",
    )
    search_fields = ("noc", "name", "external_id")
    raw_id_fields = ("region", "regions", "siblings", "colour")
    inlines = [OperatorCodeInline]
    readonly_fields = ["search_vector", "modified_at"]
    prepopulated_fields = {"slug": ("name",)}
    autocomplete_fields = ("organisation", "government_authority", "group")
    actions = ("assign_to_group", "assign_to_organisation", "assign_to_government_authority", "change_noc")
    mass_add_template_headers = (
        "operator_code",
        "external_id",
        "code",
        "fleet_number",
        "fleet_code",
        "registration",
        "prev_registration",
        "vehicle_type",
        "livery",
        "colours",
        "garage",
        "name",
        "branding",
        "notes",
        "withdrawn",
        "preserved",
        "fleet_support_vehicle",
        "vor",
        "awaiting_delivery",
        "trainer_vehicle",
        "demonstrator",
        "features",
        "slug",
    )
    mass_add_route_template_headers = (
        "operator_code",
        "service_id",
        "service_code",
        "line_name",
        "line_brand",
        "description",
        "mode",
        "current",
        "non_current_route",
        "timetable_wrong",
        "tracking",
        "public_use",
        "colour",
    )

    header_aliases = {
        "fleet_num": "fleet_number",
        "fleet": "fleet_number",
        "registration": "reg",
        "prev_reg": "prev_registration",
        "previous_reg": "prev_registration",
        "vehicle_id": "vehicle_type",
        "vehicle_type_id": "vehicle_type",
        "livery_id": "livery",
        "color": "colours",
        "colors": "colours",
        "garage_id": "garage",
        "operator_noc": "operator_code",
        "operator": "operator_code",
    }
    route_header_aliases = {
        "operator": "operator_code",
        "operator_noc": "operator_code",
        "id": "service_id",
        "route_id": "service_id",
        "service": "service_code",
        "service_name": "line_name",
        "brand": "line_brand",
        "route_brand": "line_brand",
        "route_description": "description",
        "route_colour": "colour",
        "color": "colour",
    }


    def _bulk_assign_related(self, request, queryset, form_class, field_name, template_title, submit_label):
        selected = request.POST.getlist(ACTION_CHECKBOX_NAME)
        if "apply" in request.POST:
            form = form_class(request.POST)
            if form.is_valid():
                target = form.cleaned_data[field_name]
                updated = queryset.update(
                    **{
                        field_name: target,
                        "is_manual": True,
                        "manual_updated_at": timezone.now(),
                    }
                )
                self.message_user(
                    request,
                    f"Updated {updated} operator{'s' if updated != 1 else ''}.",
                    level=messages.SUCCESS,
                )
                return None
        else:
            form = form_class()

        context = {
            **self.admin_site.each_context(request),
            "opts": self.model._meta,
            "queryset": queryset.order_by("name"),
            "form": form,
            "title": template_title,
            "field_name": field_name,
            "submit_label": submit_label,
            "selected_objects_label": "operator",
            "selected_objects_heading": "Selected operators",
            "action_checkbox_name": ACTION_CHECKBOX_NAME,
            "selected": selected,
            "action_name": request.POST.get("action", ""),
        }
        return TemplateResponse(
            request,
            "admin/busstops/operator/bulk_assign.html",
            context,
        )

    @admin.action(description="Add selected operators to a division")
    def assign_to_group(self, request, queryset):
        return self._bulk_assign_related(
            request,
            queryset,
            OperatorBulkAssignGroupForm,
            "group",
            "Assign selected operators to a division",
            "Apply division",
        )

    @admin.action(description="Add selected operators to a major operator")
    def assign_to_organisation(self, request, queryset):
        return self._bulk_assign_related(
            request,
            queryset,
            OperatorBulkAssignOrganisationForm,
            "organisation",
            "Assign selected operators to a major operator",
            "Apply major operator",
        )

    @admin.action(description="Add selected operators to a government authority")
    def assign_to_government_authority(self, request, queryset):
        return self._bulk_assign_related(
            request,
            queryset,
            OperatorBulkAssignGovernmentAuthorityForm,
            "government_authority",
            "Assign selected operators to a government authority",
            "Apply government authority",
        )

    @admin.action(description="Change NOC of selected operator")
    def change_noc(self, request, queryset):
        if queryset.count() != 1:
            self.message_user(
                request,
                "Please select exactly one operator to change its NOC.",
                level=messages.ERROR,
            )
            return None

        selected = request.POST.getlist(ACTION_CHECKBOX_NAME)
        operator = queryset.first()

        if "apply" in request.POST:
            form = ChangeNOCForm(request.POST)
            if form.is_valid():
                new_noc = form.cleaned_data["new_noc"].upper()
                old_noc = operator.noc

                if new_noc == old_noc:
                    self.message_user(
                        request,
                        "New NOC is the same as the current NOC.",
                        level=messages.WARNING,
                    )
                    return None

                if models.Operator.objects.filter(noc=new_noc).exists():
                    self.message_user(
                        request,
                        f"An operator with NOC '{new_noc}' already exists.",
                        level=messages.ERROR,
                    )
                    return None

                try:
                    with transaction.atomic():
                        from vehicles.models import Vehicle, VehicleRevision
                        from fleet.models import PinnedOperator

                        # Update all foreign key references
                        models.Depot.objects.filter(operator_id=old_noc).update(operator_id=new_noc)
                        models.OperatorVehicleColumn.objects.filter(operator_id=old_noc).update(operator_id=new_noc)
                        models.OperatorCode.objects.filter(operator_id=old_noc).update(operator_id=new_noc)
                        Vehicle.objects.filter(operator_id=old_noc).update(operator_id=new_noc)
                        Vehicle.objects.filter(operated_by_id=old_noc).update(operated_by_id=new_noc)
                        Vehicle.objects.filter(historical_fleet_id=old_noc).update(historical_fleet_id=new_noc)
                        VehicleRevision.objects.filter(from_operator_id=old_noc).update(from_operator_id=new_noc)
                        VehicleRevision.objects.filter(to_operator_id=old_noc).update(to_operator_id=new_noc)
                        VehicleRevision.objects.filter(from_operated_by_id=old_noc).update(from_operated_by_id=new_noc)
                        VehicleRevision.objects.filter(to_operated_by_id=old_noc).update(to_operated_by_id=new_noc)
                        PinnedOperator.objects.filter(operator_id=old_noc).update(operator_id=new_noc)

                        # Update the operator's primary key
                        operator.noc = new_noc
                        operator.save()

                    self.message_user(
                        request,
                        f"Successfully changed NOC from '{old_noc}' to '{new_noc}'.",
                        level=messages.SUCCESS,
                    )
                    return None
                except Exception as e:
                    self.message_user(
                        request,
                        f"Error changing NOC: {str(e)}",
                        level=messages.ERROR,
                    )
                    return None
        else:
            form = ChangeNOCForm()

        context = {
            **self.admin_site.each_context(request),
            "opts": self.model._meta,
            "queryset": queryset,
            "form": form,
            "title": f"Change NOC for {operator.name}",
            "operator": operator,
            "action_checkbox_name": ACTION_CHECKBOX_NAME,
            "selected": selected,
            "action_name": request.POST.get("action", ""),
        }
        return TemplateResponse(
            request,
            "admin/busstops/operator/change_noc.html",
            context,
        )

    def get_inlines(self, request, obj):
        inlines = [OperatorCodeInline, OperatorVehicleColumnInline]
        if request.user.is_superuser and obj and obj.vehicle_set.filter(
            **current_vehicle_filters()
        ).count() <= self.vehicle_inline_limit:
            inlines.append(OperatorVehicleInline)
        return inlines

    def get_readonly_fields(self, request, obj=None):
        fields = list(super().get_readonly_fields(request, obj))
        if request.user.is_superuser and obj:
            fields.append("mass_add_buses_link")
            fields.append("mass_add_routes_link")
            fields.append("mass_edit_buses_link")
            fields.append("vehicle_admin_link")
            fields.append("depot_admin_link")
            if obj.vehicle_set.filter(**current_vehicle_filters()).count() > self.vehicle_inline_limit:
                fields.append("vehicle_inline_notice")
        return fields

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "<path:object_id>/mass-add-buses/",
                self.admin_site.admin_view(self.mass_add_buses_view),
                name="busstops_operator_mass_add_buses",
            ),
            path(
                "<path:object_id>/mass-add-buses/template.xlsx",
                self.admin_site.admin_view(self.mass_add_buses_template_view),
                name="busstops_operator_mass_add_buses_template",
            ),
            path(
                "<path:object_id>/mass-add-buses/current-fleet.xlsx",
                self.admin_site.admin_view(self.mass_add_buses_current_fleet_view),
                name="busstops_operator_mass_add_buses_current_fleet",
            ),
            path(
                "<path:object_id>/mass-add-routes/",
                self.admin_site.admin_view(self.mass_add_routes_view),
                name="busstops_operator_mass_add_routes",
            ),
            path(
                "<path:object_id>/mass-add-routes/template.xlsx",
                self.admin_site.admin_view(self.mass_add_routes_template_view),
                name="busstops_operator_mass_add_routes_template",
            ),
            path(
                "<path:object_id>/mass-add-routes/current-routes.xlsx",
                self.admin_site.admin_view(self.mass_add_routes_current_view),
                name="busstops_operator_mass_add_routes_current",
            ),
            path(
                "<path:object_id>/mass-edit-buses/",
                self.admin_site.admin_view(self.mass_edit_buses_view),
                name="busstops_operator_mass_edit_buses",
            ),
            path(
                "<path:object_id>/mass-edit-buses/template.xlsx",
                self.admin_site.admin_view(self.mass_edit_buses_template_view),
                name="busstops_operator_mass_edit_buses_template",
            ),
            path(
                "<path:object_id>/export-basic.xlsx",
                self.admin_site.admin_view(self.export_basic_fleet_view),
                name="busstops_operator_export_basic",
            ),
            path(
                "<path:object_id>/export-advanced.xlsx",
                self.admin_site.admin_view(self.export_advanced_fleet_view),
                name="busstops_operator_export_advanced",
            ),
        ]
        return custom_urls + urls

    @admin.display(description="Mass add buses")
    def mass_add_buses_link(self, obj):
        url = reverse("admin:busstops_operator_mass_add_buses", args=(obj.pk,))
        return format_html('<a class="button" href="{}">Mass add buses</a>', url)

    @admin.display(description="Mass add routes")
    def mass_add_routes_link(self, obj):
        url = reverse("admin:busstops_operator_mass_add_routes", args=(obj.pk,))
        return format_html('<a class="button" href="{}">Mass add routes</a>', url)

    @admin.display(description="Mass edit buses")
    def mass_edit_buses_link(self, obj):
        url = reverse("admin:busstops_operator_mass_edit_buses", args=(obj.pk,))
        return format_html('<a class="button" href="{}">Mass edit buses</a>', url)

    def _build_mass_add_template_workbook(self, rows=None):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Vehicles"
        worksheet.append(self.mass_add_template_headers)
        for row in rows or ():
            worksheet.append(row)
        worksheet.freeze_panes = "A2"

        instructions = workbook.create_sheet("Instructions")
        instructions.append(["Field", "Notes"])
        instructions.append(
            [
                "operator_code",
                "Optional. Operator NOC/slug/operator code. Leave blank to use the current operator.",
            ]
        )
        instructions.append(["external_id", "Optional external id for matching or creating vehicles"])
        instructions.append(["code", "Required unless fleet_number or registration is supplied"])
        instructions.append(["fleet_number", "Integer fleet number"])
        instructions.append(["fleet_code", "Optional displayed fleet code"])
        instructions.append(["registration", "Vehicle registration"])
        instructions.append(["prev_registration", "Previous registration"])
        instructions.append(["vehicle_type", "Vehicle type id, external id, or exact name"])
        instructions.append(["livery", "Livery id, external id, or exact name"])
        instructions.append(["colours", "Optional space-separated colour values"])
        instructions.append(["garage", "Garage id, external id, or exact code"])
        instructions.append(["name", "Vehicle name"])
        instructions.append(["branding", "Branding text"])
        instructions.append(["notes", "Free-text notes"])
        instructions.append(["withdrawn", "Boolean: true/false, yes/no, 1/0"])
        instructions.append(["preserved", "Boolean: true/false, yes/no, 1/0"])
        instructions.append(["fleet_support_vehicle", "Boolean: true/false, yes/no, 1/0"])
        instructions.append(["vor", "Boolean: true/false, yes/no, 1/0"])
        instructions.append(["awaiting_delivery", "Boolean: true/false, yes/no, 1/0"])
        instructions.append(["trainer_vehicle", "Boolean: true/false, yes/no, 1/0"])
        instructions.append(["demonstrator", "Boolean: true/false, yes/no, 1/0"])
        instructions.append(["features", "Comma-separated feature names or ids"])
        instructions.append(["slug", "URL slug for the vehicle"])
        return workbook

    def _rows_text_from_workbook(self, uploaded_file):
        if not uploaded_file:
            return ""

        filename = (uploaded_file.name or "").lower()
        if filename.endswith(".csv"):
            try:
                content = uploaded_file.read().decode("utf-8-sig")
            except UnicodeDecodeError as exc:
                raise ValueError("CSV upload must be UTF-8 encoded") from exc
            return content.strip()
        if not filename.endswith(".xlsx"):
            raise ValueError("Upload must be a .xlsx or .csv file")

        workbook = load_workbook(uploaded_file, data_only=True)
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return ""

        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        if not any(headers):
            return ""

        output = StringIO()
        writer = csv.writer(output, delimiter="\t", lineterminator="\n")
        writer.writerow(headers)
        for row in rows[1:]:
            values = ["" if value is None else str(value).strip() for value in row[: len(headers)]]
            if any(values):
                writer.writerow(values)
        return output.getvalue()

    def _rows_text_from_upload(self, operator, uploaded_file):
        return shared_rows_text_from_upload(operator, uploaded_file)

    def mass_add_buses_template_view(self, request, object_id):
        if not request.user.is_superuser:
            raise PermissionDenied

        operator = self.get_object(request, object_id)
        if operator is None:
            raise PermissionDenied

        # Check if this is for historical fleet import
        is_historical = request.GET.get("historical") == "1"
        
        if is_historical:
            from vehicles.historical_fleet_bulk_import import build_template_workbook
            workbook = build_template_workbook()
        else:
            workbook = self._build_mass_add_template_workbook()
        
        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)

        response = HttpResponse(
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{operator.pk.lower()}-mass-add-template.xlsx"'
        )
        return response

    def _mass_rows_from_operator_fleet(self, operator):
        vehicles = (
            operator.vehicle_set.select_related("vehicle_type", "livery", "garage")
            .prefetch_related("features")
            .filter(**current_vehicle_filters(withdrawn=False))
            .order_by("fleet_number", "fleet_code", "reg", "code")
        )
        rows = []
        for vehicle in vehicles:
            rows.append(
                (
                    operator.noc,
                    vehicle.external_id or "",
                    vehicle.code or "",
                    vehicle.fleet_number if vehicle.fleet_number is not None else "",
                    vehicle.fleet_code or "",
                    vehicle.reg or "",
                    vehicle.prev_registration or "",
                    vehicle.vehicle_type.name if vehicle.vehicle_type else "",
                    vehicle.livery.name if vehicle.livery else "",
                    vehicle.colours or "",
vehicle.garage.name if vehicle.garage else "",
                    vehicle.name or "",
                    vehicle.branding or "",
                    vehicle.notes or "",
                    "true" if vehicle.withdrawn else "false",
                    "true" if vehicle.preserved else "false",
                    "true" if vehicle.fleet_support_vehicle else "false",
                    "true" if vehicle.vor else "false",
                    "true" if vehicle.awaiting_delivery else "false",
                    "true" if vehicle.trainer_vehicle else "false",
                    "true" if vehicle.demonstrator else "false",
                    ", ".join(feature.name for feature in vehicle.features.all()),
                    vehicle.slug or "",
                )
            )
        return rows
    def mass_add_buses_current_fleet_view(self, request, object_id):
        if not request.user.is_superuser:
            raise PermissionDenied

        operator = self.get_object(request, object_id)
        if operator is None:
            raise PermissionDenied

        workbook = self._build_mass_add_template_workbook(
            rows=self._mass_rows_from_operator_fleet(operator)
        )
        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)

        response = HttpResponse(
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{operator.pk.lower()}-current-fleet.xlsx"'
        )
        return response

    def _build_mass_add_routes_template_workbook(self, rows=None):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Routes"
        worksheet.append(self.mass_add_route_template_headers)
        for row in rows or ():
            worksheet.append(row)
        worksheet.freeze_panes = "A2"

        instructions = workbook.create_sheet("Instructions")
        instructions.append(["Field", "Notes"])
        instructions.append(["operator_code", "Optional. Defaults to the current operator."])
        instructions.append(
            [
                "service_id",
                "Optional existing service id. Best for updating a specific route without relying on matching text fields.",
            ]
        )
        instructions.append(
            [
                "service_code",
                "Used to match an existing service for this operator, or stored on new services.",
            ]
        )
        instructions.append(["line_name", "Route number or line name shown publicly."])
        instructions.append(["line_brand", "Optional brand or sub-brand."])
        instructions.append(["description", "Public route description such as origin and destination."])
        instructions.append(["mode", "Optional transport mode, defaults to bus."])
        instructions.append(
            [
                "current/non_current_route/timetable_wrong/tracking/public_use",
                "Boolean values: true/false, yes/no, 1/0.",
            ]
        )
        instructions.append(
            [
                "colour",
                "Optional service colour id or exact colour name.",
            ]
        )
        return workbook

    def _route_rows_text_from_workbook(self, uploaded_file):
        if not uploaded_file:
            return ""

        filename = (uploaded_file.name or "").lower()
        if filename.endswith(".csv"):
            try:
                content = uploaded_file.read().decode("utf-8-sig")
            except UnicodeDecodeError as exc:
                raise ValueError("CSV upload must be UTF-8 encoded") from exc
            return content.strip()
        if not filename.endswith(".xlsx"):
            raise ValueError("Upload must be a .xlsx or .csv file")

        workbook = load_workbook(uploaded_file, data_only=True)
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return ""

        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        if not any(headers):
            return ""

        output = StringIO()
        writer = csv.writer(output, delimiter="\t", lineterminator="\n")
        writer.writerow(headers)
        for row in rows[1:]:
            values = ["" if value is None else str(value).strip() for value in row[: len(headers)]]
            if any(values):
                writer.writerow(values)
        return output.getvalue()

    def _mass_route_rows_from_operator_services(self, operator):
        services = (
            models.Service.objects.filter(operator=operator)
            .select_related("colour")
            .order_by("line_name", "description", "service_code", "pk")
            .distinct()
        )
        rows = []
        for service in services:
            rows.append(
                (
                    operator.noc,
                    service.pk,
                    service.service_code or "",
                    service.line_name or "",
                    service.line_brand or "",
                    service.description or "",
                    service.mode or "",
                    "true" if service.current else "false",
                    "true" if service.non_current_route else "false",
                    "true" if service.timetable_wrong else "false",
                    "true" if service.tracking else "false",
                    "" if service.public_use is None else ("true" if service.public_use else "false"),
                    service.colour.name if service.colour else "",
                )
            )
        return rows

    def mass_add_routes_template_view(self, request, object_id):
        if not request.user.is_superuser:
            raise PermissionDenied

        operator = self.get_object(request, object_id)
        if operator is None:
            raise PermissionDenied

        workbook = self._build_mass_add_routes_template_workbook()
        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)

        response = HttpResponse(
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{operator.pk.lower()}-routes-template.xlsx"'
        )
        return response

    def mass_add_routes_current_view(self, request, object_id):
        if not request.user.is_superuser:
            raise PermissionDenied

        operator = self.get_object(request, object_id)
        if operator is None:
            raise PermissionDenied

        workbook = self._build_mass_add_routes_template_workbook(
            rows=self._mass_route_rows_from_operator_services(operator)
        )
        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)

        response = HttpResponse(
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{operator.pk.lower()}-current-routes.xlsx"'
        )
        return response

    def _normalise_route_header(self, header):
        key = header.strip().lower().replace(" ", "_")
        return self.route_header_aliases.get(key, key)

    def _resolve_service_colour(self, value):
        if value == "":
            return None
        if value.isdigit():
            try:
                return models.ServiceColour.objects.get(pk=int(value))
            except models.ServiceColour.DoesNotExist as exc:
                raise ValueError(f"Unknown colour id '{value}'") from exc
        colour = models.ServiceColour.objects.filter(name__iexact=value).first()
        if colour:
            return colour
        raise ValueError(f"Unknown colour '{value}'")

    def _find_mass_add_service(self, operator, mapped):
        service_id = mapped.get("service_id", "")
        if service_id:
            if not service_id.isdigit():
                raise ValueError(f"Invalid service_id '{service_id}'")
            try:
                return models.Service.objects.get(pk=int(service_id), operator=operator)
            except models.Service.DoesNotExist as exc:
                raise ValueError(
                    f"Unknown service_id '{service_id}' for operator {operator.noc}"
                ) from exc

        service_code = mapped.get("service_code", "")
        if service_code:
            matches = list(
                models.Service.objects.filter(
                    operator=operator, service_code__iexact=service_code
                )
                .distinct()
                .order_by("pk")[:2]
            )
            if len(matches) == 1:
                return matches[0]
            if len(matches) > 1:
                raise ValueError(
                    f"Multiple services match service_code '{service_code}'. Use service_id."
                )

        line_name = mapped.get("line_name", "")
        description = mapped.get("description", "")
        if line_name or description:
            filters = Q(operator=operator)
            if line_name:
                filters &= Q(line_name__iexact=line_name)
            if description:
                filters &= Q(description__iexact=description)
            matches = list(models.Service.objects.filter(filters).distinct().order_by("pk")[:2])
            if len(matches) == 1:
                return matches[0]
            if len(matches) > 1:
                raise ValueError(
                    "Multiple services match this line_name/description combination. Use service_id."
                )

        return None

    def _parse_mass_add_route_rows(self, operator, rows_text):
        parsed_rows = []
        if not rows_text.strip():
            return parsed_rows

        delimiter = "\t" if "\t" in rows_text.splitlines()[0] else ","
        reader = csv.DictReader(StringIO(rows_text), delimiter=delimiter)
        if not reader.fieldnames:
            return parsed_rows

        for row_number, original_row in enumerate(reader, start=2):
            mapped = {}
            for key, value in original_row.items():
                if not key:
                    continue
                mapped[self._normalise_route_header(key)] = (value or "").strip()
            if not any(mapped.values()):
                continue

            row = {
                "row_number": row_number,
                "raw": mapped,
                "errors": [],
                "service": None,
                "action": "create",
                "operator": operator,
                "operator_preview_label": f"Match {operator.noc} - {operator.name}",
                "values": {},
            }
            try:
                row["operator"] = self._resolve_operator(operator, mapped.get("operator_code", ""))
                row["operator_preview_label"] = (
                    f"Match {row['operator'].noc} - {row['operator'].name}"
                )
                row["service"] = self._find_mass_add_service(row["operator"], mapped)
                row["action"] = "update" if row["service"] else "create"

                values = {
                    "service_code": mapped.get("service_code", ""),
                    "line_name": mapped.get("line_name", ""),
                    "line_brand": mapped.get("line_brand", ""),
                    "description": mapped.get("description", ""),
                    "mode": mapped.get("mode", "") or "bus",
                    "current": self._coerce_bool(mapped.get("current", "true")),
                    "non_current_route": self._coerce_bool(
                        mapped.get("non_current_route", "false")
                    ),
                    "timetable_wrong": self._coerce_bool(
                        mapped.get("timetable_wrong", "false")
                    ),
                    "tracking": self._coerce_bool(mapped.get("tracking", "false")),
                    "public_use": self._coerce_bool(mapped.get("public_use", "")),
                    "colour": self._resolve_service_colour(mapped.get("colour", "")),
                }
                if not any(
                    [
                        values["service_code"],
                        values["line_name"],
                        values["line_brand"],
                        values["description"],
                    ]
                ):
                    raise ValueError(
                        "Provide at least one of service_code, line_name, line_brand, or description."
                    )
                row["values"] = values
            except ValueError as exc:
                row["errors"].append(str(exc))

            parsed_rows.append(row)

        return parsed_rows

    def _commit_mass_add_routes(self, rows):
        created = updated = errors = 0
        for row in rows:
            if row["errors"]:
                errors += 1
                continue
            try:
                with transaction.atomic():
                    service = row["service"]
                    values = row["values"].copy()
                    colour = values.pop("colour")
                    if service is None:
                        service = models.Service.objects.create(**values, colour=colour)
                        row["operator"].service_set.add(service)
                        service.update_search_vector()
                        row["service"] = service
                        created += 1
                    else:
                        for field, value in values.items():
                            setattr(service, field, value)
                        service.colour = colour
                        service.save(update_fields=[*values.keys(), "colour", "modified_at"])
                        service.update_search_vector()
                        if not service.operator.filter(pk=row["operator"].pk).exists():
                            service.operator.add(row["operator"])
                        updated += 1
            except Exception as exc:
                row["errors"].append(str(exc))
                errors += 1
        return created, updated, errors

    def mass_add_routes_view(self, request, object_id):
        if not request.user.is_superuser:
            raise PermissionDenied

        operator = self.get_object(request, object_id)
        if operator is None:
            raise PermissionDenied

        form = MassAddRoutesForm(request.POST or None, request.FILES or None)
        rows = []
        created = updated = errors = 0

        if request.method == "POST" and form.is_valid():
            rows_text = form.cleaned_data["rows_text"]
            workbook = form.cleaned_data.get("workbook")
            if workbook:
                try:
                    rows_text = self._route_rows_text_from_workbook(workbook)
                    form.cleaned_data["rows_text"] = rows_text
                except ValueError as exc:
                    form.add_error("workbook", str(exc))
            if rows_text and form.is_valid():
                rows = self._parse_mass_add_route_rows(operator, rows_text)
                if request.POST.get("action") == "commit":
                    created, updated, errors = self._commit_mass_add_routes(rows)
                    self.message_user(
                        request,
                        f"Mass add routes complete: created {created}, updated {updated}, errors {errors}",
                        level=messages.SUCCESS if errors == 0 else messages.WARNING,
                    )
                elif rows:
                    self.message_user(request, "Preview generated", level=messages.INFO)

        context = {
            **self.admin_site.each_context(request),
            "opts": self.model._meta,
            "operator": operator,
            "title": f"Mass add routes for {operator}",
            "form": form,
            "rows": rows,
            "created": created,
            "updated": updated,
            "errors": errors,
            "template_download_url": reverse(
                "admin:busstops_operator_mass_add_routes_template", args=(operator.pk,)
            ),
            "export_download_url": reverse(
                "admin:busstops_operator_mass_add_routes_current", args=(operator.pk,)
            ),
            "can_commit": bool(rows),
            "commit_label": "Commit routes",
        }
        return TemplateResponse(
            request,
            "admin/busstops/operator/mass_routes.html",
            context,
        )

    def mass_edit_buses_template_view(self, request, object_id):
        if not request.user.is_superuser:
            raise PermissionDenied

        operator = self.get_object(request, object_id)
        if operator is None:
            raise PermissionDenied

        workbook = self._build_mass_add_template_workbook()
        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)

        response = HttpResponse(
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{operator.pk.lower()}-mass-edit-template.xlsx"'
        )
        return response

    @admin.display(description="Vehicles")
    @admin.display(description="Vehicles")
    def vehicle_admin_link(self, obj):
        url = reverse("admin:vehicles_vehicle_changelist")
        return format_html(
            '<a href="{}?operator__noc__exact={}">Open vehicle list ({})</a>',
            url,
            obj.noc,
            obj.vehicle_set.filter(**current_vehicle_filters()).count(),
        )

    @admin.display(description="Garages")
    def depot_admin_link(self, obj):
        url = reverse("admin:bustimes_garage_changelist")
        add_url = reverse("admin:bustimes_garage_add")
        return format_html(
            '<a href="{}?operator__noc__exact={}">Manage garages</a> | '
            '<a href="{}?operator={}">Add garage on map</a>',
            url,
            obj.noc,
            add_url,
            obj.noc,
        )

    @admin.display(description="Vehicle inline")
    def vehicle_inline_notice(self, obj):
        return (
            f"Hidden for operators with more than {self.vehicle_inline_limit} vehicles "
            "to avoid oversized admin form submissions. Use the vehicle list link instead."
        )

    def _normalise_header(self, header):
        key = header.strip().lower().replace(" ", "_")
        return self.header_aliases.get(key, key)

    @staticmethod
    def _coerce_bool(value):
        value = value.strip().lower()
        if value in {"", "none", "null"}:
            return None
        if value in {"1", "true", "yes", "y", "on"}:
            return True
        if value in {"0", "false", "no", "n", "off"}:
            return False
        raise ValueError(f"Invalid boolean value '{value}'")

    def _resolve_reference(self, model, value, label):
        if value == "":
            return None
        if value.isdigit():
            try:
                return model.objects.get(pk=int(value))
            except model.DoesNotExist as exc:
                raise ValueError(f"Unknown {label} id '{value}'") from exc

        try:
            return model.objects.get(external_id=value)
        except (model.DoesNotExist, model.MultipleObjectsReturned):
            pass

        if hasattr(model, "name"):
            item = model.objects.filter(name__iexact=value).first()
            if item:
                return item

        if hasattr(model, "code"):
            item = model.objects.filter(code__iexact=value).first()
            if item:
                return item

        raise ValueError(f"Unknown {label} '{value}'")

    def _resolve_operator(self, default_operator, value):
        if not value:
            return default_operator
        text = str(value).strip()
        if not text:
            return default_operator

        operator = models.Operator.objects.filter(
            Q(noc__iexact=text)
            | Q(slug__iexact=text)
            | Q(operatorcode__code__iexact=text)
        ).first()
        if operator:
            return operator
        raise ValueError(f"Unknown operator_code '{text}'")

    def _resolve_operator_preview(self, default_operator, value, *, allow_create):
        if not value:
            return default_operator, "match", f"Match {default_operator.noc} - {default_operator.name}"
        text = str(value).strip()
        if not text:
            return default_operator, "match", f"Match {default_operator.noc} - {default_operator.name}"

        operator = models.Operator.objects.filter(
            Q(noc__iexact=text)
            | Q(slug__iexact=text)
            | Q(operatorcode__code__iexact=text)
        ).first()
        if operator:
            return operator, "match", f"Match {operator.noc} - {operator.name}"
        if allow_create:
            return None, "create", f"Create operator {text}"
        raise ValueError(f"Unknown operator_code '{text}'")

    def _resolve_garage_preview(self, operator, value, *, allow_create):
        if not value:
            return None, "", ""
        text = str(value).strip()
        if not text:
            return None, "", ""

        garages = Garage.objects.all()
        if operator:
            garages = garages.filter(operators=operator)
        trimmed = text[4:].strip() if text.upper().startswith("GSC ") else text
        garage = garages.filter(
            Q(name__iexact=text)
            | Q(code__iexact=text)
            | Q(name__iexact=trimmed)
            | Q(code__iexact=trimmed)
        ).first()
        if garage:
            label = garage.name or garage.code or str(garage.pk)
            first_operator = garage.operators.first()
            operator_label = first_operator.noc if first_operator else "unknown operator"
            return garage, "match", f"Match depot {label} for {operator_label}"
        if allow_create:
            operator_label = operator.noc if operator else "new operator"
            return None, "create", f"Create depot {text} for {operator_label}"
        raise ValueError(f"Unknown garage '{text}'")

    @staticmethod
    def _historical_code_with_year(code, historical_year):
        if not code or historical_year is None:
            return code
        return f"{code}-{historical_year}"
    def _vehicle_match_queryset(
        self, operator, historical_fleet=None, historical_year=None
    ):
        if historical_year is not None:
            filters = {"historical_fleet": historical_fleet or operator}
            columns = vehicle_db_columns()
            if "historical_fleet_year" in columns:
                filters["historical_fleet_year"] = historical_year
            return operator.vehicle_set.filter(**filters)

        if operator.preserved or historical_fleet:
            filters = {}
            columns = vehicle_db_columns()
            if historical_fleet and "historical_fleet_id" in columns:
                filters["historical_fleet"] = historical_fleet
            return operator.vehicle_set.filter(**filters)
        return operator.vehicle_set.filter(**current_vehicle_filters(preserved=False))

    def _parse_mass_rows(
        self,
        operator,
        rows_text,
        *,
        allow_create=True,
        historical_fleet=None,
        historical_year=None,
    ):
        rows = []

        if not rows_text.strip():
            return rows

        delimiter = "	" if "	" in rows_text.splitlines()[0] else ","
        reader = csv.DictReader(StringIO(rows_text), delimiter=delimiter)
        if not reader.fieldnames:
            return rows

        for index, original_row in enumerate(reader, start=2):
            mapped = {}
            for key, value in original_row.items():
                if not key:
                    continue
                mapped[self._normalise_header(key)] = (value or "").strip()

            if not any(mapped.values()):
                continue

            row = {
                "row_number": index,
                "raw": mapped,
                "errors": [],
                "action": "skip",
                "operator": operator,
                "operator_preview_action": "match",
                "operator_preview_label": f"Match {operator.noc} - {operator.name}",
                "pending_operator_code": "",
                "vehicle": None,
                "values": {},
                "features": None,
                "has_features": False,
                "garage_preview_action": "",
                "garage_preview_label": "",
                "pending_garage_name": "",
                "historical_fleet": historical_fleet,
                "historical_fleet_preview_label": (
                    f"Attach to {historical_fleet.noc} - {historical_fleet.name}"
                    if historical_fleet
                    else ""
                ),
                "historical_year": historical_year,
                "final_code_preview": "",
            }

            if "operator_code" in mapped:
                try:
                    (
                        row["operator"],
                        row["operator_preview_action"],
                        row["operator_preview_label"],
                    ) = self._resolve_operator_preview(
                        operator, mapped.get("operator_code"), allow_create=allow_create
                    )
                    if row["operator_preview_action"] == "create":
                        row["pending_operator_code"] = mapped.get("operator_code", "").strip()
                except ValueError as exc:
                    row["errors"].append(str(exc))
            if historical_year is not None and row["operator"] and row["operator"].pk != operator.pk:
                row["errors"].append(
                    "Historical fleet imports must stay on the operator you opened."
                )
                row["operator"] = operator
                row["operator_preview_action"] = "match"
                row["operator_preview_label"] = f"Match {operator.noc} - {operator.name}"
            provided_code = mapped.get("code", "")
            provided_fleet_code = mapped.get("fleet_code", "")
            code = provided_code or provided_fleet_code
            fleet_number = None
            if mapped.get("fleet_number"):
                try:
                    fleet_number = int(mapped["fleet_number"])
                except ValueError:
                    row["errors"].append("fleet_number must be an integer")
            elif mapped.get("fleet_number") == "":
                fleet_number = None

            reg = mapped.get("reg", "").upper().replace(" ", "")
            prev_registration = mapped.get("prev_registration", "").upper().replace(
                " ", ""
            )

            if not code:
                if fleet_number is not None:
                    code = str(fleet_number)
                elif reg:
                    code = reg

            external_id = mapped.get("external_id") or None
            vehicle = None
            if external_id:
                vehicle = Vehicle.objects.filter(external_id=external_id).first()
                if (
                    vehicle
                    and historical_year is not None
                    and (
                        vehicle.operator_id != operator.pk
                        or vehicle.historical_fleet_id != operator.pk
                        or vehicle.historical_fleet_year != historical_year
                    )
                ):
                    max_len = Vehicle._meta.get_field("external_id").max_length
                    suffix = f"-{historical_year}"
                    base = str(external_id)
                    if len(base) + len(suffix) > max_len:
                        base = base[: max_len - len(suffix)]
                    external_id = f"{base}{suffix}"
                    vehicle = Vehicle.objects.filter(external_id=external_id).first()
                if (
                    vehicle
                    and row["operator"]
                    and row["operator"].preserved
                    and vehicle.operator_id != row["operator"].pk
                ):
                    # Preserved fleets must not update other fleets. If the source provides an external_id
                    # that collides with a live fleet, namespace it using the preserved operator code.
                    suffix = (row["operator"].noc or "").strip().upper()
                    if not suffix:
                        row["errors"].append(
                            "Preserved fleets are isolated but operator noc is missing; cannot namespace external_id"
                        )
                        vehicle = None
                    else:
                        base = str(external_id)
                        # external_id field max_length is 100 on Vehicle
                        max_len = Vehicle._meta.get_field("external_id").max_length
                        extra = f"-{suffix}"
                        if len(base) + len(extra) > max_len:
                            base = base[: max_len - len(extra)]
                        external_id = f"{base}{extra}"
                        # If we've already imported this preserved variant before, treat it as an update.
                        vehicle = Vehicle.objects.filter(external_id=external_id).first()
            if not vehicle and code:
                if row["operator"]:
                    queryset = self._vehicle_match_queryset(
                        row["operator"],
                        row.get("historical_fleet"),
                        row.get("historical_year"),
                    )
                    if row.get("historical_year") is not None:
                        vehicle = queryset.filter(
                            Q(code__iexact=code)
                            | Q(code__iexact=self._historical_code_with_year(code, row["historical_year"]))
                            | Q(fleet_code__iexact=provided_fleet_code or code)
                        ).first()
                    else:
                        vehicle = queryset.filter(code__iexact=code).first()
            if not vehicle and reg:
                if row["operator"]:
                    vehicle = self._vehicle_match_queryset(
                        row["operator"],
                        row.get("historical_fleet"),
                        row.get("historical_year"),
                    ).filter(reg__iexact=reg).first()

            if not vehicle and not code:
                row["errors"].append(
                    "Could not determine vehicle identifier (code/fleet_num/registration)"
                )
            elif not vehicle and not allow_create:
                row["errors"].append(
                    "No existing vehicle matched this row for mass edit"
                )

            row["vehicle"] = vehicle
            row["action"] = "update" if vehicle else "create"

            if row["action"] == "create":
                if code:
                    row["values"]["code"] = code
            elif provided_code:
                row["values"]["code"] = provided_code
            if mapped.get("external_id"):
                row["values"]["external_id"] = external_id
            if mapped.get("fleet_number"):
                row["values"]["fleet_number"] = fleet_number
            if provided_fleet_code:
                row["values"]["fleet_code"] = provided_fleet_code
            if reg:
                row["values"]["reg"] = reg
            if prev_registration:
                row["values"]["prev_registration"] = prev_registration
            if mapped.get("name"):
                row["values"]["name"] = mapped["name"]
            if mapped.get("notes"):
                row["values"]["notes"] = mapped["notes"]
            if mapped.get("branding"):
                row["values"]["branding"] = mapped["branding"]
            if mapped.get("colours"):
                row["values"]["colours"] = mapped["colours"]

            if row.get("historical_fleet"):
                row["values"]["historical_fleet"] = row["historical_fleet"]
            if row.get("historical_year") is not None:
                row["values"]["historical_fleet"] = operator
                row["values"]["historical_fleet_year"] = row["historical_year"]
                if code and not provided_fleet_code:
                    row["values"]["fleet_code"] = code
                row["final_code_preview"] = self._historical_code_with_year(
                    code,
                    row["historical_year"],
                )
            elif row.get("operator") and row["operator"].preserved and "preserved" not in row["values"]:
                row["values"]["preserved"] = True
            else:
                row["final_code_preview"] = code

            for field in (
                "withdrawn",
                "preserved",
                "fleet_support_vehicle",
                "vor",
                "awaiting_delivery",
                "trainer_vehicle",
                "demonstrator",
            ):
                if not mapped.get(field):
                    continue
                try:
                    parsed_bool = self._coerce_bool(mapped[field])
                except ValueError as exc:
                    row["errors"].append(str(exc))
                else:
                    if parsed_bool is not None:
                        row["values"][field] = parsed_bool

            # Year-based historical imports always stay attached to the
            # operator's historical fleet, but they are not implicitly preserved.
            if row.get("historical_year") is not None:
                row["values"]["historical_fleet"] = operator
                row["values"]["historical_fleet_year"] = row["historical_year"]

            for field, model, label in (
                ("vehicle_type", VehicleType, "vehicle_type"),
                ("livery", Livery, "livery"),
            ):
                if mapped.get(field):
                    try:
                        row["values"][field] = self._resolve_reference(
                            model, mapped.get(field, ""), label
                        )
                    except ValueError as exc:
                        if field == "vehicle_type" and allow_create:
                            row["values"]["vehicle_type"] = mapped.get(field, "")
                        elif field == "livery" and allow_create:
                            row["values"]["notes"] = "\n".join(
                                part
                                for part in (
                                    row["values"].get("notes", ""),
                                    f"Imported livery text: {mapped.get(field, '')}",
                                )
                                if part
                            )
                        else:
                            row["errors"].append(str(exc))

            if mapped.get("garage"):
                try:
                    (
                        garage,
                        row["garage_preview_action"],
                        row["garage_preview_label"],
                    ) = self._resolve_garage_preview(
                        row["operator"], mapped.get("garage", ""), allow_create=allow_create
                    )
                    if garage:
                        row["values"]["garage"] = garage
                    elif row["garage_preview_action"] == "create":
                        row["pending_garage_name"] = mapped.get("garage", "").strip()
                except ValueError as exc:
                    row["errors"].append(str(exc))

            if mapped.get("features"):
                row["has_features"] = True
                feature_values = [
                    item.strip()
                    for item in mapped.get("features", "").replace(";", ",").split(","
                    )
                    if item.strip()
                ]
                row["features"] = feature_values

            rows.append(row)

        return rows

    def _get_or_create_vehicle_type(self, value):
        text = str(value or "").strip()
        if not text:
            return None
        vehicle_type = VehicleType.objects.filter(name__iexact=text).first()
        if vehicle_type:
            return vehicle_type
        return VehicleType.objects.create(
            name=text,
            is_manual=True,
            manual_updated_at=timezone.now(),
        )

    def _commit_mass_rows(
        self,
        operator,
        rows,
        *,
        allow_create=True,
        historical_fleet=None,
        historical_year=None,
        historical_creator="",
    ):
        created = 0
        updated = 0
        errors = 0

        for row in rows:
            if row["errors"]:
                errors += 1
                continue

            if row["action"] == "create" and not allow_create:
                row["errors"].append("Mass edit only supports existing vehicles")
                errors += 1
                continue

            try:
                with transaction.atomic():
                    row_operator = row.get("operator") or operator
                    row_historical_fleet = row.get("historical_fleet") or historical_fleet
                    row_historical_year = row.get("historical_year")
                    if row_historical_year is None:
                        row_historical_year = historical_year
                    if not row.get("operator") and row.get("pending_operator_code"):
                        row_operator = self._create_operator_from_code(
                            row["pending_operator_code"]
                        )
                        row["operator"] = row_operator
                        row["operator_preview_action"] = "match"
                        row["operator_preview_label"] = (
                            f"Created {row_operator.noc} - {row_operator.name}"
                        )
                    if row.get("pending_garage_name") and not row["values"].get("garage"):
                        created_garage = shared_create_garage_for_operator(
                            row_operator,
                            row["pending_garage_name"],
                        )
                        row["values"]["garage"] = created_garage
                        row["garage_preview_action"] = "match"
                        row["garage_preview_label"] = f"Created depot {created_garage}"
                    vehicle = row["vehicle"] or Vehicle(operator=row_operator)
                    if (
                        row_operator.preserved
                        and vehicle.pk
                        and vehicle.operator_id != row_operator.pk
                    ):
                        raise ValueError(
                            "Preserved fleets are isolated: refusing to update a vehicle belonging to another operator"
                        )
                    vehicle.operator = row_operator
                    if row_historical_fleet:
                        vehicle.historical_fleet = row_historical_fleet
                    if row_historical_year is not None:
                        vehicle.historical_fleet = operator
                        vehicle.historical_fleet_year = row_historical_year
                        vehicle.historical_fleet_creator = historical_creator
                        incoming_code = row["values"].get("code") or vehicle.code
                        if incoming_code:
                            desired_code = self._historical_code_with_year(
                                incoming_code,
                                row_historical_year,
                            )
                            if row["action"] == "create":
                                row["values"]["code"] = desired_code
                            elif vehicle.code != desired_code:
                                row["values"]["code"] = desired_code
                        vehicle.slug = slugify(vehicle_slug(vehicle))
                    elif row["values"].get("withdrawn") is True and not row_historical_fleet:
                        vehicle.historical_fleet = row_operator

                    for field, value in row["values"].items():
                        if field == "vehicle_type" and isinstance(value, str):
                            value = self._get_or_create_vehicle_type(value)
                        setattr(vehicle, field, value)

                    vehicle.is_manual = True
                    vehicle.manual_updated_at = timezone.now()
                    vehicle.save()

                    if row["has_features"]:
                        feature_objs = []
                        for feature in row["features"] or []:
                            if feature.isdigit():
                                obj = VehicleFeature.objects.filter(pk=int(feature)).first()
                                if not obj:
                                    raise ValueError(
                                        f"Unknown feature id '{feature}'"
                                    )
                            else:
                                obj, _ = VehicleFeature.objects.get_or_create(name=feature)
                            feature_objs.append(obj)
                        vehicle.features.set(feature_objs)

                    if row["action"] == "create":
                        created += 1
                    else:
                        updated += 1
            except Exception as exc:
                row["errors"].append(str(exc))
                errors += 1

        return created, updated, errors

    def mass_add_buses_view(self, request, object_id):
        if not request.user.is_superuser:
            raise PermissionDenied

        operator = self.get_object(request, object_id)
        if operator is None:
            raise PermissionDenied

        rows = []
        created = 0
        updated = 0
        errors = 0

        if request.method == "POST":
            form = MassAddBusesForm(request.POST, request.FILES)
            if form.is_valid():
                rows_text = form.cleaned_data.get("rows_text") or ""
                workbook = form.cleaned_data.get("workbook")
                try:
                    if workbook:
                        rows_text = self._rows_text_from_upload(operator, workbook)
                except ValueError as exc:
                    form.add_error("workbook", str(exc))

                if not form.errors and not rows_text.strip():
                    form.add_error(None, "Paste rows or upload a completed workbook.")

                if not form.errors:
                    rows = self._parse_mass_rows(operator, rows_text)
                    form = MassAddBusesForm(initial={"rows_text": rows_text})

                    if request.POST.get("action") == "commit":
                        created, updated, errors = self._commit_mass_rows(operator, rows)
                        if created or updated:
                            self.message_user(
                                request,
                                f"Mass add complete: created {created}, updated {updated}, errors {errors}",
                            )
                        elif errors:
                            self.message_user(
                                request,
                                f"No rows imported. {errors} row(s) had errors.",
                                level=messages.WARNING,
                            )
                    else:
                        self.message_user(
                            request,
                            "Preview generated. Review rows and click Commit import when ready.",
                        )
        else:
            form = MassAddBusesForm()

        context = {
            **self.admin_site.each_context(request),
            "opts": self.model._meta,
            "original": operator,
            "operator": operator,
            "title": f"Mass add buses for {operator}",
            "form": form,
            "rows": rows,
            "can_commit": any(not row["errors"] for row in rows),
            "created": created,
            "updated": updated,
            "errors": errors,
            "template_download_url": reverse(
                "admin:busstops_operator_mass_add_buses_template", args=(operator.pk,)
            ),
            "export_download_url": reverse(
                "admin:busstops_operator_mass_add_buses_current_fleet", args=(operator.pk,)
            ),
            "allow_create": True,
            "commit_label": "Commit import",
        }

        return TemplateResponse(
            request,
            "admin/busstops/operator/mass_buses.html",
            context,
        )

    # HistoricalFleet model was removed - this view is no longer functional
    # def new_historical_fleet_view(self, request, object_id):
    #     if not request.user.is_superuser:
    #         raise PermissionDenied
    #
    #     operator = self.get_object(request, object_id)
    #     if operator is None:
    #         raise PermissionDenied
    #
    #     rows = []
    #     created = 0
    #     updated = 0
    #     errors = 0
    #
    #     if request.method == "POST":
    #         form = NewHistoricalFleetForm(request.POST, request.FILES)
    #         if form.is_valid():
    #             rows_text = form.cleaned_data.get("rows_text") or ""
    #             workbook = form.cleaned_data.get("workbook")
    #             try:
    #                 if workbook:
    #                     rows_text = self._rows_text_from_workbook(workbook, service=service)
    #             except ValueError as exc:
    #                 form.add_error("workbook", str(exc))
    #
    #             if not form.errors and not rows_text.strip():
    #                 form.add_error(None, "Paste rows or upload a completed workbook.")
    #
    #             if not form.errors:
    #                 from vehicles.historical_fleet_bulk_import import (
    #                     bulk_import_historical_vehicles,
    #                     parse_pasted_rows,
    #                 )
    #
    #                 rows, parse_err = parse_pasted_rows(rows_text)
    #                 if parse_err:
    #                     form.add_error(None, parse_err)
    #                 else:
    #                     # Parse rows for preview
    #                     from vehicles.historical_fleet_bulk_import import build_historical_vehicles
    #
    #                     instances, row_errors = build_historical_vehicles(operator.id, rows)
    #
    #                     # Build preview rows
    #                     rows = []
    #                     for i, (instance, error) in enumerate(zip(instances, row_errors), 1):
    #                         rows.append({
    #                             "row_number": i,
    #                             "fleet_number": instance.fleet_number,
    #                             "fleet_code": instance.fleet_code,
    #                             "reg": instance.reg,
    #                             "vehicle_type": instance.vehicle_type.name if instance.vehicle_type else "",
    #                             "livery": instance.livery.name if instance.livery else "",
    #                             "joined_fleet_date": instance.joined_fleet_date.strftime("%d-%m-%Y") if instance.joined_fleet_date else "",
    #                             "left_fleet_date": instance.left_fleet_date.strftime("%d-%m-%Y") if instance.left_fleet_date else "",
    #                             "errors": error if error else "",
    #                         })
    #
    #                     form = NewHistoricalFleetForm(
    #                         initial={
    #                             "rows_text": rows_text,
    #                         }
    #                     )
    #
    #                     if request.POST.get("action") == "commit":
    #                         created, import_errors = bulk_import_historical_vehicles(operator.id, rows_text)
    #                         errors = len(import_errors)
    #                         if created:
    #                             self.message_user(
    #                                 request,
    #                                 f"Historical fleet import complete: created {created}, errors {errors}",
    #                             )
    #                         elif errors:
    #                             self.message_user(
    #                                 request,
    #                                 f"No rows imported. {errors} row(s) had errors.",
    #                                 level=messages.WARNING,
    #                             )
    #                     else:
    #                         self.message_user(
    #                             request,
    #                             "Preview generated. Review rows and click Commit import when ready.",
    #                         )
    #     else:
    #         form = NewHistoricalFleetForm()
    #
    #     context = {
    #         **self.admin_site.each_context(request),
    #         "opts": self.model._meta,
    #         "original": operator,
    #         "operator": operator,
    #         "title": f"New historical fleet for {operator}",
    #         "form": form,
    #         "rows": rows,
    #         "can_commit": any(not row["errors"] for row in rows),
    #         "created": created,
    #         "updated": updated,
    #         "errors": errors,
    #         "template_download_url": reverse(
    #             "admin:busstops_operator_mass_add_buses_template", args=(operator.pk,)
    #         ) + "?historical=1",
    #         "export_download_url": "",
    #         "allow_create": True,
    #         "commit_label": "Commit historical fleet",
    #         "historical_import": True,
    #     }
    #
    #     return TemplateResponse(
    #         request,
    #         "admin/busstops/operator/mass_buses.html",
    #         context,
    #     )

    def mass_edit_buses_view(self, request, object_id):
        if not request.user.is_superuser:
            raise PermissionDenied

        operator = self.get_object(request, object_id)
        if operator is None:
            raise PermissionDenied

        rows = []
        created = 0
        updated = 0
        errors = 0

        if request.method == "POST":
            form = MassEditBusesForm(request.POST, request.FILES)
            if form.is_valid():
                rows_text = form.cleaned_data.get("rows_text") or ""
                workbook = form.cleaned_data.get("workbook")
                try:
                    if workbook:
                        rows_text = self._rows_text_from_workbook(workbook)
                except ValueError as exc:
                    form.add_error("workbook", str(exc))

                if not form.errors and not rows_text.strip():
                    form.add_error(None, "Paste rows or upload a completed workbook.")

                if not form.errors:
                    rows = self._parse_mass_rows(operator, rows_text, allow_create=False)
                    form = MassEditBusesForm(initial={"rows_text": rows_text})

                    if request.POST.get("action") == "commit":
                        created, updated, errors = self._commit_mass_rows(
                            operator,
                            rows,
                            allow_create=False,
                        )
                        if updated:
                            self.message_user(
                                request,
                                f"Mass edit complete: updated {updated}, errors {errors}",
                            )
                        elif errors:
                            self.message_user(
                                request,
                                f"No rows updated. {errors} row(s) had errors.",
                                level=messages.WARNING,
                            )
                    else:
                        self.message_user(
                            request,
                            "Preview generated. Review rows and click Commit changes when ready.",
                        )
        else:
            form = MassEditBusesForm()

        context = {
            **self.admin_site.each_context(request),
            "opts": self.model._meta,
            "original": operator,
            "operator": operator,
            "title": f"Mass edit buses for {operator}",
            "form": form,
            "rows": rows,
            "can_commit": any(not row["errors"] for row in rows),
            "created": created,
            "updated": updated,
            "errors": errors,
            "allow_create": False,
            "commit_label": "Commit changes",
            "template_download_url": reverse(
                "admin:busstops_operator_mass_edit_buses_template", args=(operator.pk,)
            ),
            "export_download_url": reverse(
                "admin:busstops_operator_mass_add_buses_current_fleet", args=(operator.pk,)
            ),
        }

        return TemplateResponse(
            request,
            "admin/busstops/operator/mass_buses.html",
            context,
        )

    def get_queryset(self, request):
        queryset = super().get_queryset(request)
        if "changelist" in request.resolver_match.view_name:
            queryset = queryset.annotate(
                services=SubqueryCount("service", filter=Q(service__current=True)),
                vehicles=SubqueryCount("vehicle"),
            ).prefetch_related("operatorcode_set")
        return queryset

    @admin.display(ordering="services")
    def services(self, obj):
        url = reverse("admin:busstops_service_changelist")
        return format_html(
            '<a href="{}?operator__noc__exact={}">{}</a>', url, obj.noc, obj.services
        )

    @admin.display(ordering="vehicles")
    def vehicles(self, obj):
        url = reverse("admin:vehicles_vehicle_changelist")
        return format_html(
            '<a href="{}?operator__noc__exact={}">{}</a>', url, obj.noc, obj.vehicles
        )

    def get_search_results(self, request, queryset, search_term):
        queryset, use_distinct = super().get_search_results(
            request, queryset, search_term
        )

        if (
            request.path.endswith("/autocomplete/")
            and not (
                request.GET.get("app_label") == "busstops"
                and request.GET.get("model_name") == "depot"
                and request.GET.get("field_name") == "operator"
            )
        ):
            queryset = queryset.filter(
                Q(noc=search_term)
                | Exists(
                    models.Service.objects.filter(operator=OuterRef("pk"), current=True)
                )
            )

        return queryset, use_distinct

    @staticmethod
    def payment(obj):
        return ", ".join(str(code) for code in obj.payment_methods.all())

    @staticmethod
    def operator_codes(obj):
        return ", ".join(str(code) for code in obj.operatorcode_set.all())


    def save_model(self, request, obj, form, change):
        if request.user.is_superuser:
            obj.is_manual = True
            obj.manual_updated_at = timezone.now()
        super().save_model(request, obj, form, change)

    def export_basic_fleet_view(self, request, object_id):
        from fleet.exporters.xlsx import build_basic_fleet_workbook, workbook_bytes
        from vehicles.models import Vehicle
        from vehicles.views import current_vehicle_filters
        
        operator = self.get_object(request, object_id)
        if operator is None:
            raise PermissionDenied
        
        vehicles = operator.vehicle_set.filter(**current_vehicle_filters()).select_related("vehicle_type", "livery").prefetch_related("features")
        if "withdrawn" not in request.GET:
            vehicles = vehicles.filter(**current_vehicle_filters(withdrawn=False))
        
        workbook = build_basic_fleet_workbook(operator, vehicles, advanced=False)
        response = HttpResponse(
            workbook_bytes(workbook),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{operator.slug}-fleet-basic.xlsx"'
        return response

    def export_advanced_fleet_view(self, request, object_id):
        from fleet.exporters.xlsx import build_basic_fleet_workbook, workbook_bytes
        from vehicles.models import Vehicle
        from vehicles.views import current_vehicle_filters
        
        if not request.user.advanced_mode:
            raise PermissionDenied
        
        operator = self.get_object(request, object_id)
        if operator is None:
            raise PermissionDenied
        
        vehicles = operator.vehicle_set.filter(**current_vehicle_filters()).select_related("vehicle_type", "livery").prefetch_related("features")
        if "withdrawn" not in request.GET:
            vehicles = vehicles.filter(**current_vehicle_filters(withdrawn=False))
        
        workbook = build_basic_fleet_workbook(operator, vehicles, advanced=True)
        response = HttpResponse(
            workbook_bytes(workbook),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{operator.slug}-fleet-advanced.xlsx"'
        return response


class ServiceCodeInline(admin.TabularInline):
    model = models.ServiceCode


class RouteInline(admin.TabularInline):
    model = Route
    show_change_link = True
    fields = ["source", "code", "service_code", "timetable_counts"]
    raw_id_fields = ["source"]
    readonly_fields = ["timetable_counts"]

    @admin.display(description="Timetable")
    def timetable_counts(self, obj):
        if not obj.pk:
            return ""
        trip_count = getattr(obj, "trip_count", None)
        stop_time_count = getattr(obj, "stop_time_count", None)
        if trip_count is None:
            trip_count = Trip.objects.filter(route=obj).count()
        if stop_time_count is None:
            stop_time_count = StopTime.objects.filter(trip__route=obj).count()
        return format_html("{} trips, {} stop times", trip_count, stop_time_count)

    def get_queryset(self, request):
        queryset = super().get_queryset(request)
        return queryset.annotate(
            trip_count=Count("trip", distinct=True),
            stop_time_count=Count("trip__stoptime", distinct=True),
        )


class FromServiceLinkInline(admin.TabularInline):
    model = models.ServiceLink
    fk_name = "from_service"
    autocomplete_fields = ["to_service"]

    def get_formset(self, request, obj=None, **kwargs):
        formset = super().get_formset(request, obj, **kwargs)
        if obj:
            # Filter to_service to only show services with the same operator(s)
            operator_ids = obj.operator.values_list('noc', flat=True)
            if operator_ids:
                formset.form.base_fields['to_service'].queryset = (
                    models.Service.objects.filter(operator__in=operator_ids)
                )
        return formset


class ToServiceLinkInline(FromServiceLinkInline):
    fk_name = "to_service"
    autocomplete_fields = ["from_service"]

    def get_formset(self, request, obj=None, **kwargs):
        formset = super().get_formset(request, obj, **kwargs)
        if obj:
            # Filter from_service to only show services with the same operator(s)
            operator_ids = obj.operator.values_list('noc', flat=True)
            if operator_ids:
                formset.form.base_fields['from_service'].queryset = (
                    models.Service.objects.filter(operator__in=operator_ids)
                )
        return formset


class ServicePaymentMethodInline(admin.TabularInline):
    model = models.ServicePaymentMethod
    autocomplete_fields = ["payment_method"]
    extra = 0


class SplitServiceFilter(DuplicateOperatorFilter):
    title = "split"
    parameter_name = "split"

    def queryset(self, request, queryset):
        if self.value():
            exists = Exists(
                Route.objects.filter(
                    Exists(
                        Route.objects.filter(
                            ~Q(service=OuterRef("service")),
                            source=OuterRef("source"),
                            service_code=OuterRef("service_code"),
                        )
                    ),
                    ~Q(service_code=""),
                    service=OuterRef("id"),
                )
            )
            queryset = queryset.filter(exists)

        return queryset


class DuplicateLineNameFilter(DuplicateOperatorFilter):
    title = "duplicate line name"
    parameter_name = "duplicate_line_name"

    def queryset(self, request, queryset):
        if self.value():
            exists = Exists(
                models.Service.objects.filter(
                    ~Q(id=OuterRef("id")),
                    line_name__iexact=OuterRef("line_name"),
                    source=OuterRef("source"),
                    # current=True,
                )
            )
            queryset = queryset.filter(exists)

        return queryset


@admin.register(models.Service)
class ServiceAdmin(GISModelAdmin):
    list_display = (
        "line_name",
        "line_brand",
        "description",
        "mode",
        "operators",
        "colour",
        "region_id",
        "current",
        "non_current_route",
        "event_specific",
        "school_route",
    )
    list_filter = (
        SplitServiceFilter,
        DuplicateLineNameFilter,
        "current",
        "non_current_route",
        "event_specific",
        "school_route",
        "timetable_wrong",
        "public_use",
        "tracking",
        "mode",
        "region",
        ("source", admin.RelatedOnlyFieldListFilter),
        ("operator", admin.RelatedOnlyFieldListFilter),
        "is_rail_replacement",
    )
    search_fields = ("service_code", "line_name", "line_brand", "description")
    raw_id_fields = ("operator", "stops", "colour", "source")
    inlines = [
        ServiceCodeInline,
        RouteInline,
        FromServiceLinkInline,
        ToServiceLinkInline,
        ServicePaymentMethodInline,
    ]
    readonly_fields = [
        "search_vector",
        "modified_at",
        "timetable_data",
        "timetable_bulk_edit_link",
        "route_editor_link",
    ]
    list_editable = ["colour", "line_brand"]
    list_select_related = ["colour"]
    actions = [
        "assign_to_operator",
        "current_false",
        "current_true",
        "public_use_true",
        "merge",
        "unmerge",
    ]

    def get_queryset(self, request):
        queryset = super().get_queryset(request)
        # Prevent querying disruptions tables if they don't exist
        try:
            # Test if disruptions tables exist by attempting a simple query
            from disruptions.models import Situation
            Situation.objects.none()
        except (ImportError, Exception):
            # If disruptions models are not available, don't prefetch related disruptions
            # by not allowing Django to automatically prefetch related objects
            queryset = queryset.select_related(None).prefetch_related(None)
        return queryset

    def get_deleted_objects(self, objs, request):
        """
        Override to handle missing disruptions tables during deletion.
        """
        try:
            # Test if disruptions tables exist
            from disruptions.models import Situation
            Situation.objects.none()
            return super().get_deleted_objects(objs, request)
        except (ImportError, Exception):
            # If disruptions tables don't exist, use a simplified deletion collector
            from django.contrib.admin.utils import get_deleted_objects
            from django.db.models import Q
            
            # Create a custom deletion collector that skips disruptions
            deleted_objects = []
            model_count = {self.model.__name__.lower(): len(objs)}
            perms_needed = set()
            
            # Add the objects being deleted
            for obj in objs:
                deleted_objects.append(str(obj))
            
            return deleted_objects, model_count, perms_needed

    @staticmethod
    def _touch_service(service):
        service.save(update_fields=["modified_at"])

    def _bulk_assign_operator(self, request, queryset):
        selected = request.POST.getlist(ACTION_CHECKBOX_NAME)
        if "apply" in request.POST:
            form = ServiceBulkAssignOperatorForm(request.POST)
            if form.is_valid():
                operator = form.cleaned_data["operator"]
                updated = 0
                for service in queryset:
                    service.operator.set([operator])
                    self._touch_service(service)
                    updated += 1
                self.message_user(
                    request,
                    f"Assigned {operator} to {updated} service{'s' if updated != 1 else ''}.",
                    level=messages.SUCCESS,
                )
                return None
        else:
            form = ServiceBulkAssignOperatorForm()

        context = {
            **self.admin_site.each_context(request),
            "opts": self.model._meta,
            "queryset": queryset.order_by("line_name", "id"),
            "form": form,
            "title": "Assign selected services to an operator",
            "submit_label": "Apply operator",
            "selected_objects_label": "service",
            "selected_objects_heading": "Selected services",
            "action_checkbox_name": ACTION_CHECKBOX_NAME,
            "selected": selected,
            "action_name": request.POST.get("action", ""),
        }
        return TemplateResponse(
            request,
            "admin/busstops/operator/bulk_assign.html",
            context,
        )

    @admin.action(description="Assign selected services to an operator")
    def assign_to_operator(self, request, queryset):
        return self._bulk_assign_operator(request, queryset)

    @admin.display(ordering="service_codes")
    def service_codes(self, obj):
        return obj.service_codes

    @admin.display(ordering="routes")
    def routes(self, obj):
        return obj.route_set.count()

    @admin.display(description="Operators")
    def operators(self, obj):
        return ", ".join(str(operator) for operator in obj.operator.all())

    @admin.display(description="Timetable data")
    def timetable_data(self, obj):
        if not obj.pk:
            return ""

        route_count = Route.objects.filter(service=obj).count()
        trip_count = Trip.objects.filter(route__service=obj).count()
        stop_time_count = StopTime.objects.filter(trip__route__service=obj).count()
        if not route_count:
            return format_html(
                '<p>No routes are linked to this service yet. Public URL: <a href="{}">{}</a></p>',
                obj.get_absolute_url(),
                obj.get_absolute_url(),
            )

        routes = (
            Route.objects.filter(service=obj)
            .select_related("source", "source__source")
            .annotate(
                trip_count=Count("trip", distinct=True),
                stop_time_count=Count("trip__stoptime", distinct=True),
            )
            .order_by("line_name", "id")[:30]
        )

        rows = format_html_join(
            "",
            '<tr><td><a href="{}">{}</a></td><td>{}</td><td>{}</td><td>{}</td><td>{}</td></tr>',
            (
                (
                    reverse("admin:bustimes_route_change", args=(route.pk,)),
                    route.line_name or route.pk,
                    route.description,
                    route.source,
                    route.trip_count,
                    route.stop_time_count,
                )
                for route in routes
            ),
        )
        return format_html(
            '<p><a href="{}">Open public timetable</a></p>'
            "<p>{} routes, {} trips, {} stop times.</p>"
            '<table><thead><tr><th>Route</th><th>Description</th><th>Data source</th><th>Trips</th><th>Stop times</th></tr></thead><tbody>{}</tbody></table>',
            obj.get_absolute_url(),
            route_count,
            trip_count,
            stop_time_count,
            rows,
        )

    @admin.display(description="Route editor")
    def route_editor_link(self, obj):
        if not obj or not obj.pk:
            return ""
        url = reverse("admin:busstops_service_route_editor")
        return format_html(
            '<a class="button" href="{}?service={}">Open route editor</a>',
            url,
            obj.pk,
        )

    @admin.display(description="Timetable bulk edit")
    def timetable_bulk_edit_link(self, obj):
        if not obj or not obj.pk:
            return ""
        url = reverse("admin:busstops_service_mass_edit_timetable", args=(obj.pk,))
        return format_html(
            '<a class="button" href="{}">Open timetable workbook editor</a>',
            url,
        )

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "<path:object_id>/mass-edit-timetable/",
                self.admin_site.admin_view(self.mass_edit_timetable_view),
                name="busstops_service_mass_edit_timetable",
            ),
            path(
                "<path:object_id>/mass-edit-timetable/template.xlsx",
                self.admin_site.admin_view(self.mass_edit_timetable_template_view),
                name="busstops_service_mass_edit_timetable_template",
            ),
            path(
                "<path:object_id>/mass-edit-timetable/simple-template.xlsx",
                self.admin_site.admin_view(self.mass_edit_timetable_simple_template_view),
                name="busstops_service_mass_edit_timetable_simple_template",
            ),
            path(
                "<path:object_id>/mass-edit-timetable/current.xlsx",
                self.admin_site.admin_view(self.mass_edit_timetable_current_view),
                name="busstops_service_mass_edit_timetable_current",
            ),
            path(
                "<path:object_id>/mass-edit-timetable/delete/",
                self.admin_site.admin_view(self.mass_delete_timetable_view),
                name="busstops_service_mass_delete_timetable",
            ),
            path(
                "tools/route-editor/",
                self.admin_site.admin_view(self.route_editor_view),
                name="busstops_service_route_editor",
            ),
        ]
        return custom_urls + urls

    timetable_header_aliases = {
        "trip": "trip_id",
        "route": "route_id",
        "calendar": "calendar_id",
        "stop_code": "stop_atco_code",
        "stop": "stop_atco_code",
        "destination": "destination_atco_code",
        "operator": "operator_noc",
        "garage": "garage_code",
        "vehicle_type": "vehicle_type_code",
        "direction": "inbound",
    }
    simple_timetable_trip_id_row = 1
    simple_timetable_route_id_row = 2
    simple_timetable_calendar_id_row = 3
    simple_timetable_inbound_row = 4
    simple_timetable_headsign_row = 5
    simple_timetable_import_key_row = 6
    simple_timetable_header_row = 8
    simple_timetable_data_row = 9

    def _normalise_timetable_header(self, header):
        key = header.strip().lower().replace(" ", "_")
        return self.timetable_header_aliases.get(key, key)

    @staticmethod
    def _simple_timetable_string(value):
        if value is None:
            return ""
        return str(value).strip()

    def _simple_timetable_time_string(self, value):
        if value is None:
            return ""
        if isinstance(value, str):
            text = value.strip()
            if re.fullmatch(r"\d{1,3}:\d{2}:\d{2}", text):
                return text[:-3]
            return text
        if hasattr(value, "hour") and hasattr(value, "minute"):
            return f"{value.hour:02d}:{value.minute:02d}"
        return str(value).strip()

    def _service_simple_timetable_data(self, service):
        trips = (
            Trip.objects.filter(route__service=service)
            .select_related(
                "route",
                "calendar",
                "destination",
                "operator",
                "garage",
                "vehicle_type",
            )
            .prefetch_related(
                Prefetch(
                    "stoptime_set",
                    queryset=StopTime.objects.select_related("stop").order_by("sequence", "id"),
                )
            )
            .order_by("route__line_name", "inbound", "start", "id")
        )

        ordered_stops = {False: [], True: []}
        seen_stop_codes = {False: set(), True: set()}
        trip_columns = {False: [], True: []}

        for trip in trips:
            direction = bool(trip.inbound)
            stop_map = {}
            for stop_time in trip.stoptime_set.all():
                stop_code = stop_time.stop_id or stop_time.stop_code
                stop_name = stop_time.display_name or (
                    stop_time.stop.common_name if stop_time.stop_id else stop_time.stop_code
                )
                if stop_code and stop_code not in seen_stop_codes[direction]:
                    seen_stop_codes[direction].add(stop_code)
                    ordered_stops[direction].append({"code": stop_code, "name": stop_name})
                if stop_code:
                    stop_map[stop_code] = stop_time

            trip_columns[direction].append(
                {
                    "trip_id": str(trip.pk),
                    "route_id": str(trip.route_id or ""),
                    "calendar_id": str(trip.calendar_id or ""),
                    "inbound": "true" if trip.inbound else "false",
                    "headsign": trip.headsign or "",
                    "import_key": f"trip-{trip.pk}",
                    "label": self._format_timetable_time(trip.start) or f"Trip {trip.pk}",
                    "stops": stop_map,
                }
            )

        if not ordered_stops[False] and not ordered_stops[True]:
            for route_link in RouteLink.objects.filter(service=service).select_related(
                "from_stop", "to_stop"
            ):
                for stop in (route_link.from_stop, route_link.to_stop):
                    if stop.atco_code not in seen_stop_codes[False]:
                        seen_stop_codes[False].add(stop.atco_code)
                        ordered_stops[False].append(
                            {"code": stop.atco_code, "name": stop.common_name}
                        )

        if not trip_columns[False] and not trip_columns[True]:
            route = Route.objects.filter(service=service).order_by("id").first()
            trip_columns[False].append(
                {
                    "trip_id": "",
                    "route_id": str(route.pk) if route else "",
                    "calendar_id": "",
                    "inbound": "false",
                    "headsign": "",
                    "import_key": "new-outbound-1",
                    "label": "Trip 1",
                    "stops": {},
                }
            )

        return ordered_stops, trip_columns

    def _build_simple_timetable_workbook(self, service):
        workbook = Workbook()
        ordered_stops, trip_columns = self._service_simple_timetable_data(service)
        timing_point_font = Font(bold=True)

        def populate_simple_direction_sheet(worksheet, *, inbound):
            sheet_name = "Inbound" if inbound else "Outbound"
            worksheet.title = sheet_name
            direction_stops = ordered_stops[inbound]
            direction_trips = trip_columns[inbound]

            if not direction_stops:
                fallback_stops = ordered_stops[False] or ordered_stops[True]
                direction_stops = fallback_stops
            if not direction_trips:
                route = Route.objects.filter(service=service).order_by("id").first()
                direction_trips = [
                    {
                        "trip_id": "",
                        "route_id": str(route.pk) if route else "",
                        "calendar_id": "",
                        "inbound": "true" if inbound else "false",
                        "headsign": "",
                        "import_key": f"new-{'inbound' if inbound else 'outbound'}-1",
                        "label": "Trip 1",
                        "stops": {},
                    }
                ]

            metadata_rows = (
                (self.simple_timetable_trip_id_row, "trip_id"),
                (self.simple_timetable_route_id_row, "route_id"),
                (self.simple_timetable_calendar_id_row, "calendar_id"),
                (self.simple_timetable_inbound_row, "inbound"),
                (self.simple_timetable_headsign_row, "headsign"),
                (self.simple_timetable_import_key_row, "import_key"),
            )
            for row_number, label in metadata_rows:
                worksheet.cell(row=row_number, column=1, value=label)
                worksheet.row_dimensions[row_number].hidden = True

            worksheet.cell(row=self.simple_timetable_header_row, column=1, value="Stop name")
            worksheet.cell(
                row=self.simple_timetable_header_row, column=2, value="stop_atco_code"
            )

            for row_number, stop in enumerate(
                direction_stops, start=self.simple_timetable_data_row
            ):
                name_cell = worksheet.cell(row=row_number, column=1, value=stop["name"])
                if any(
                    trip["stops"].get(stop["code"])
                    and trip["stops"][stop["code"]].timing_status == "PTP"
                    for trip in direction_trips
                ):
                    name_cell.font = timing_point_font
                worksheet.cell(row=row_number, column=2, value=stop["code"])

            for column_number, trip in enumerate(direction_trips, start=3):
                worksheet.cell(
                    row=self.simple_timetable_trip_id_row,
                    column=column_number,
                    value=trip["trip_id"],
                )
                worksheet.cell(
                    row=self.simple_timetable_route_id_row,
                    column=column_number,
                    value=trip["route_id"],
                )
                worksheet.cell(
                    row=self.simple_timetable_calendar_id_row,
                    column=column_number,
                    value=trip["calendar_id"],
                )
                worksheet.cell(
                    row=self.simple_timetable_inbound_row,
                    column=column_number,
                    value=trip["inbound"],
                )
                worksheet.cell(
                    row=self.simple_timetable_headsign_row,
                    column=column_number,
                    value=trip["headsign"],
                )
                worksheet.cell(
                    row=self.simple_timetable_import_key_row,
                    column=column_number,
                    value=trip["import_key"],
                )
                worksheet.cell(
                    row=self.simple_timetable_header_row,
                    column=column_number,
                    value=trip["label"],
                )

                for row_number, stop in enumerate(
                    direction_stops, start=self.simple_timetable_data_row
                ):
                    stop_time = trip["stops"].get(stop["code"])
                    if not stop_time:
                        continue
                    cell = worksheet.cell(
                        row=row_number,
                        column=column_number,
                        value=self._format_timetable_time(
                            stop_time.departure or stop_time.arrival
                        ),
                    )

            worksheet.freeze_panes = f"C{self.simple_timetable_data_row}"

        worksheet = workbook.active
        populate_simple_direction_sheet(worksheet, inbound=False)
        workbook.create_sheet("Inbound")
        populate_simple_direction_sheet(workbook["Inbound"], inbound=True)

        instructions = workbook.create_sheet("Instructions")
        instructions.append(["How to use", "Notes"])
        instructions.append(
            [
                "Stop names",
                "Column A is the visible stop list and is always used in the published timetable. Column B contains the ATCO stop code used for validation and location.",
            ]
        )
        instructions.append(
            [
                "Trip columns",
                "Each column from C onward is one trip. Edit the times directly in the grid. On this service-specific screen, blank route_id values will use the current service route automatically.",
            ]
        )
        instructions.append(
            [
                "Timing points",
                "Bold stop names in column A are imported as timing points. Non-bold stop names are imported as non-timing points.",
            ]
        )
        instructions.append(
            [
                "New trips",
                "To create a new trip, copy an existing trip column, then unhide the top rows and clear trip_id while giving import_key a new value.",
            ]
        )

        return workbook

    def _rows_text_from_simple_timetable_workbook(self, workbook, service=None):
        output = StringIO()
        writer = csv.writer(output, delimiter="\t", lineterminator="\n")
        headers = self._timetable_headers()
        writer.writerow(headers)
        default_route_id = ""
        if service is not None:
            default_route_id = str(
                Route.objects.filter(service=service).order_by("id").values_list("pk", flat=True).first()
                or ""
            )

        for sheet_name in ("Outbound", "Inbound", "Simple Timetable"):
            if sheet_name not in workbook.sheetnames:
                continue
            worksheet = workbook[sheet_name]
            default_inbound = ""
            if sheet_name == "Outbound":
                default_inbound = "false"
            elif sheet_name == "Inbound":
                default_inbound = "true"
            for column_number in range(3, worksheet.max_column + 1):
                trip_id = self._simple_timetable_string(
                    worksheet.cell(
                        row=self.simple_timetable_trip_id_row, column=column_number
                    ).value
                )
                route_id = self._simple_timetable_string(
                    worksheet.cell(
                        row=self.simple_timetable_route_id_row, column=column_number
                    ).value
                )
                if not route_id:
                    route_id = default_route_id
                calendar_id = self._simple_timetable_string(
                    worksheet.cell(
                        row=self.simple_timetable_calendar_id_row, column=column_number
                    ).value
                )
                inbound = self._simple_timetable_string(
                    worksheet.cell(
                        row=self.simple_timetable_inbound_row, column=column_number
                    ).value
                )
                if not inbound:
                    inbound = default_inbound
                headsign = self._simple_timetable_string(
                    worksheet.cell(
                        row=self.simple_timetable_headsign_row, column=column_number
                    ).value
                )
                import_key = self._simple_timetable_string(
                    worksheet.cell(
                        row=self.simple_timetable_import_key_row, column=column_number
                    ).value
                )

                trip_rows = []
                for row_number in range(self.simple_timetable_data_row, worksheet.max_row + 1):
                    stop_name = self._simple_timetable_string(
                        worksheet.cell(row=row_number, column=1).value
                    )
                    stop_code = self._simple_timetable_string(
                        worksheet.cell(row=row_number, column=2).value
                    )
                    cell = worksheet.cell(row=row_number, column=column_number)
                    time_value = self._simple_timetable_time_string(cell.value)
                    if not time_value:
                        continue
                    is_timing_point = bool(
                        worksheet.cell(row=row_number, column=1).font.bold
                        or cell.font.bold
                    )
                    trip_rows.append((stop_name, stop_code, time_value, is_timing_point))

                if not trip_rows and not any(
                    [trip_id, route_id, calendar_id, inbound, headsign, import_key]
                ):
                    continue

                if not import_key:
                    import_key = f"simple-{sheet_name.lower()}-{column_number - 2}"
                elif import_key.startswith("new-"):
                    import_key = f"{sheet_name.lower()}-{import_key}"

                for sequence, (stop_name, stop_code, time_value, is_timing_point) in enumerate(
                    trip_rows, start=1
                ):
                    writer.writerow(
                        [
                            import_key,
                            trip_id,
                            route_id,
                            "",
                            calendar_id,
                            inbound,
                            sequence,
                            stop_code,
                            stop_name,
                            time_value,
                            time_value,
                            "true",
                            "true",
                            "PTP" if is_timing_point else "OTH",
                            "",
                            headsign,
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                        ]
                    )

        return output.getvalue().strip()

    def _rows_text_from_workbook(self, uploaded_file, service=None):
        if not uploaded_file:
            return ""

        filename = (uploaded_file.name or "").lower()
        if filename.endswith(".csv"):
            try:
                content = uploaded_file.read().decode("utf-8-sig")
            except UnicodeDecodeError as exc:
                raise ValueError("CSV upload must be UTF-8 encoded") from exc
            return content.strip()
        if not filename.endswith(".xlsx"):
            raise ValueError("Upload must be a .xlsx or .csv file")

        workbook = load_workbook(uploaded_file, data_only=True)
        if (
            "Simple Timetable" in workbook.sheetnames
            or "Outbound" in workbook.sheetnames
            or "Inbound" in workbook.sheetnames
        ):
            return self._rows_text_from_simple_timetable_workbook(workbook, service=service)
        worksheet = workbook["Timetable"] if "Timetable" in workbook.sheetnames else workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return ""

        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        if not any(headers):
            return ""

        output = StringIO()
        writer = csv.writer(output, delimiter="\t", lineterminator="\n")
        writer.writerow(headers)
        for row in rows[1:]:
            values = ["" if value is None else str(value).strip() for value in row[: len(headers)]]
            if any(values):
                writer.writerow(values)
        return output.getvalue()

    def _rows_text_from_pdf(self, operator, uploaded_file):
        rows = parse_pdf(uploaded_file, default_operator_code=operator.noc)
        output = StringIO()
        writer = csv.writer(output, delimiter="\t", lineterminator="\n")
        writer.writerow(list(TARGET_COLUMNS) + ["features"])
        for row in rows:
            writer.writerow(
                [
                    row.operator_code or operator.noc,
                    row.external_id,
                    row.code,
                    row.fleet_number,
                    row.fleet_code,
                    row.registration,
                    row.prev_registration,
                    row.vehicle_type,
                    row.livery,
                    row.colours,
                    row.garage,
                    row.name,
                    row.branding,
                    row.notes,
                    "true" if row.withdrawn else "false",
                    "true" if row.preserved else "false",
                    "true" if row.fleet_support_vehicle else "false",
                    "true" if row.vor else "false",
                    "true" if row.awaiting_delivery else "false",
                    "true" if row.trainer_vehicle else "false",
                    "true" if row.demonstrator else "false",
                    "",
                ]
            )
        return output.getvalue()

    def _rows_text_from_upload(self, operator, uploaded_file):
        if not uploaded_file:
            return ""
        filename = (uploaded_file.name or "").lower()
        if filename.endswith(".pdf"):
            return self._rows_text_from_pdf(operator, uploaded_file)
        return self._rows_text_from_workbook(uploaded_file)

    @staticmethod
    def _parse_timetable_bool(value, *, default=None):
        value = (value or "").strip().lower()
        if value in {"", "none", "null"}:
            return default
        if value in {"1", "true", "yes", "y", "on"}:
            return True
        if value in {"0", "false", "no", "n", "off"}:
            return False
        raise ValueError(f"Invalid boolean value '{value}'")

    @staticmethod
    def _parse_timetable_time(value):
        text = (value or "").strip()
        if not text:
            return None
        match = re.fullmatch(r"(?P<hours>\d{1,3}):(?P<minutes>\d{2})", text)
        if not match:
            raise ValueError(f"Invalid time '{text}'. Use HH:MM.")
        hours = int(match.group("hours"))
        minutes = int(match.group("minutes"))
        if minutes >= 60:
            raise ValueError(f"Invalid time '{text}'. Minutes must be below 60.")
        return timedelta(hours=hours, minutes=minutes)

    @staticmethod
    def _format_timetable_time(value):
        if value is None:
            return ""
        total_seconds = int(value.total_seconds())
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        return f"{hours:02d}:{minutes:02d}"

    @staticmethod
    def _timetable_headers():
        return [
            "import_key",
            "trip_id",
            "route_id",
            "line_name",
            "calendar_id",
            "inbound",
            "sequence",
            "stop_atco_code",
            "stop_name",
            "arrival",
            "departure",
            "pick_up",
            "set_down",
            "timing_status",
            "destination_atco_code",
            "headsign",
            "block",
            "ticket_machine_code",
            "vehicle_journey_code",
            "operator_noc",
            "garage_code",
            "vehicle_type_code",
        ]

    def _build_timetable_workbook(self, rows=None):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Timetable"
        headers = self._timetable_headers()
        worksheet.append(headers)
        if rows:
            for row in rows:
                worksheet.append([row.get(header, "") for header in headers])

        instructions = workbook.create_sheet("Instructions")
        instructions.append(["column", "notes"])
        instructions.append(["import_key", "Required for new trips. Reuse the same key on every row belonging to the same new trip."])
        instructions.append(["trip_id", "Existing trip id for edits. Leave blank for new trips."])
        instructions.append(["route_id", "Required for new trips. Must belong to this service."])
        instructions.append(["calendar_id", "Optional existing calendar id. Leave blank for no calendar."])
        instructions.append(["inbound", "Boolean: true/false, yes/no, 1/0."])
        instructions.append(["sequence", "Required stop order within the trip."])
        instructions.append(["stop_atco_code", "Required ATCO stop code."])
        instructions.append(["arrival/departure", "Use HH:MM. Hours may be 24+ for after-midnight trips."])
        instructions.append(["pick_up/set_down", "Boolean values. Blank defaults to true."])
        instructions.append(["timing_status", "Optional timing status such as PTP."])
        instructions.append(["destination_atco_code", "Optional destination stop ATCO code for the trip."])
        instructions.append(["headsign/block/ticket_machine_code/vehicle_journey_code", "Optional trip-level fields. Repeat the same values on each row in a trip."])
        instructions.append(["operator_noc", "Optional operator NOC/slug for the trip."])
        instructions.append(["garage_code", "Optional garage id, code, external id, or name."])
        instructions.append(["vehicle_type_code", "Optional timetable vehicle type id, code, or description."])
        instructions.append(["line_name", "Informational export column. It is ignored on import."])
        instructions.append(["stop_name", "Optional published stop label. It is retained on import."])
        return workbook

    def _service_timetable_export_rows(self, service):
        trips = (
            Trip.objects.filter(route__service=service)
            .select_related(
                "route",
                "calendar",
                "destination",
                "operator",
                "garage",
                "vehicle_type",
            )
            .prefetch_related(
                Prefetch(
                    "stoptime_set",
                    queryset=StopTime.objects.select_related("stop").order_by("sequence", "id"),
                )
            )
            .order_by("inbound", "route__line_name", "start", "id")
        )

        rows = []
        for trip in trips:
            # Ensure stops are in sequence order
            stop_times = sorted(trip.stoptime_set.all(), key=lambda st: (st.sequence or 0, st.id))
            for stop_time in stop_times:
                rows.append(
                    {
                        "import_key": f"trip-{trip.id}",
                        "trip_id": trip.id,
                        "route_id": trip.route_id or "",
                        "line_name": trip.route.line_name if trip.route_id else "",
                        "calendar_id": trip.calendar_id or "",
                        "inbound": "true" if trip.inbound else "false",
                        "sequence": stop_time.sequence if stop_time.sequence is not None else "",
                        "stop_atco_code": stop_time.stop_id or stop_time.stop_code,
                        "stop_name": stop_time.display_name
                        or (stop_time.stop.common_name if stop_time.stop_id else ""),
                        "arrival": self._format_timetable_time(stop_time.arrival),
                        "departure": self._format_timetable_time(stop_time.departure),
                        "pick_up": "true" if stop_time.pick_up else "false",
                        "set_down": "true" if stop_time.set_down else "false",
                        "timing_status": stop_time.timing_status or "",
                        "destination_atco_code": trip.destination_id or "",
                        "headsign": trip.headsign or "",
                        "block": trip.block or "",
                        "ticket_machine_code": trip.ticket_machine_code or "",
                        "vehicle_journey_code": trip.vehicle_journey_code or "",
                        "operator_noc": trip.operator.noc if trip.operator_id else "",
                        "garage_code": trip.garage.code if trip.garage_id else "",
                        "vehicle_type_code": trip.vehicle_type.code if trip.vehicle_type_id else "",
                    }
                )
        return rows

    def mass_edit_timetable_template_view(self, request, object_id):
        if not request.user.is_superuser:
            raise PermissionDenied
        service = self.get_object(request, object_id)
        if service is None:
            raise PermissionDenied

        workbook = self._build_timetable_workbook()
        stream = BytesIO()
        workbook.save(stream)
        response = HttpResponse(
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{service.slug}-timetable-template.xlsx"'
        )
        return response

    def mass_edit_timetable_simple_template_view(self, request, object_id):
        if not request.user.is_superuser:
            raise PermissionDenied
        service = self.get_object(request, object_id)
        if service is None:
            raise PermissionDenied

        workbook = self._build_simple_timetable_workbook(service)
        stream = BytesIO()
        workbook.save(stream)
        response = HttpResponse(
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{service.slug}-simple-timetable-template.xlsx"'
        )
        return response

    def mass_edit_timetable_current_view(self, request, object_id):
        if not request.user.is_superuser:
            raise PermissionDenied
        service = self.get_object(request, object_id)
        if service is None:
            raise PermissionDenied

        workbook = self._build_timetable_workbook(self._service_timetable_export_rows(service))
        stream = BytesIO()
        workbook.save(stream)
        response = HttpResponse(
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{service.slug}-timetable.xlsx"'
        )
        return response

    def mass_delete_timetable_view(self, request, object_id):
        if not request.user.is_superuser:
            raise PermissionDenied
        service = self.get_object(request, object_id)
        if service is None:
            raise PermissionDenied

        if request.method == "POST":
            deleted_trips = Trip.objects.filter(route__service=service).count()
            deleted_routes = Route.objects.filter(service=service).count()
            
            with transaction.atomic():
                Trip.objects.filter(route__service=service).delete()
                Route.objects.filter(service=service).delete()
                service.update_geometry()
                self._touch_service(service)

            self.message_user(
                request,
                f"Deleted {deleted_trips} trips and {deleted_routes} routes for {service}.",
                level=messages.SUCCESS,
            )
            return redirect("admin:busstops_service_mass_edit_timetable", args=(object_id,))

        context = {
            **self.admin_site.each_context(request),
            "opts": self.model._meta,
            "original": service,
            "service": service,
            "title": f"Delete timetable for {service}",
        }
        return TemplateResponse(
            request,
            "admin/busstops/service/delete_timetable.html",
            context,
        )

    def _resolve_timetable_operator(self, value):
        if value == "":
            return None
        if value.isdigit():
            return models.Operator.objects.filter(pk=int(value)).first()
        return models.Operator.objects.filter(
            Q(noc__iexact=value) | Q(slug__iexact=value) | Q(operatorcode__code__iexact=value)
        ).first()

    def _resolve_timetable_garage(self, value):
        if value == "":
            return None
        if value.isdigit():
            return Garage.objects.filter(pk=int(value)).first()
        return Garage.objects.filter(
            Q(code__iexact=value) | Q(external_id__iexact=value) | Q(name__iexact=value)
        ).first()

    def _resolve_timetable_vehicle_type(self, value):
        if value == "":
            return None
        if value.isdigit():
            return TimetableVehicleType.objects.filter(pk=int(value)).first()
        return TimetableVehicleType.objects.filter(
            Q(code__iexact=value) | Q(description__iexact=value)
        ).first()

    def _get_or_create_manual_timetable_route(self, service):
        existing_route = Route.objects.filter(service=service).order_by("id").first()
        if existing_route:
            return existing_route

        source = service.source
        if source is None:
            source, _ = models.DataSource.objects.get_or_create(name="Manual timetable")

        code = f"manual-service-{service.pk}"
        route, _ = Route.objects.get_or_create(
            source=source,
            code=code,
            defaults={
                "service": service,
                "service_code": service.service_code or code,
                "line_name": service.line_name or "",
                "description": service.description or "",
            },
        )
        if route.service_id != service.pk:
            route.service = service
            if not route.service_code:
                route.service_code = service.service_code or code
            if not route.line_name:
                route.line_name = service.line_name or ""
            if not route.description:
                route.description = service.description or ""
            route.save(
                update_fields=["service", "service_code", "line_name", "description"]
            )
        return route

    def _parse_timetable_rows(self, service, rows_text):
        parsed_rows = []
        groups = {}
        default_route = self._get_or_create_manual_timetable_route(service)
        default_route_id = default_route.pk if default_route else None

        if not rows_text.strip():
            return parsed_rows

        delimiter = "\t" if "\t" in rows_text.splitlines()[0] else ","
        reader = csv.DictReader(StringIO(rows_text), delimiter=delimiter)
        if not reader.fieldnames:
            return parsed_rows

        for index, original_row in enumerate(reader, start=2):
            mapped = {}
            for key, value in original_row.items():
                if not key:
                    continue
                mapped[self._normalise_timetable_header(key)] = (value or "").strip()
            if not any(mapped.values()):
                continue

            trip_id_value = mapped.get("trip_id", "")
            import_key = mapped.get("import_key", "")
            group_key = trip_id_value or import_key
            entry = groups.setdefault(
                group_key or f"row-{index}",
                {
                    "row_numbers": [],
                    "raw_rows": [],
                    "errors": [],
                    "trip": None,
                    "route": None,
                    "calendar": None,
                    "destination": None,
                    "operator": None,
                    "garage": None,
                    "vehicle_type": None,
                    "stop_rows": [],
                    "action": "create",
                    "identifier": group_key or f"row-{index}",
                },
            )
            entry["row_numbers"].append(index)
            entry["raw_rows"].append(mapped)

        for entry in groups.values():
            first = entry["raw_rows"][0]
            trip_id_value = first.get("trip_id", "")
            import_key_value = first.get("import_key", "")
            route_id_value = first.get("route_id", "")
            calendar_value = first.get("calendar_id", "")
            inbound_value = first.get("inbound", "")
            destination_value = first.get("destination_atco_code", "")
            headsign_value = first.get("headsign", "")
            block_value = first.get("block", "")
            ticket_machine_code_value = first.get("ticket_machine_code", "")
            vehicle_journey_code_value = first.get("vehicle_journey_code", "")
            operator_value = first.get("operator_noc", "")
            garage_value = first.get("garage_code", "")
            vehicle_type_value = first.get("vehicle_type_code", "")

            for row in entry["raw_rows"][1:]:
                for field, expected in (
                    ("trip_id", trip_id_value),
                    ("route_id", route_id_value),
                    ("calendar_id", calendar_value),
                    ("inbound", inbound_value),
                    ("destination_atco_code", destination_value),
                    ("headsign", headsign_value),
                    ("block", block_value),
                    ("ticket_machine_code", ticket_machine_code_value),
                    ("vehicle_journey_code", vehicle_journey_code_value),
                    ("operator_noc", operator_value),
                    ("garage_code", garage_value),
                    ("vehicle_type_code", vehicle_type_value),
                ):
                    if (row.get(field, "") or "") != (expected or ""):
                        entry["errors"].append(
                            f"Trip-level field '{field}' must be the same on every row for {entry['identifier']}."
                        )
                        break

            if trip_id_value:
                if not trip_id_value.isdigit():
                    entry["errors"].append(f"trip_id must be numeric for {entry['identifier']}.")
                else:
                    entry["trip"] = Trip.objects.filter(
                        pk=int(trip_id_value), route__service=service
                    ).select_related("route").first()
                    if not entry["trip"]:
                        entry["errors"].append(
                            f"Trip {trip_id_value} does not belong to this service."
                        )
                    else:
                        entry["action"] = "update"
            elif not import_key_value:
                entry["errors"].append(
                    "import_key is required for new trips so rows can be grouped together."
                )

            route_id_to_use = route_id_value or (entry["trip"].route_id if entry["trip"] else "")
            if not route_id_to_use and default_route_id:
                route_id_to_use = default_route_id
            if not route_id_to_use:
                entry["errors"].append(
                    f"route_id is required for new trip {entry['identifier']}."
                )
            elif not str(route_id_to_use).isdigit():
                entry["errors"].append(f"route_id must be numeric for {entry['identifier']}.")
            else:
                entry["route"] = Route.objects.filter(
                    pk=int(route_id_to_use), service=service
                ).first()
                if not entry["route"]:
                    entry["errors"].append(
                        f"Route {route_id_to_use} does not belong to this service."
                    )
            if entry["route"] is None and default_route is not None:
                entry["route"] = default_route

            calendar_to_use = calendar_value
            if calendar_to_use:
                if not calendar_to_use.isdigit():
                    entry["errors"].append(
                        f"calendar_id must be numeric for {entry['identifier']}."
                    )
                else:
                    entry["calendar"] = Calendar.objects.filter(pk=int(calendar_to_use)).first()
                    if not entry["calendar"]:
                        entry["errors"].append(
                            f"Calendar {calendar_to_use} was not found."
                        )

            if destination_value:
                entry["destination"] = models.StopPoint.objects.filter(
                    atco_code=destination_value
                ).first()
                if not entry["destination"]:
                    entry["errors"].append(
                        f"Unknown destination_atco_code '{destination_value}'."
                    )

            if operator_value:
                entry["operator"] = self._resolve_timetable_operator(operator_value)
                if not entry["operator"]:
                    entry["errors"].append(f"Unknown operator_noc '{operator_value}'.")

            if garage_value:
                entry["garage"] = self._resolve_timetable_garage(garage_value)
                if not entry["garage"]:
                    entry["errors"].append(f"Unknown garage_code '{garage_value}'.")

            if vehicle_type_value:
                entry["vehicle_type"] = self._resolve_timetable_vehicle_type(vehicle_type_value)
                if not entry["vehicle_type"]:
                    entry["errors"].append(
                        f"Unknown vehicle_type_code '{vehicle_type_value}'."
                    )

            try:
                entry["inbound"] = self._parse_timetable_bool(
                    inbound_value,
                    default=entry["trip"].inbound if entry["trip"] else False,
                )
            except ValueError as exc:
                entry["errors"].append(str(exc))

            seen_sequences = set()
            for row_number, row in zip(entry["row_numbers"], entry["raw_rows"]):
                sequence_value = row.get("sequence", "")
                if not sequence_value:
                    entry["errors"].append(f"sequence is required on row {row_number}.")
                    continue
                try:
                    sequence = int(sequence_value)
                except ValueError:
                    entry["errors"].append(f"sequence must be an integer on row {row_number}.")
                    continue
                if sequence in seen_sequences:
                    entry["errors"].append(
                        f"Duplicate sequence {sequence} in {entry['identifier']}."
                    )
                    continue
                seen_sequences.add(sequence)

                stop_code = row.get("stop_atco_code", "")
                stop = models.StopPoint.objects.filter(atco_code=stop_code).first()
                if not stop:
                    entry["errors"].append(
                        f"Unknown stop_atco_code '{stop_code}' on row {row_number}."
                    )
                    continue

                try:
                    arrival = self._parse_timetable_time(row.get("arrival", ""))
                    departure = self._parse_timetable_time(row.get("departure", ""))
                except ValueError as exc:
                    entry["errors"].append(f"{exc} Row {row_number}.")
                    continue

                if arrival is None and departure is None:
                    entry["errors"].append(
                        f"Either arrival or departure is required on row {row_number}."
                    )
                    continue

                try:
                    pick_up = self._parse_timetable_bool(row.get("pick_up", ""), default=True)
                    set_down = self._parse_timetable_bool(row.get("set_down", ""), default=True)
                except ValueError as exc:
                    entry["errors"].append(f"{exc} Row {row_number}.")
                    continue

                entry["stop_rows"].append(
                    {
                        "row_number": row_number,
                        "sequence": sequence,
                        "stop": stop,
                        "arrival": arrival,
                        "departure": departure,
                        "pick_up": pick_up,
                        "set_down": set_down,
                        "timing_status": row.get("timing_status", ""),
                        "display_name": row.get("stop_name", ""),
                    }
                )

            entry["stop_rows"].sort(key=lambda row: (row["sequence"], row["row_number"]))
            if not entry["stop_rows"]:
                entry["errors"].append(f"No valid stop rows were provided for {entry['identifier']}.")
            else:
                first_row = entry["stop_rows"][0]
                last_row = entry["stop_rows"][-1]
                entry["start"] = first_row["departure"] or first_row["arrival"]
                entry["end"] = last_row["arrival"] or last_row["departure"]
                if entry["start"] is None or entry["end"] is None:
                    entry["errors"].append(
                        f"Could not determine start/end times for {entry['identifier']}."
                    )

            entry["headsign"] = headsign_value or None
            entry["block"] = block_value or None
            entry["ticket_machine_code"] = ticket_machine_code_value or None
            entry["vehicle_journey_code"] = vehicle_journey_code_value or None

            parsed_rows.append(entry)

        return parsed_rows

    def _commit_timetable_rows(self, service, parsed_rows):
        created = 0
        updated = 0
        errors = 0

        for entry in parsed_rows:
            if entry["errors"]:
                errors += 1
                continue

            try:
                with transaction.atomic():
                    trip = entry["trip"] or Trip()
                    trip.route = entry["route"]
                    trip.calendar = entry["calendar"]
                    trip.inbound = entry["inbound"]
                    trip.destination = entry["destination"]
                    trip.headsign = entry["headsign"]
                    trip.block = entry["block"]
                    trip.ticket_machine_code = entry["ticket_machine_code"]
                    trip.vehicle_journey_code = entry["vehicle_journey_code"]
                    trip.operator = entry["operator"]
                    trip.garage = entry["garage"]
                    trip.vehicle_type = entry["vehicle_type"]
                    trip.start = entry["start"]
                    trip.end = entry["end"]
                    trip.save()

                    StopTime.objects.filter(trip=trip).delete()
                    StopTime.objects.bulk_create(
                        [
                            StopTime(
                                trip=trip,
                                stop=stop_row["stop"],
                                stop_code=stop_row["stop"].atco_code,
                                display_name=stop_row["display_name"],
                                arrival=stop_row["arrival"],
                                departure=stop_row["departure"],
                                sequence=stop_row["sequence"],
                                timing_status=stop_row["timing_status"],
                                pick_up=stop_row["pick_up"],
                                set_down=stop_row["set_down"],
                            )
                            for stop_row in entry["stop_rows"]
                        ]
                    )

                    if entry["action"] == "create":
                        created += 1
                    else:
                        updated += 1
            except Exception as exc:
                entry["errors"].append(str(exc))
                errors += 1

        if created or updated:
            service.do_stop_usages()
            service.update_geometry()
            self._touch_service(service)

        return created, updated, errors

    def _create_operator_from_code(self, code):
        code = str(code or "").strip().upper()
        slug_base = slugify(code) or "operator"
        slug = slug_base
        suffix = 2
        while models.Operator.objects.filter(slug=slug).exists():
            slug = f"{slug_base}-{suffix}"
            suffix += 1
        return models.Operator.objects.create(
            noc=code,
            name=code,
            slug=slug,
            is_manual=True,
            manual_updated_at=timezone.now(),
        )

    def _create_garage_for_operator(self, operator, garage_name):
        garage_name = str(garage_name or "").strip()
        trimmed = garage_name[4:].strip() if garage_name.upper().startswith("GSC ") else garage_name
        garage = Garage.objects.filter(operators=operator).filter(
            Q(name__iexact=trimmed) | Q(code__iexact=trimmed)
        ).first()
        if garage:
            return garage
        garage = Garage.objects.create(
            name=trimmed,
            code=trimmed,
            is_manual=True,
            manual_updated_at=timezone.now(),
        )
        garage.operators.add(operator)
        return garage

    def mass_edit_timetable_view(self, request, object_id):
        if not request.user.is_superuser:
            raise PermissionDenied

        service = self.get_object(request, object_id)
        if service is None:
            raise PermissionDenied

        rows = []
        created = 0
        updated = 0
        errors = 0

        if request.method == "POST":
            form = MassEditTimetableForm(request.POST, request.FILES)
            if form.is_valid():
                rows_text = form.cleaned_data.get("rows_text") or ""
                workbook = form.cleaned_data.get("workbook")
                try:
                    if workbook:
                        rows_text = self._rows_text_from_workbook(workbook, service=service)
                except ValueError as exc:
                    form.add_error("workbook", str(exc))

                if not form.errors and not rows_text.strip():
                    form.add_error(None, "Paste rows or upload a completed workbook.")

                if not form.errors:
                    rows = self._parse_timetable_rows(service, rows_text)
                    form = MassEditTimetableForm(initial={"rows_text": rows_text})
                    if request.POST.get("action") == "commit":
                        created, updated, errors = self._commit_timetable_rows(service, rows)
                        if created or updated:
                            self.message_user(
                                request,
                                f"Timetable import complete: created {created}, updated {updated}, errors {errors}",
                            )
                        elif errors:
                            self.message_user(
                                request,
                                f"No trips imported. {errors} item(s) had errors.",
                                level=messages.WARNING,
                            )
                    else:
                        self.message_user(
                            request,
                            "Preview generated. Review changes and click Commit changes when ready.",
                        )
        else:
            form = MassEditTimetableForm()

        context = {
            **self.admin_site.each_context(request),
            "opts": self.model._meta,
            "original": service,
            "service": service,
            "title": f"Mass edit timetable for {service}",
            "form": form,
            "rows": rows,
            "can_commit": any(not row["errors"] for row in rows),
            "created": created,
            "updated": updated,
            "errors": errors,
            "commit_label": "Commit changes",
            "template_download_url": reverse(
                "admin:busstops_service_mass_edit_timetable_template",
                args=(service.pk,),
            ),
            "simple_template_download_url": reverse(
                "admin:busstops_service_mass_edit_timetable_simple_template",
                args=(service.pk,),
            ),
            "export_download_url": reverse(
                "admin:busstops_service_mass_edit_timetable_current",
                args=(service.pk,),
            ),
            "delete_timetable_url": reverse(
                "admin:busstops_service_mass_delete_timetable",
                args=(service.pk,),
            ),
        }

        return TemplateResponse(
            request,
            "admin/busstops/service/mass_timetable.html",
            context,
        )

    def route_editor_view(self, request):
        if not request.user.is_staff:
            raise PermissionDenied

        query = (request.GET.get("q") or "").strip()
        service_id = request.GET.get("service") or request.POST.get("service")
        inbound = request.GET.get("inbound") or request.POST.get("inbound")
        inbound = inbound == "1"
        service = None
        if service_id and service_id.isdigit():
            service = models.Service.objects.filter(pk=service_id).first()

        if request.method == "POST" and service:
            action = request.POST.get("action") or "save_geometry"

            if action == "save_stops":
                raw_stops = (
                    request.POST.get("stops_selection")
                    or request.POST.get("stop_codes")
                    or ""
                ).strip()
                stop_ids = []
                seen = set()
                for stop_id in re.split(r"[\s,]+", raw_stops):
                    stop_id = stop_id.strip()
                    if not stop_id or stop_id in seen:
                        continue
                    seen.add(stop_id)
                    stop_ids.append(stop_id)

                if not stop_ids:
                    self.message_user(
                        request,
                        "Add at least one stop code to create the stop chain.",
                        level=messages.ERROR,
                    )
                    return TemplateResponse(
                        request,
                        "admin/busstops/service/route_editor.html",
                        self.admin_site.each_context(request)
                        | self._route_editor_context(query, service, raw_stops, "", inbound),
                    )

                stops = models.StopPoint.objects.in_bulk(stop_ids)
                missing = [stop_id for stop_id in stop_ids if stop_id not in stops]
                if missing:
                    self.message_user(
                        request,
                        f"Unknown stop code(s): {', '.join(missing[:10])}",
                        level=messages.ERROR,
                    )
                    return TemplateResponse(
                        request,
                        "admin/busstops/service/route_editor.html",
                        self.admin_site.each_context(request)
                        | self._route_editor_context(query, service, raw_stops, "", inbound),
                    )

                line_name = service.get_line_name() or service.line_name or ""
                with transaction.atomic():
                    service.stopusage_set.filter(inbound=inbound).delete()
                    models.StopUsage.objects.bulk_create(
                        [
                            models.StopUsage(
                                service=service,
                                stop_id=stop_id,
                                order=index,
                                timing_point=True,
                                inbound=inbound,
                                line_name=line_name,
                            )
                            for index, stop_id in enumerate(stop_ids)
                        ]
                    )
                    service.update_geometry()
                    self._touch_service(service)

                direction = "inbound" if inbound else "outbound"
                self.message_user(
                    request,
                    f"Saved {len(stop_ids)} stops for the {direction} chain.",
                    level=messages.SUCCESS,
                )

            elif action == "generate_geometry":
                # Parse waypoints from form data
                waypoints_text = request.POST.get("waypoints", "").strip()
                waypoints = []
                if waypoints_text:
                    for line in waypoints_text.splitlines():
                        line = line.strip()
                        if not line:
                            continue
                        parts = line.split(",")
                        if len(parts) == 2:
                            try:
                                lon, lat = float(parts[0]), float(parts[1])
                                waypoints.append((lon, lat))
                            except ValueError:
                                continue

                result = self._generate_route_editor_geometry(service, waypoints, inbound)
                if result.get("error"):
                    self.message_user(
                        request,
                        result["error"],
                        level=messages.ERROR,
                    )
                else:
                    self.message_user(
                        request,
                        (
                            "Generated snapped route geometry. "
                            f"Updated {result['updated']}, created {result['created']}, "
                            f"skipped {result['skipped']}."
                        ),
                        level=messages.SUCCESS,
                    )

            segments = busstops_views._route_editor_segments(service)
            existing_links = {
                (route_link.from_stop_id, route_link.to_stop_id): route_link
                for route_link in service.routelink_set.all()
            }
            updated = 0
            created = 0
            deleted = 0

            if action == "save_geometry":
                with transaction.atomic():
                    for segment in segments:
                        field_name = f"segment__{segment['from_stop_id']}__{segment['to_stop_id']}"
                        raw_value = (request.POST.get(field_name) or "").strip()
                        pair = (segment["from_stop_id"], segment["to_stop_id"])

                        if not raw_value:
                            route_link = existing_links.get(pair)
                            if route_link:
                                route_link.delete()
                                deleted += 1
                            continue

                        coordinates = []
                        for line in raw_value.splitlines():
                            line = line.strip()
                            if not line:
                                continue
                            parts = [part.strip() for part in line.split(",")]
                            if len(parts) != 2:
                                self.message_user(
                                    request,
                                    f"Invalid coordinate line for {segment['from_stop_name']} -> {segment['to_stop_name']}: {line}",
                                    level=messages.ERROR,
                                )
                                return TemplateResponse(
                                    request,
                                    "admin/busstops/service/route_editor.html",
                                    self.admin_site.each_context(request)
                                    | self._route_editor_context(query, service, "", "", inbound),
                                )
                            try:
                                coordinates.append((float(parts[0]), float(parts[1])))
                            except ValueError:
                                self.message_user(
                                    request,
                                    f"Invalid number for {segment['from_stop_name']} -> {segment['to_stop_name']}: {line}",
                                    level=messages.ERROR,
                                )
                                return TemplateResponse(
                                    request,
                                    "admin/busstops/service/route_editor.html",
                                    self.admin_site.each_context(request)
                                    | self._route_editor_context(query, service, "", "", inbound),
                                )

                        if len(coordinates) < 2:
                            self.message_user(
                                request,
                                f"{segment['from_stop_name']} -> {segment['to_stop_name']} needs at least two points.",
                                level=messages.ERROR,
                            )
                            return TemplateResponse(
                                request,
                                "admin/busstops/service/route_editor.html",
                                self.admin_site.each_context(request)
                                | self._route_editor_context(query, service),
                            )

                        geometry = LineString(coordinates, srid=4326)
                        route_link = existing_links.get(pair)
                        if route_link:
                            route_link.geometry = geometry
                            route_link.override = True
                            route_link.save(update_fields=["geometry", "override"])
                            updated += 1
                        else:
                            RouteLink.objects.create(
                                service=service,
                                from_stop_id=segment["from_stop_id"],
                                to_stop_id=segment["to_stop_id"],
                                geometry=geometry,
                                override=True,
                            )
                            created += 1

                self.message_user(
                    request,
                    f"Saved route geometry. Updated {updated}, created {created}, deleted {deleted}.",
                    level=messages.SUCCESS,
                )
                self._touch_service(service)

        return TemplateResponse(
            request,
            "admin/busstops/service/route_editor.html",
            self.admin_site.each_context(request)
            | self._route_editor_context(query, service, "", "", inbound),
        )

    def _generate_route_editor_geometry(self, service, waypoints=None, inbound=False):
        stop_usages = list(
            service.stopusage_set.filter(inbound=inbound, stop__latlong__isnull=False)
            .select_related("stop")
            .order_by("order", "id")
        )
        if len(stop_usages) < 2:
            return {"error": "Add at least two mapped stops before generating route geometry."}

        session = requests.Session()
        existing_links = {
            (route_link.from_stop_id, route_link.to_stop_id): route_link
            for route_link in service.routelink_set.all()
        }
        created = 0
        updated = 0
        skipped = 0

        with transaction.atomic():
            for from_usage, to_usage in pairwise(stop_usages):
                try:
                    line_substring = self._route_editor_segment_geometry(
                        session, from_usage, to_usage, waypoints
                    )
                except requests.RequestException as exc:
                    return {"error": f"Could not generate snapped geometry: {exc}"}
                except ValueError as exc:
                    return {"error": str(exc)}
                except (KeyError, IndexError, TypeError):
                    skipped += 1
                    continue

                from_point = ShapelyPoint(from_usage.stop.latlong.coords)
                to_point = ShapelyPoint(to_usage.stop.latlong.coords)
                if (
                    line_substring.geom_type != "LineString"
                    or len(line_substring.coords) < 2
                    or from_point.distance(
                        ShapelyPoint(line_substring.coords[0])
                    ) > 0.01
                    or to_point.distance(
                        ShapelyPoint(line_substring.coords[-1])
                    ) > 0.01
                ):
                    skipped += 1
                    continue

                pair = (from_usage.stop_id, to_usage.stop_id)
                route_link = existing_links.get(pair)
                if route_link:
                    route_link.geometry = line_substring.wkt
                    route_link.override = True
                    route_link.save(update_fields=["geometry", "override"])
                    updated += 1
                else:
                    RouteLink.objects.create(
                        service=service,
                        from_stop_id=from_usage.stop_id,
                        to_stop_id=to_usage.stop_id,
                        geometry=line_substring.wkt,
                        override=True,
                    )
                    created += 1

        service.update_geometry()
        self._touch_service(service)

        return {
            "created": created,
            "updated": updated,
            "skipped": skipped,
        }

    def _route_editor_segment_geometry(self, session, from_usage, to_usage, waypoints=None):
        router = getattr(settings, "ROUTE_EDITOR_ROUTER", "osrm")
        if router == "stadia":
            return self._route_editor_segment_geometry_stadia(
                session, from_usage, to_usage
            )
        return self._route_editor_segment_geometry_osrm(session, from_usage, to_usage, waypoints)

    def _route_editor_segment_geometry_osrm(self, session, from_usage, to_usage, waypoints=None):
        base_url = getattr(settings, "ROUTE_EDITOR_OSRM_URL", "").strip().rstrip("/")
        if not base_url:
            raise ValueError("ROUTE_EDITOR_OSRM_URL is not configured on this environment.")

        coordinates = (
            f"{from_usage.stop.latlong.x},{from_usage.stop.latlong.y};"
        )
        
        # Add waypoints if provided (for routing guidance only)
        if waypoints:
            for waypoint in waypoints:
                coordinates += f"{waypoint[0]},{waypoint[1]};"
        
        coordinates += f"{to_usage.stop.latlong.x},{to_usage.stop.latlong.y}"
        
        response = session.get(
            f"{base_url}/route/v1/driving/{coordinates}",
            params={
                "overview": "full",
                "geometries": "geojson",
                "steps": "false",
                "continue_straight": "true",
            },
            timeout=30,
        )
        response.raise_for_status()
        payload = response.json()

        route = (payload.get("routes") or [None])[0]
        geometry = (route or {}).get("geometry") or {}
        route_coordinates = geometry.get("coordinates") or []
        if len(route_coordinates) < 2:
            raise ValueError("Local router did not return a usable route geometry.")

        return ShapelyLineString(route_coordinates)

    def _route_editor_segment_geometry_stadia(self, session, from_usage, to_usage):
        api_key = getattr(settings, "STADIA_MAPS_API_KEY", "")
        if not api_key:
            raise ValueError("STADIA_MAPS_API_KEY is not configured on this environment.")

        import polyline

        response = session.post(
            "https://api.stadiamaps.com/route/v1",
            params={"api_key": api_key},
            json={
                "locations": [
                    {
                        "lat": from_usage.stop.latlong.y,
                        "lon": from_usage.stop.latlong.x,
                        "type": "break",
                    },
                    {
                        "lat": to_usage.stop.latlong.y,
                        "lon": to_usage.stop.latlong.x,
                        "type": "break",
                    },
                ],
                "costing": "bus",
            },
            timeout=30,
        )
        response.raise_for_status()
        payload = response.json()
        leg = payload["trip"]["legs"][0]
        return ShapelyLineString(
            [(lon, lat) for lat, lon in polyline.decode(leg["shape"], precision=6)]
        )

    def _route_editor_context(self, query, service, stop_codes_override=None, waypoints_text="", inbound=False):
        results = []
        if query:
            results = list(
                models.Service.objects.with_line_names()
                .filter(
                    Q(line_name__icontains=query)
                    | Q(description__icontains=query)
                    | Q(service_code__icontains=query)
                    | Q(line_brand__icontains=query)
                    | Q(operator__name__icontains=query)
                )
                .distinct()
                .order_by("line_name", "description")[:25]
            )

        segments = []
        stop_codes_text = stop_codes_override
        selected_stop_rows = []
        preview_stops = []
        if service:
            if stop_codes_text is None:
                stop_usages = list(
                    service.stopusage_set.filter(inbound=inbound)
                    .select_related("stop", "stop__locality")
                    .order_by("order", "id")
                )
                stop_codes_text = "\n".join(stop_usage.stop_id for stop_usage in stop_usages)
                selected_stop_rows = [
                    {
                        "stop_id": stop_usage.stop_id,
                        "stop_name": (
                            stop_usage.stop.get_qualified_name()
                            if stop_usage.stop_id and stop_usage.stop
                            else stop_usage.stop_id
                        ),
                    }
                    for stop_usage in stop_usages
                ]
                preview_stops = [
                    {
                        "stop_id": stop_usage.stop_id,
                        "stop_name": (
                            stop_usage.stop.get_qualified_name()
                            if stop_usage.stop_id and stop_usage.stop
                            else stop_usage.stop_id
                        ),
                        "coordinates": list(stop_usage.stop.latlong.coords)
                        if stop_usage.stop and stop_usage.stop.latlong
                        else None,
                    }
                    for stop_usage in stop_usages
                ]
            elif stop_codes_text:
                stop_ids = [
                    item.strip()
                    for item in re.split(r"[\s,]+", stop_codes_text)
                    if item.strip()
                ]
                stops = models.StopPoint.objects.filter(atco_code__in=stop_ids).select_related(
                    "locality"
                )
                stops_by_id = {stop.atco_code: stop for stop in stops}
                selected_stop_rows = [
                    {
                        "stop_id": stop_id,
                        "stop_name": (
                            stops_by_id[stop_id].get_qualified_name()
                            if stop_id in stops_by_id
                            else stop_id
                        ),
                    }
                    for stop_id in stop_ids
                ]
                preview_stops = [
                    {
                        "stop_id": stop_id,
                        "stop_name": (
                            stops_by_id[stop_id].get_qualified_name()
                            if stop_id in stops_by_id
                            else stop_id
                        ),
                        "coordinates": list(stops_by_id[stop_id].latlong.coords)
                        if stop_id in stops_by_id and stops_by_id[stop_id].latlong
                        else None,
                    }
                    for stop_id in stop_ids
                ]
            for segment in busstops_views._route_editor_segments(service):
                segment = segment.copy()
                segment["field_name"] = (
                    f"segment__{segment['from_stop_id']}__{segment['to_stop_id']}"
                )
                segment["coordinates_text"] = "\n".join(
                    f"{lng:.6f}, {lat:.6f}" for lng, lat in segment["coordinates"]
                )
                segments.append(segment)

        return {
            "title": "Route editor",
            "search_query": query,
            "results": results,
            "service": service,
            "segments": segments,
            "stop_codes_text": stop_codes_text or "",
            "selected_stop_ids_csv": ",".join(
                row["stop_id"] for row in selected_stop_rows if row.get("stop_id")
            ),
            "selected_stop_rows": selected_stop_rows,
            "preview_stops": preview_stops,
            "waypoints_text": waypoints_text,
        }

    def get_queryset(self, request):
        queryset = super().get_queryset(request)
        if "changelist" in request.resolver_match.view_name:
            queryset = queryset.annotate(routes=SubqueryCount("route"))

            queryset = queryset.annotate(
                service_codes=StringAgg("route__service_code", Value(" "))
            )

        return queryset

    def get_search_results(self, request, queryset, search_term):
        if search_term and request.path.endswith("/autocomplete/"):
            queryset = queryset.filter(current=True)

            query = SearchQuery(search_term, search_type="websearch", config="english")
            rank = SearchRank(F("search_vector"), query)
            queryset = (
                queryset.annotate(rank=rank)
                .filter(Q(search_vector=query) | Q(service_code=search_term))
                .order_by("-rank")
            )
            return queryset, False

        return super().get_search_results(request, queryset, search_term)

    def current_false(self, request, queryset):
        result = queryset.order_by().update(current=False)
        log_change(request, queryset, ["current"])
        self.message_user(request, f"{result}")

    def current_true(self, request, queryset):
        result = queryset.order_by().update(current=True)
        log_change(request, queryset, ["current"])
        self.message_user(request, f"{result}")

    def public_use_true(self, request, queryset):
        result = queryset.order_by().update(public_use=True)
        log_change(request, queryset, ["public_use"])
        self.message_user(request, f"{result}")

    @transaction.atomic
    def merge(self, request, queryset):
        first = queryset[0]
        others = queryset[1:]

        first.current = True

        for other in others:
            if other.current:
                other.route_set.update(service=first)
                first.operator.add(*other.operator.all())
            other.vehiclejourney_set.update(service=first)
            other.servicecode_set.filter(
                ~Exists(
                    models.ServiceCode.objects.filter(
                        scheme=OuterRef("scheme"), code=OuterRef("code"), service=first
                    )
                )
            ).update(service=first)

            other.routelink_set.filter(
                ~Exists(
                    RouteLink.objects.filter(
                        from_stop=OuterRef("from_stop"),
                        to_stop=OuterRef("to_stop"),
                        service=first,
                    )
                )
            ).update(service=first)

            if not first.servicecode_set.filter(
                code=other.slug, scheme="slug"
            ).exists():
                models.ServiceCode.objects.create(
                    code=other.slug, service=first, scheme="slug"
                )

            if (
                other.service_code
                and other.service_code != first.service_code
                and not first.servicecode_set.filter(
                    code=other.service_code, scheme="ServiceCode"
                ).exists()
            ):
                models.ServiceCode.objects.create(
                    code=other.service_code, service=first, scheme="ServiceCode"
                )

            other.delete()

        first.do_stop_usages()
        first.update_geometry()
        first.save(force_update=True)

        first.update_description()
        first.update_search_vector()

        self.message_user(request, f"merged {others} into {first}")

    def unmerge(self, request, queryset):
        for service in queryset:
            with transaction.atomic():
                services_by_line_name = {service.line_name: service.id}
                service_id = service.id  # for use later
                operators = service.operator.all()
                routes = service.route_set.all()
                journeys = service.vehiclejourney_set.all()
                bool(journeys)  # force evaluation
                service_codes = service.servicecode_set.all()
                bool(service_codes)  # force evaluation
                for route in routes:
                    if route.line_name not in services_by_line_name:
                        service.id = None
                        service.line_name = route.line_name
                        service.description = route.description
                        service.search_vector = None
                        service.slug = ""
                        service.save(force_insert=True)
                        service.operator.set(operators)
                        services_by_line_name[route.line_name] = service.id
                    route.service_id = services_by_line_name[route.line_name]
                    route.save(update_fields=["service_id"])

                for service in models.Service.objects.filter(
                    id__in=services_by_line_name.values()
                ):
                    service.do_stop_usages()
                    service.update_geometry()
                    service.update_search_vector()
                    if service.id != service_id:
                        journeys.filter(
                            Q(trip__route__service=service)
                            | Q(route_name__iexact=service.line_name)
                        ).update(service=service)
                        service_codes.filter(
                            code__istartswith=f"{service.line_name}-"
                        ).update(service=service)


@admin.register(models.ServiceLink)
class ServiceLinkAdmin(admin.ModelAdmin):
    save_as = True
    list_display = (
        "from_service",
        "from_service__current",
        "to_service",
        "to_service__current",
        "how",
    )
    list_filter = (
        "from_service__current",
        "to_service__current",
        "from_service__source",
        "to_service__source",
    )
    autocomplete_fields = ("from_service", "to_service")

    @staticmethod
    def from_service__current(obj):
        return obj.from_service.current

    @staticmethod
    def to_service__current(obj):
        return obj.to_service.current


@admin.register(models.Locality)
class LocalityAdmin(GISModelAdmin):
    list_display = ("id", "name", "slug", "modified_at", "created_at")
    search_fields = ("id", "name")
    raw_id_fields = ("adjacent", "parent")
    list_filter = ("modified_at", "created_at", "admin_area__region", "admin_area")
    readonly_fields = ["search_vector"]


@admin.register(models.OperatorCode)
class OperatorCodeAdmin(admin.ModelAdmin):
    save_as = True
    list_display = ("id", "operator", "source", "code")
    list_filter = [("source", admin.RelatedOnlyFieldListFilter)]
    search_fields = ("code",)
    raw_id_fields = ("operator",)


@admin.register(models.ServiceCode)
class ServiceCodeAdmin(admin.ModelAdmin):
    list_display = ["id", "service", "scheme", "code"]
    list_filter = [
        "scheme",
        "service__current",
        ("service__operator", admin.RelatedOnlyFieldListFilter),
        "service__stops__admin_area",
    ]
    search_fields = ["code", "service__line_name", "service__description"]
    autocomplete_fields = ["service"]


@admin.register(models.ServiceColour)
class ServiceColourAdmin(admin.ModelAdmin):
    list_display = ["preview", "foreground", "background", "services"]
    search_fields = ["name"]
    list_filter = [("service__operator", admin.EmptyFieldListFilter)]

    def get_queryset(self, request):
        queryset = super().get_queryset(request)
        if "changelist" in request.resolver_match.view_name:
            queryset = queryset.annotate(
                services=SubqueryCount("service", filter=Q(current=True))
            )
        return queryset

    @admin.display(ordering="services")
    def services(self, obj):
        return obj.services


@admin.register(models.DataSource)
class DataSourceAdmin(admin.ModelAdmin):
    search_fields = ("name", "url")
    list_display = (
        "name",
        "description",
        "url",
        "sha1",
        "datetime",
        "settings",
        "routes",
        "services",
        "source",
        # "journeys",
    )
    list_filter = (
        ("route", admin.EmptyFieldListFilter),
        ("service", admin.EmptyFieldListFilter),
        ("vehiclejourney", admin.EmptyFieldListFilter),
    )
    actions = ["delete_routes", "remove_datetimes"]
    show_full_result_count = False

    def get_queryset(self, request):
        queryset = super().get_queryset(request)
        if "changelist" in request.resolver_match.view_name:
            queryset = queryset.annotate(
                routes=SubqueryCount("route", filter=~Q(service=None)),
                services=SubqueryCount("service", filter=Q(current=True)),
                # journeys=Exists(VehicleJourney.objects.filter(source=OuterRef("id"))),
            ).prefetch_related("operatorcode_set")
        return queryset

    @admin.display(ordering="routes")
    def routes(self, obj):
        url = reverse("admin:bustimes_route_changelist")
        return format_html(
            '<a href="{}?source__id__exact={}">{}</a>', url, obj.id, obj.routes
        )

    @admin.display(ordering="services")
    def services(self, obj):
        url = reverse("admin:busstops_service_changelist")
        return format_html(
            '<a href="{}?source__id__exact={}">{}</a>', url, obj.id, obj.services
        )

    # @admin.display(ordering="journeys")
    # def journeys(self, obj):
    #     url = reverse("admin:vehicles_vehiclejourney_changelist")
    #     return format_html(
    #         '<a href="{}?source__id__exact={}">{}</a>', url, obj.id, obj.journeys
    #     )

    def delete_routes(self, request, queryset):
        result = Route.objects.filter(source__in=queryset).update(service=None)
        self.message_user(request, result)

    def remove_datetimes(self, request, queryset):
        result = queryset.order_by().update(datetime=None, sha1="")
        log_change(request, queryset, ["datetime", "sha1"])
        self.message_user(request, result)


@admin.register(models.SIRISource)
class SIRISourceAdmin(admin.ModelAdmin):
    list_display = ("name", "url", "requestor_ref", "areas", "is_poorly")
    autocomplete_fields = ("operators", "admin_areas")

    def get_queryset(self, request):
        queryset = super().get_queryset(request)
        if "changelist" in request.resolver_match.view_name:
            queryset = queryset.annotate(
                areas=StringAgg(
                    Cast("admin_areas__atco_code", output_field=CharField()),
                    Value(", "),
                )
            )
        return queryset

    @staticmethod
    def areas(obj):
        return obj.areas


class PaymentMethodOperatorInline(admin.TabularInline):
    model = models.PaymentMethod.operator_set.through
    autocomplete_fields = ["operator"]


class PaymentMethodServiceInline(admin.TabularInline):
    model = models.PaymentMethod.service_set.through
    autocomplete_fields = ["service"]



@admin.register(models.HomepageNotice)
class HomepageNoticeAdmin(admin.ModelAdmin):
    list_display = ("title", "from_date", "to_date", "is_active", "modified_at")
    list_filter = ("is_active", "from_date", "to_date")
    search_fields = ("title", "message")

    def formfield_for_dbfield(self, db_field, request, **kwargs):
        formfield = super().formfield_for_dbfield(db_field, request, **kwargs)
        if db_field.name == "message":
            formfield.widget.attrs.setdefault("rows", 6)
        return formfield


@admin.register(models.RouteNotice)
class RouteNoticeAdmin(admin.ModelAdmin):
    list_display = (
        "title",
        "service",
        "start",
        "end",
        "planned",
        "diversion",
        "diversion_num",
        "route_map_id",
    )
    list_filter = ("planned", "diversion", "start", "end")
    search_fields = ("title", "description", "service__line_name", "service__description")
    autocomplete_fields = ("service", "other_services")
    readonly_fields = ("route_map_id",)

    def formfield_for_dbfield(self, db_field, request, **kwargs):
        formfield = super().formfield_for_dbfield(db_field, request, **kwargs)
        if db_field.name == "description":
            formfield.widget.attrs.setdefault("rows", 6)
        return formfield


@admin.register(models.PaymentMethod)
class PaymentMethodAdmin(admin.ModelAdmin):
    list_display = ("name", "url", "operators")
    search_fields = ("name", "url")
    inlines = [PaymentMethodOperatorInline, PaymentMethodServiceInline]

    def get_queryset(self, request):
        queryset = super().get_queryset(request)
        if "changelist" in request.resolver_match.view_name:
            queryset = queryset.annotate(
                operators=StringAgg("operator", Value(", "), distinct=True)
            )
        return queryset

    @admin.display(description="Operators")
    def operators(self, obj):
        return ", ".join(str(operator) for operator in obj.operator.all())


admin.site.register(models.Region)
admin.site.register(models.District)












