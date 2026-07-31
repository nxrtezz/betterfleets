from io import BytesIO
from unittest.mock import patch

from django.contrib.admin.helpers import ACTION_CHECKBOX_NAME
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase

from openpyxl import load_workbook
from openpyxl.workbook import Workbook

from accounts.models import User
from busstops.models import Operator, OperatorGroup, Organisation, Service, ServiceColour
from bustimes.models import Garage
from fleet.parsers.pdf_fleet_parser import ParsedFleetRecord
from vehicles.models import Livery, Vehicle


class OperatorAdminMassAddTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.operator = Operator.objects.create(noc="MASS", name="Mass Operator", slug="mass-operator")
        cls.existing_vehicle = Vehicle.objects.create(
            operator=cls.operator,
            code="100",
            reg="AB12CDE",
            name="Old Name",
        )
        cls.existing_service = Service.objects.create(
            service_code="MASS-1",
            line_name="1",
            description="Town - Station",
            current=True,
        )
        cls.existing_service.operator.add(cls.operator)
        cls.blue = ServiceColour.objects.create(name="Blue", background="#0000ff")
        cls.superuser = User.objects.create_superuser(
            username="root",
            email="root@example.com",
            password="pass",
        )
        cls.staff_user = User.objects.create_user(
            username="staff",
            email="staff@example.com",
            password="pass",
            is_staff=True,
        )

    def mass_add_url(self):
        return f"/admin/busstops/operator/{self.operator.pk}/mass-add-buses/"

    def mass_add_template_url(self):
        return f"/admin/busstops/operator/{self.operator.pk}/mass-add-buses/template.xlsx"

    def mass_edit_url(self):
        return f"/admin/busstops/operator/{self.operator.pk}/mass-edit-buses/"

    def mass_edit_template_url(self):
        return f"/admin/busstops/operator/{self.operator.pk}/mass-edit-buses/template.xlsx"

    def mass_add_routes_url(self):
        return f"/admin/busstops/operator/{self.operator.pk}/mass-add-routes/"

    def mass_add_routes_template_url(self):
        return f"/admin/busstops/operator/{self.operator.pk}/mass-add-routes/template.xlsx"

    def current_routes_url(self):
        return f"/admin/busstops/operator/{self.operator.pk}/mass-add-routes/current-routes.xlsx"

    def new_historical_fleet_url(self):
        return f"/admin/busstops/operator/{self.operator.pk}/new-historical-fleet/"
    def current_fleet_url(self):
        return f"/admin/busstops/operator/{self.operator.pk}/mass-add-buses/current-fleet.xlsx"

    def fleet_import_url(self):
        return "/fleet/import"

    def admin_vehicle_mass_import_url(self):
        return "/admin/vehicles/vehicle/mass-import/"

    def make_workbook_upload(self, rows):
        workbook = Workbook()
        worksheet = workbook.active
        for row in rows:
            worksheet.append(row)
        output = BytesIO()
        workbook.save(output)
        return SimpleUploadedFile(
            "mass-add-template.xlsx",
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_superuser_can_access_mass_add_page(self):
        self.client.force_login(self.superuser)
        response = self.client.get(self.mass_add_url())
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Mass add buses")
        self.assertContains(response, "Download XLSX template")

    def test_superuser_can_download_mass_add_template(self):
        self.client.force_login(self.superuser)
        response = self.client.get(self.mass_add_template_url())
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(workbook.active.title, "Vehicles")
        headers = [cell.value for cell in workbook.active[1]]
        self.assertEqual(
            headers,
            [
                "operator_code",
                "external_id",
                "code",
                "fleet_number",
                "fleet_code",
                "registration",
                "prev_registration",
                "vehicle_type",
                "livery",
                "colours",
                "garage",
                "name",
                "branding",
                "notes",
                "withdrawn",
                "preserved",
                "fleet_support_vehicle",
                "vor",
                "awaiting_delivery",
                "trainer_vehicle",
                "demonstrator",
                "features",
            ],
        )
    def test_superuser_can_download_current_fleet_workbook(self):
        self.client.force_login(self.superuser)
        response = self.client.get(self.current_fleet_url())
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        headers = [cell.value for cell in workbook.active[1]]
        self.assertEqual(headers[0], "operator_code")
        data_row = [cell.value for cell in workbook.active[2]]
        self.assertEqual(data_row[0], self.operator.noc)
        self.assertEqual(data_row[2], self.existing_vehicle.code)
        self.assertEqual(data_row[9], "")

    def test_superuser_can_access_mass_add_routes_page(self):
        self.client.force_login(self.superuser)
        response = self.client.get(self.mass_add_routes_url())
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Mass add routes")
        self.assertContains(response, "Download XLSX template")

    def test_superuser_can_download_mass_add_routes_template(self):
        self.client.force_login(self.superuser)
        response = self.client.get(self.mass_add_routes_template_url())
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(workbook.active.title, "Routes")
        headers = [cell.value for cell in workbook.active[1]]
        self.assertEqual(
            headers,
            [
                "operator_code",
                "service_id",
                "service_code",
                "line_name",
                "line_brand",
                "description",
                "mode",
                "current",
                "non_current_route",
                "timetable_wrong",
                "tracking",
                "public_use",
                "colour",
            ],
        )

    def test_superuser_can_download_current_routes_workbook(self):
        self.client.force_login(self.superuser)
        response = self.client.get(self.current_routes_url())
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        headers = [cell.value for cell in workbook.active[1]]
        self.assertEqual(headers[0], "operator_code")
        data_row = [cell.value for cell in workbook.active[2]]
        self.assertEqual(data_row[0], self.operator.noc)
        self.assertEqual(data_row[1], self.existing_service.pk)
        self.assertEqual(data_row[2], self.existing_service.service_code)
        self.assertEqual(data_row[3], self.existing_service.line_name)

    def test_superuser_can_access_mass_edit_page(self):
        self.client.force_login(self.superuser)
        response = self.client.get(self.mass_edit_url())
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Mass edit buses")
        self.assertContains(
            response,
            "Mass edit only updates vehicles that already exist for this operator.",
        )

    def test_superuser_can_access_new_historical_fleet_page(self):
        self.client.force_login(self.superuser)
        response = self.client.get(self.new_historical_fleet_url())
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "New historical fleet")
        self.assertContains(response, "Year this historical fleet snapshot is from")

    def test_superuser_can_download_mass_edit_template(self):
        self.client.force_login(self.superuser)
        response = self.client.get(self.mass_edit_template_url())
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(workbook.active.title, "Vehicles")

    def test_non_superuser_cannot_access_mass_add_page(self):
        self.client.force_login(self.staff_user)
        response = self.client.get(self.mass_add_url())
        self.assertEqual(response.status_code, 403)

    def test_non_superuser_cannot_access_mass_edit_page(self):
        self.client.force_login(self.staff_user)
        response = self.client.get(self.mass_edit_url())
        self.assertEqual(response.status_code, 403)

    def test_non_superuser_cannot_access_mass_add_routes_page(self):
        self.client.force_login(self.staff_user)
        response = self.client.get(self.mass_add_routes_url())
        self.assertEqual(response.status_code, 403)

    def test_mass_add_commit_creates_updates_and_keeps_row_errors_isolated(self):
        self.client.force_login(self.superuser)

        payload = {
            "rows_text": "\n".join(
                [
                    "code,registration,name,notes,withdrawn",
                    "100,AB12CDE,Updated Bus,Updated notes,false",
                    "200,CD34EFG,New Bus,Fresh,true",
                    "300,EF56HIJ,Bad Bus,,maybe",
                ]
            ),
            "action": "commit",
        }

        response = self.client.post(self.mass_add_url(), payload)
        self.assertEqual(response.status_code, 200)

        self.existing_vehicle.refresh_from_db()
        self.assertEqual(self.existing_vehicle.name, "Updated Bus")
        self.assertEqual(self.existing_vehicle.notes, "Updated notes")
        self.assertTrue(self.existing_vehicle.is_manual)

        created = Vehicle.objects.get(operator=self.operator, code="200")
        self.assertEqual(created.reg, "CD34EFG")
        self.assertEqual(created.name, "New Bus")
        self.assertTrue(created.withdrawn)
        self.assertTrue(created.is_manual)

        self.assertFalse(Vehicle.objects.filter(operator=self.operator, code="300").exists())

    def test_mass_add_routes_commit_creates_updates_and_keeps_row_errors_isolated(self):
        self.client.force_login(self.superuser)

        payload = {
            "rows_text": "\n".join(
                [
                    "service_code,line_name,description,current,public_use,colour",
                    "MASS-1,1 Updated,Town - Station via Centre,false,true,Blue",
                    "MASS-2,2,Town - Hospital,true,false,Blue",
                    ",,,,maybe,Unknown",
                ]
            ),
            "action": "commit",
        }

        response = self.client.post(self.mass_add_routes_url(), payload)
        self.assertEqual(response.status_code, 200)

        self.existing_service.refresh_from_db()
        self.assertEqual(self.existing_service.line_name, "1 Updated")
        self.assertEqual(self.existing_service.description, "Town - Station via Centre")
        self.assertFalse(self.existing_service.current)
        self.assertTrue(self.existing_service.public_use)
        self.assertEqual(self.existing_service.colour, self.blue)

        created = Service.objects.get(service_code="MASS-2")
        self.assertEqual(created.line_name, "2")
        self.assertEqual(created.description, "Town - Hospital")
        self.assertFalse(created.public_use)
        self.assertEqual(created.colour, self.blue)
        self.assertTrue(created.operator.filter(pk=self.operator.pk).exists())

        self.assertEqual(Service.objects.filter(service_code="").count(), 0)

    def test_mass_add_preview_accepts_uploaded_workbook(self):
        self.client.force_login(self.superuser)
        workbook = self.make_workbook_upload(
            [
                ["code", "registration", "name", "withdrawn"],
                ["250", "XY12ZZZ", "Workbook Bus", "false"],
            ]
        )

        response = self.client.post(
            self.mass_add_url(),
            {"rows_text": "", "action": "preview", "workbook": workbook},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Preview generated")
        self.assertContains(response, "250")
        self.assertContains(response, "create")
        self.assertContains(response, "XY12ZZZ")
        self.assertFalse(Vehicle.objects.filter(operator=self.operator, code="250").exists())

    def test_mass_add_routes_preview_accepts_uploaded_workbook(self):
        self.client.force_login(self.superuser)
        workbook = self.make_workbook_upload(
            [
                ["service_code", "line_name", "description", "current"],
                ["MASS-3", "3", "Town - College", "true"],
            ]
        )

        response = self.client.post(
            self.mass_add_routes_url(),
            {"rows_text": "", "action": "preview", "workbook": workbook},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Preview generated")
        self.assertContains(response, "MASS-3")
        self.assertContains(response, "create")
        self.assertFalse(Service.objects.filter(service_code="MASS-3").exists())
    def test_mass_add_preserved_operator_does_not_update_other_operators_by_external_id(self):
        preserved = Operator.objects.create(
            noc="PRES",
            name="Preserved Operator",
            slug="preserved-operator",
            preserved=True,
        )
        other = Operator.objects.create(noc="OTHR", name="Other Operator", slug="other-operator")
        other_vehicle = Vehicle.objects.create(
            operator=other,
            code="X1",
            reg="AA11AAA",
            external_id="shared-external",
            name="Other Fleet Bus",
        )

        self.client.force_login(self.superuser)
        response = self.client.post(
            f"/admin/busstops/operator/{preserved.pk}/mass-add-buses/",
            {
                "rows_text": "\n".join(
                    [
                        "external_id,code,registration,name,withdrawn",
                        "shared-external,999,BB22BBB,Preserved Bus,false",
                    ]
                ),
                "action": "commit",
            },
        )
        self.assertEqual(response.status_code, 200)

        other_vehicle.refresh_from_db()
        self.assertEqual(other_vehicle.operator_id, other.pk)
        self.assertEqual(other_vehicle.code, "X1")
        self.assertEqual(other_vehicle.name, "Other Fleet Bus")

        created = Vehicle.objects.get(operator=preserved, code="999")
        self.assertEqual(created.external_id, "shared-external-PRES")
        self.assertEqual(created.reg, "BB22BBB")
        self.assertEqual(created.name, "Preserved Bus")

    def test_mass_add_commit_after_workbook_preview_uses_hidden_rows_text(self):
        self.client.force_login(self.superuser)
        workbook = self.make_workbook_upload(
            [
                ["code", "registration", "name", "withdrawn"],
                ["251", "XY13ZZZ", "Preview Then Commit", "false"],
            ]
        )

        preview_response = self.client.post(
            self.mass_add_url(),
            {"rows_text": "", "action": "preview", "workbook": workbook},
        )
        self.assertEqual(preview_response.status_code, 200)
        self.assertContains(preview_response, "251")

        commit_response = self.client.post(
            self.mass_add_url(),
            {
                "rows_text": (
                    "code\tregistration\tname\twithdrawn\n"
                    "251\tXY13ZZZ\tPreview Then Commit\tfalse\n"
                ),
                "action": "commit",
            },
        )
        self.assertEqual(commit_response.status_code, 200)
        created = Vehicle.objects.get(operator=self.operator, code="251")
        self.assertEqual(created.reg, "XY13ZZZ")
        self.assertEqual(created.name, "Preview Then Commit")
    def test_mass_add_commit_accepts_uploaded_workbook(self):
        self.client.force_login(self.superuser)
        workbook = self.make_workbook_upload(
            [
                ["code", "registration", "name", "withdrawn"],
                ["260", "ZZ12YYY", "Uploaded Bus", "true"],
            ]
        )

        response = self.client.post(
            self.mass_add_url(),
            {"rows_text": "", "action": "commit", "workbook": workbook},
        )
        self.assertEqual(response.status_code, 200)

        created = Vehicle.objects.get(operator=self.operator, code="260")
        self.assertEqual(created.reg, "ZZ12YYY")
        self.assertEqual(created.name, "Uploaded Bus")
        self.assertTrue(created.withdrawn)
        self.assertTrue(created.is_manual)

    def test_mass_add_can_target_another_operator_with_operator_code(self):
        self.client.force_login(self.superuser)
        other_operator = Operator.objects.create(
            noc="OT01",
            name="Other Mass Operator",
            slug="other-mass-operator",
        )
        response = self.client.post(
            self.mass_add_url(),
            {
                "rows_text": "\n".join(
                    [
                        "operator_code,code,registration,name,withdrawn",
                        "OT01,777,QQ12WWW,From Multi Operator Import,false",
                    ]
                ),
                "action": "commit",
            },
        )
        self.assertEqual(response.status_code, 200)
        created = Vehicle.objects.get(operator=other_operator, code="777")
        self.assertEqual(created.reg, "QQ12WWW")
        self.assertEqual(created.name, "From Multi Operator Import")

    def test_new_historical_fleet_commit_creates_year_scoped_historical_vehicle(self):
        self.client.force_login(self.superuser)
        response = self.client.post(
            self.new_historical_fleet_url(),
            {
                "historical_year": 2005,
                "rows_text": "\n".join(
                    [
                        "code,registration,name",
                        "700,AA55AAA,History Bus",
                    ]
                ),
                "action": "commit",
            },
        )
        self.assertEqual(response.status_code, 200)

        created = Vehicle.objects.get(operator=self.operator, code="700-2005")
        self.assertEqual(created.historical_fleet_id, self.operator.pk)
        self.assertEqual(created.historical_fleet_year, 2005)
        self.assertFalse(created.preserved)
        self.assertIn("-2005", created.slug)
        self.assertEqual(created.code, "700-2005")
        self.assertEqual(created.fleet_code, "700")

    def test_new_historical_fleet_reimport_updates_same_year_record(self):
        self.client.force_login(self.superuser)
        payload = {
            "historical_year": 2001,
            "rows_text": "\n".join(
                [
                    "external_id,code,registration,name",
                    "history-701,701,BB51BBB,Historic Fleet Bus",
                ]
            ),
            "action": "commit",
        }
        response = self.client.post(self.new_historical_fleet_url(), payload)
        self.assertEqual(response.status_code, 200)

        response = self.client.post(
            self.new_historical_fleet_url(),
            {
                "historical_year": 2001,
                "rows_text": "\n".join(
                    [
                        "external_id,code,registration,name",
                        "history-701,701,BB51BBB,Historic Fleet Bus Updated",
                    ]
                ),
                "action": "commit",
            },
        )
        self.assertEqual(response.status_code, 200)

        created = Vehicle.objects.get(operator=self.operator, code="701-2001")
        self.assertEqual(created.name, "Historic Fleet Bus Updated")
        self.assertEqual(
            Vehicle.objects.filter(
                operator=self.operator,
                historical_fleet=self.operator,
                historical_fleet_year=2001,
                code="701-2001",
            ).count(),
            1,
        )

    def test_new_historical_fleet_can_coexist_with_live_code(self):
        self.client.force_login(self.superuser)
        live_vehicle = Vehicle.objects.create(
            operator=self.operator,
            code="10001",
            fleet_code="10001",
            reg="AA11AAA",
        )

        response = self.client.post(
            self.new_historical_fleet_url(),
            {
                "historical_year": 2022,
                "rows_text": "\n".join(
                    [
                        "code,registration,name",
                        "10001,BB22BBB,Historic Clash Bus",
                    ]
                ),
                "action": "commit",
            },
        )
        self.assertEqual(response.status_code, 200)

        live_vehicle.refresh_from_db()
        self.assertEqual(live_vehicle.code, "10001")

        created = Vehicle.objects.get(
            operator=self.operator,
            historical_fleet=self.operator,
            historical_fleet_year=2022,
            code="10001-2022",
        )
        self.assertEqual(created.fleet_code, "10001")
        self.assertEqual(created.name, "Historic Clash Bus")
        self.assertFalse(created.preserved)

    def test_mass_add_withdrawn_vehicle_is_attached_to_historical_fleet(self):
        self.client.force_login(self.superuser)
        response = self.client.post(
            self.mass_add_url(),
            {
                "rows_text": "\n".join(
                    [
                        "code,registration,name,withdrawn",
                        "880,CC22CCC,Withdrawn Bus,true",
                    ]
                ),
                "action": "commit",
            },
        )
        self.assertEqual(response.status_code, 200)

        created = Vehicle.objects.get(operator=self.operator, code="880")
        self.assertTrue(created.withdrawn)
        self.assertEqual(created.historical_fleet_id, self.operator.pk)
        self.assertIsNone(created.historical_fleet_year)

    def test_mass_edit_blank_cells_do_not_override_existing_values(self):
        self.client.force_login(self.superuser)
        garage = Garage.objects.create(operator=self.operator, code="GAR1", name="Garage 1")
        self.existing_vehicle.garage = garage
        self.existing_vehicle.name = "Existing Name"
        self.existing_vehicle.branding = "Existing Brand"
        self.existing_vehicle.notes = "Keep Me"
        self.existing_vehicle.save()

        response = self.client.post(
            self.mass_edit_url(),
            {
                "rows_text": "\n".join(
                    [
                        "code,garage,name,branding,notes,withdrawn",
                        "100,,,,,true",
                    ]
                ),
                "action": "commit",
            },
        )
        self.assertEqual(response.status_code, 200)

        self.existing_vehicle.refresh_from_db()
        self.assertEqual(self.existing_vehicle.garage, garage)
        self.assertEqual(self.existing_vehicle.name, "Existing Name")
        self.assertEqual(self.existing_vehicle.branding, "Existing Brand")
        self.assertEqual(self.existing_vehicle.notes, "Keep Me")
        self.assertTrue(self.existing_vehicle.withdrawn)

    def test_mass_add_preview_accepts_pdf_and_shows_operator_and_depot_intentions(self):
        self.client.force_login(self.superuser)
        pdf_upload = SimpleUploadedFile(
            "fleet.pdf",
            b"%PDF-1.4 fake pdf",
            content_type="application/pdf",
        )

        with patch("busstops.admin.parse_pdf") as mocked_parse_pdf:
            mocked_parse_pdf.return_value = [
                ParsedFleetRecord(
                    operator_code="OT99",
                    code="901",
                    fleet_number="901",
                    fleet_code="901",
                    registration="HF17AZA",
                    vehicle_type="Volvo B11R Plaxton Elite",
                    livery="Damory",
                    garage="GSC Pimperne",
                )
            ]
            response = self.client.post(
                self.mass_add_url(),
                {"rows_text": "", "action": "preview", "workbook": pdf_upload},
            )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Create operator OT99")
        self.assertContains(response, "Create depot GSC Pimperne for new operator")
        self.assertContains(response, "901")

    def test_mass_add_commit_from_pdf_creates_operator_depot_and_vehicle(self):
        self.client.force_login(self.superuser)
        pdf_upload = SimpleUploadedFile(
            "fleet.pdf",
            b"%PDF-1.4 fake pdf",
            content_type="application/pdf",
        )

        with patch("busstops.admin.parse_pdf") as mocked_parse_pdf:
            mocked_parse_pdf.return_value = [
                ParsedFleetRecord(
                    operator_code="OT98",
                    code="902",
                    fleet_number="902",
                    fleet_code="902",
                    registration="HF17AZB",
                    vehicle_type="Volvo B11R Plaxton Elite",
                    livery="Damory",
                    garage="GSC Pimperne",
                    name="Imported PDF Bus",
                )
            ]
            response = self.client.post(
                self.mass_add_url(),
                {"rows_text": "", "action": "commit", "workbook": pdf_upload},
            )

        self.assertEqual(response.status_code, 200)
        created_operator = Operator.objects.get(noc="OT98")
        created_garage = Garage.objects.get(operator=created_operator, name="Pimperne")
        created_vehicle = Vehicle.objects.get(operator=created_operator, code="902")
        self.assertEqual(created_vehicle.reg, "HF17AZB")
        self.assertEqual(created_vehicle.garage, created_garage)
        self.assertEqual(created_vehicle.name, "Imported PDF Bus")

    def test_frontend_fleet_import_page_accepts_pdf_preview(self):
        self.client.force_login(self.superuser)
        pdf_upload = SimpleUploadedFile(
            "fleet.pdf",
            b"%PDF-1.4 fake pdf",
            content_type="application/pdf",
        )

        with patch("busstops.fleet_imports.parse_pdf") as mocked_parse_pdf:
            mocked_parse_pdf.return_value = [
                ParsedFleetRecord(
                    operator_code="OT97",
                    code="903",
                    fleet_number="903",
                    fleet_code="903",
                    registration="HF17AZC",
                    vehicle_type="Volvo B11R Plaxton Elite",
                    livery="Excelsior (light)",
                    garage="GSC Pimperne",
                )
            ]
            response = self.client.post(
                self.fleet_import_url(),
                {
                    "operator": self.operator.pk,
                    "manual_livery_selection": "on",
                    "rows_text": "",
                    "action": "preview",
                    "upload": pdf_upload,
                },
            )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Fleet import")
        self.assertContains(response, "Create operator OT97")
        self.assertContains(response, "Excelsior (light)")

    def test_frontend_fleet_import_shows_clean_error_when_pdf_support_is_missing(self):
        self.client.force_login(self.superuser)
        pdf_upload = SimpleUploadedFile(
            "fleet.pdf",
            b"%PDF-1.4 fake pdf",
            content_type="application/pdf",
        )

        with patch("busstops.fleet_imports.parse_pdf") as mocked_parse_pdf:
            mocked_parse_pdf.side_effect = RuntimeError("pdfplumber is required")
            response = self.client.post(
                self.fleet_import_url(),
                {
                    "operator": self.operator.pk,
                    "rows_text": "",
                    "action": "preview",
                    "upload": pdf_upload,
                },
            )

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            "PDF import is not available on this server yet because the PDF extraction dependency is missing.",
        )

    def test_frontend_fleet_import_commit_uses_manual_livery_mapping(self):
        self.client.force_login(self.superuser)
        mapped_livery = Livery.objects.create(
            name="Excelsior Light",
            colour="#ffffff",
            left_css="#ffffff",
        )
        pdf_upload = SimpleUploadedFile(
            "fleet.pdf",
            b"%PDF-1.4 fake pdf",
            content_type="application/pdf",
        )

        with patch("busstops.fleet_imports.parse_pdf") as mocked_parse_pdf:
            mocked_parse_pdf.return_value = [
                ParsedFleetRecord(
                    operator_code=self.operator.noc,
                    code="904",
                    fleet_number="904",
                    fleet_code="904",
                    registration="HF17AZD",
                    vehicle_type="Volvo B11R Plaxton Elite",
                    livery="Excelsior (light)",
                    garage="GSC Pimperne",
                )
            ]
            response = self.client.post(
                self.fleet_import_url(),
                {
                    "operator": self.operator.pk,
                    "manual_livery_selection": "on",
                    "rows_text": "",
                    "action": "commit",
                    "livery_map_0": str(mapped_livery.pk),
                    "upload": pdf_upload,
                },
            )

        self.assertEqual(response.status_code, 200)
        created_vehicle = Vehicle.objects.get(operator=self.operator, code="904")
        self.assertEqual(created_vehicle.livery, mapped_livery)

    def test_frontend_fleet_import_uses_depot_to_match_operator(self):
        self.client.force_login(self.superuser)
        depot_operator = Operator.objects.create(
            noc="DP01",
            name="Depot Matched Operator",
            slug="depot-matched-operator",
        )
        Garage.objects.create(operator=depot_operator, code="PIM", name="Pimperne")
        pdf_upload = SimpleUploadedFile(
            "fleet.pdf",
            b"%PDF-1.4 fake pdf",
            content_type="application/pdf",
        )

        with patch("busstops.fleet_imports.parse_pdf") as mocked_parse_pdf:
            mocked_parse_pdf.return_value = [
                ParsedFleetRecord(
                    operator_code="",
                    code="905",
                    fleet_number="905",
                    fleet_code="905",
                    registration="HF17AZE",
                    vehicle_type="Volvo B11R Plaxton Elite",
                    livery="Excelsior (dark)",
                    garage="GSC Pimperne",
                )
            ]
            response = self.client.post(
                self.fleet_import_url(),
                {
                    "operator": self.operator.pk,
                    "rows_text": "",
                    "action": "preview",
                    "upload": pdf_upload,
                },
            )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Match DP01 - Depot Matched Operator via depot")

    def test_admin_vehicle_mass_import_page_is_available(self):
        self.client.force_login(self.superuser)
        response = self.client.get(self.admin_vehicle_mass_import_url())
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Mass import vehicles")

    def test_admin_vehicle_mass_import_blank_operator_does_not_override_existing_operator(self):
        self.client.force_login(self.superuser)
        response = self.client.post(
            self.admin_vehicle_mass_import_url(),
            {
                "operator": "",
                "rows_text": "\n".join(
                    [
                        "code,name,operator_code",
                        "100,Still Same Operator,",
                    ]
                ),
                "action": "commit",
            },
        )
        self.assertEqual(response.status_code, 200)
        self.existing_vehicle.refresh_from_db()
        self.assertEqual(self.existing_vehicle.operator, self.operator)
        self.assertEqual(self.existing_vehicle.name, "Still Same Operator")

    def test_admin_vehicle_mass_import_persists_external_link_when_match_is_by_registration(self):
        self.client.force_login(self.superuser)
        self.existing_vehicle.external_id = None
        self.existing_vehicle.save(update_fields=["external_id"])

        response = self.client.post(
            self.admin_vehicle_mass_import_url(),
            {
                "operator": self.operator.pk,
                "rows_text": "\n".join(
                    [
                        "external_id,registration,name",
                        "import-vehicle-100,AB12CDE,Linked By Registration",
                    ]
                ),
                "action": "commit",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.existing_vehicle.refresh_from_db()
        self.assertEqual(self.existing_vehicle.external_id, "import-vehicle-100")
        self.assertEqual(self.existing_vehicle.name, "Linked By Registration")



    def operator_changelist_url(self):
        return "/admin/busstops/operator/"

    def test_bulk_assign_to_group_action_updates_selected_operators(self):
        self.client.force_login(self.superuser)
        operator_two = Operator.objects.create(noc="MAS2", name="Mass Operator 2", slug="mass-operator-2")
        group = OperatorGroup.objects.create(name="Bulk Group", slug="bulk-group")

        response = self.client.post(
            self.operator_changelist_url(),
            {
                "action": "assign_to_group",
                ACTION_CHECKBOX_NAME: [self.operator.pk, operator_two.pk],
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Assign selected operators to a group")

        response = self.client.post(
            self.operator_changelist_url(),
            {
                "action": "assign_to_group",
                "apply": "1",
                "group": group.pk,
                ACTION_CHECKBOX_NAME: [self.operator.pk, operator_two.pk],
            },
            follow=True,
        )
        self.assertEqual(response.status_code, 200)

        self.operator.refresh_from_db()
        operator_two.refresh_from_db()
        self.assertEqual(self.operator.group, group)
        self.assertEqual(operator_two.group, group)
        self.assertTrue(self.operator.is_manual)
        self.assertTrue(operator_two.is_manual)

    def test_bulk_assign_to_organisation_action_updates_selected_operators(self):
        self.client.force_login(self.superuser)
        operator_two = Operator.objects.create(noc="MAS3", name="Mass Operator 3", slug="mass-operator-3")
        organisation = Organisation.objects.create(name="Bulk Org", slug="bulk-org")

        response = self.client.post(
            self.operator_changelist_url(),
            {
                "action": "assign_to_organisation",
                ACTION_CHECKBOX_NAME: [self.operator.pk, operator_two.pk],
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Assign selected operators to an organisation")

        response = self.client.post(
            self.operator_changelist_url(),
            {
                "action": "assign_to_organisation",
                "apply": "1",
                "organisation": organisation.pk,
                ACTION_CHECKBOX_NAME: [self.operator.pk, operator_two.pk],
            },
            follow=True,
        )
        self.assertEqual(response.status_code, 200)

        self.operator.refresh_from_db()
        operator_two.refresh_from_db()
        self.assertEqual(self.operator.organisation, organisation)
        self.assertEqual(operator_two.organisation, organisation)
        self.assertTrue(self.operator.is_manual)
        self.assertTrue(operator_two.is_manual)


