from io import BytesIO
from unittest.mock import Mock, patch

from django.contrib.admin import site
from django.contrib.contenttypes.models import ContentType
from django.contrib.gis.geos import Point
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse
from django.test.utils import override_settings

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from accounts.models import User
from bustimes.models import Calendar, Route, RouteLink, StopTime, Trip
from vehicles.models import Vehicle, VehicleType, VehicleTypeGroup

from .models import DataSource, Manufacturer, Operator, PreservationGroup, Service, StopGroup, StopPoint


class BusStopsAdminTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.source = DataSource.objects.create()

        cls.service_a = Service.objects.create(
            line_name="129", description="Frankby Cemetery - Liscard"
        )
        cls.service_b = Service.objects.create(
            line_name="129A", description="Frankby - Moreton - Liscard"
        )

        stop_a = StopPoint.objects.create(
            atco_code="2902", active=True, common_name="Sandy Corner"
        )
        stop_b = StopPoint.objects.create(
            atco_code="2903", active=True, common_name="Leafy Hollow"
        )
        RouteLink.objects.create(
            from_stop=stop_a,
            to_stop=stop_b,
            service=cls.service_a,
            geometry="LINESTRING(1.2 51.1,1.1 51.2)",
        )
        RouteLink.objects.create(
            from_stop=stop_a,
            to_stop=stop_b,
            service=cls.service_b,
            geometry="LINESTRING(1.3 51.1,1.1 51.3)",
        )
        Route.objects.create(
            source=cls.source,
            service_code="129",
            code="129",
            line_name="129",
            service=cls.service_a,
        )
        Route.objects.create(
            source=cls.source,
            service_code="129",
            code="129A",
            line_name="129A",
            service=cls.service_b,
        )
        cls.calendar = Calendar.objects.create(
            mon=True,
            start_date="2026-01-01",
        )
        cls.stop_timetable_a = StopPoint.objects.create(
            atco_code="3901", active=True, common_name="Alpha"
        )
        cls.stop_timetable_b = StopPoint.objects.create(
            atco_code="3902", active=True, common_name="Bravo"
        )
        cls.service_timetable = Service.objects.create(
            line_name="55", description="Station - Depot"
        )
        cls.route_timetable = Route.objects.create(
            source=cls.source,
            service_code="55",
            code="55-out",
            line_name="55",
            service=cls.service_timetable,
        )
        cls.trip_timetable = Trip.objects.create(
            route=cls.route_timetable,
            calendar=cls.calendar,
            inbound=False,
            start="08:00",
            end="08:10",
            headsign="Depot",
        )
        StopTime.objects.bulk_create(
            [
                StopTime(
                    trip=cls.trip_timetable,
                    stop=cls.stop_timetable_a,
                    stop_code=cls.stop_timetable_a.atco_code,
                    departure="08:00",
                    sequence=1,
                    timing_status="PTP",
                ),
                StopTime(
                    trip=cls.trip_timetable,
                    stop=cls.stop_timetable_b,
                    stop_code=cls.stop_timetable_b.atco_code,
                    arrival="08:10",
                    sequence=2,
                    timing_status="PTP",
                ),
            ]
        )

        cls.staff_user = User.objects.create(
            username="josh", is_staff=True, is_superuser=True, email="j@example.com"
        )

    def setUp(self):
        ContentType.objects.clear_cache()

    def test_preservation_group_admin(self):
        group = PreservationGroup.objects.create(
            name="Admin Preservation Group",
            slug="admin-preservation-group",
            description="Admin searchable group",
        )
        Vehicle.objects.create(code="ADMINGROUP", preserved=True, preservation_group=group)
        self.client.force_login(self.staff_user)

        response = self.client.get("/admin/busstops/preservationgroup/?q=searchable")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Admin Preservation Group")
        self.assertContains(response, "1")
        self.assertIn(PreservationGroup, site._registry)

    def make_url(self, site, model, page: str) -> str:
        return reverse(
            f"{site.name}:{model._meta.app_label}_{model._meta.model_name}_{page}"
        )

    def make_timetable_upload(self, rows):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Timetable"
        for row in rows:
            worksheet.append(row)
        output = BytesIO()
        workbook.save(output)
        return SimpleUploadedFile(
            "timetable.xlsx",
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_everything(self):
        self.client.force_login(self.staff_user)

        for model, model_admin in site._registry.items():
            url = self.make_url(site, model, "changelist")
            response = self.client.get(url, {"q": "blah"})
            self.assertEqual(response.status_code, 200)

            url = self.make_url(site, model, "add")
            response = self.client.get(url, {"q": "blah"})
            self.assertEqual(response.status_code, 200)

    def test_merge_and_unmerge(self):
        self.client.force_login(self.staff_user)

        self.client.post(
            "/admin/busstops/service/",
            {
                "action": "merge",
                "_selected_action": [self.service_a.id, self.service_b.id],
            },
        )
        response = self.client.get("/admin/busstops/service/")
        self.assertEqual(
            list(response.context["messages"])[0].message,
            "merged <QuerySet [<Service: 129A - Frankby - Moreton - Liscard>]> into 129 - Frankby Cemetery - Liscard",
        )

        # merged into 1:
        self.assertEqual(Service.objects.count(), 1)

        # unmerge back into 2:
        self.client.post(
            "/admin/busstops/service/",
            {
                "action": "unmerge",
                "_selected_action": [self.service_a.id],
            },
        )
        self.assertEqual(Service.objects.count(), 2)

    def test_split_service_filter(self):
        self.client.force_login(self.staff_user)
        response = self.client.get(
            "/admin/busstops/service/?split=1",
        )
        self.assertContains(response, "Frankby Cemetery - Liscard")

    def test_data_source_admin(self):
        url = "/admin/busstops/datasource/"

        self.client.force_login(self.staff_user)
        response = self.client.get(url)
        self.assertContains(response, ">0<")  # services
        self.assertContains(response, ">2<")  # routes

        self.client.post(
            url,
            {
                "action": "delete_routes",
                "_selected_action": [self.source.id],
            },
        )
        response = self.client.get(url)

        with self.assertNumQueries(8):
            self.client.post(
                url,
                {
                    "action": "remove_datetimes",
                    "_selected_action": [self.source.id],
                },
            )

    def test_routelink_admin(self):
        url = "/admin/bustimes/routelink/"

        self.client.force_login(self.staff_user)

        response = self.client.get(url + "?dodgy=from_stop")
        self.assertEqual(response.context_data["cl"].result_count, 0)

        # move the stop away from the route link so it's "dodgy"
        StopPoint.objects.all().update(latlong="POINT(1.9 51.9)")

        response = self.client.get(url + "?dodgy=from_stop")
        self.assertEqual(response.context_data["cl"].result_count, 2)

        response = self.client.get(url + "?dodgy=to_stop")
        self.assertEqual(response.context_data["cl"].result_count, 2)

        url = RouteLink.objects.first().get_absolute_url()

        res = self.client.get(url)
        self.assertContains(res, "from Sandy Corner")
        self.assertContains(res, "to Leafy Hollow")

    def test_stop_group_admin_renders_selector_map(self):
        self.client.force_login(self.staff_user)
        response = self.client.get("/admin/busstops/stopgroup/add/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="stop-group-selector-map"')
        self.assertContains(response, 'name="stops_selection"')

    def test_stop_group_admin_saves_selected_stops_in_order(self):
        self.client.force_login(self.staff_user)
        stop_c = StopPoint.objects.create(
            atco_code="2904",
            active=True,
            common_name="Map Stop",
            latlong=Point(-2.9, 53.4),
        )
        stop_d = StopPoint.objects.create(
            atco_code="2905",
            active=True,
            common_name="Map Stop 2",
            latlong=Point(-2.8, 53.5),
        )

        response = self.client.post(
            "/admin/busstops/stopgroup/add/",
            {
                "name": "Test group",
                "slug": "test-group",
                "active": "on",
                "stops_selection": f"{stop_d.atco_code},{stop_c.atco_code}",
                "_save": "Save",
            },
        )

        self.assertEqual(response.status_code, 302)
        group = StopGroup.objects.get(slug="test-group")
        self.assertEqual(
            list(group.stopgroupstop_set.order_by("order").values_list("stop_id", flat=True)),
            [stop_d.atco_code, stop_c.atco_code],
        )

    @override_settings(STADIA_MAPS_API_KEY="test-key")
    @patch("busstops.admin.requests.Session")
    def test_service_route_editor_generates_snapped_geometry(self, session_cls):
        self.client.force_login(self.staff_user)
        stop_c = StopPoint.objects.create(
            atco_code="2904",
            active=True,
            common_name="Map Stop",
            latlong=Point(-2.9, 53.4),
        )
        stop_d = StopPoint.objects.create(
            atco_code="2905",
            active=True,
            common_name="Map Stop 2",
            latlong=Point(-2.8, 53.5),
        )
        self.service_a.stopusage_set.create(
            stop=stop_c,
            order=0,
            timing_point=True,
            inbound=False,
            line_name=self.service_a.line_name,
        )
        self.service_a.stopusage_set.create(
            stop=stop_d,
            order=1,
            timing_point=True,
            inbound=False,
            line_name=self.service_a.line_name,
        )

        response = Mock()
        response.raise_for_status.return_value = None
        response.json.return_value = {
            "trip": {
                "legs": [
                    {
                        "shape": "_xdrI~bws@_pR_pR",
                    }
                ]
            }
        }
        session = Mock()
        session.post.return_value = response
        session_cls.return_value = session

        res = self.client.post(
            "/admin/busstops/service/tools/route-editor/",
            {
                "action": "generate_geometry",
                "service": self.service_a.id,
            },
        )

        self.assertEqual(res.status_code, 200)
        route_link = RouteLink.objects.get(
            service=self.service_a,
            from_stop=stop_c,
            to_stop=stop_d,
        )
        self.assertTrue(route_link.override)

    def test_superuser_can_download_service_timetable_template(self):
        self.client.force_login(self.staff_user)
        response = self.client.get(
            f"/admin/busstops/service/{self.service_timetable.pk}/mass-edit-timetable/template.xlsx"
        )
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(workbook.active.title, "Timetable")
        headers = [cell.value for cell in workbook.active[1]]
        self.assertEqual(
            headers[:8],
            [
                "import_key",
                "trip_id",
                "route_id",
                "line_name",
                "calendar_id",
                "inbound",
                "sequence",
                "stop_atco_code",
            ],
        )

    def test_superuser_can_download_current_service_timetable_workbook(self):
        self.client.force_login(self.staff_user)
        response = self.client.get(
            f"/admin/busstops/service/{self.service_timetable.pk}/mass-edit-timetable/current.xlsx"
        )
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        rows = list(workbook["Timetable"].iter_rows(values_only=True))
        self.assertEqual(rows[1][1], self.trip_timetable.pk)
        self.assertEqual(rows[1][7], self.stop_timetable_a.atco_code)

    def test_superuser_can_download_simple_service_timetable_template(self):
        self.client.force_login(self.staff_user)
        response = self.client.get(
            f"/admin/busstops/service/{self.service_timetable.pk}/mass-edit-timetable/simple-template.xlsx"
        )
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(workbook.sheetnames, ["Outbound", "Inbound", "Instructions"])
        worksheet = workbook["Outbound"]
        self.assertEqual(worksheet["A8"].value, "Stop name")
        self.assertEqual(worksheet["A9"].value, self.stop_timetable_a.common_name)
        self.assertEqual(worksheet["B9"].value, self.stop_timetable_a.atco_code)
        self.assertEqual(worksheet["C9"].value, "08:00")
        self.assertTrue(worksheet["C9"].font.bold)

    def test_service_timetable_preview_accepts_uploaded_simple_workbook(self):
        self.client.force_login(self.staff_user)
        response = self.client.get(
            f"/admin/busstops/service/{self.service_timetable.pk}/mass-edit-timetable/simple-template.xlsx"
        )
        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook["Outbound"]
        worksheet["C9"] = "08:05"
        worksheet["C9"].font = Font(bold=True)
        worksheet["C10"] = "08:15"
        worksheet["C10"].font = Font(bold=False)

        output = BytesIO()
        workbook.save(output)
        upload = SimpleUploadedFile(
            "simple-timetable.xlsx",
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        preview = self.client.post(
            f"/admin/busstops/service/{self.service_timetable.pk}/mass-edit-timetable/",
            {"workbook": upload, "action": "preview"},
        )

        self.assertEqual(preview.status_code, 200)
        self.assertContains(preview, str(self.trip_timetable.pk))

    def test_simple_timetable_uses_stop_labels_and_bold_names(self):
        self.client.force_login(self.staff_user)
        response = self.client.get(
            f"/admin/busstops/service/{self.service_timetable.pk}/mass-edit-timetable/simple-template.xlsx"
        )
        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook["Outbound"]
        worksheet["A9"] = "Bus Station (stand A)"
        worksheet["A9"].font = Font(bold=True)
        worksheet["C9"] = "08:05"
        worksheet["A10"] = "Terminus"
        worksheet["A10"].font = Font(bold=False)
        worksheet["C10"] = "08:15"

        output = BytesIO()
        workbook.save(output)
        upload = SimpleUploadedFile(
            "simple-timetable.xlsx",
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response = self.client.post(
            f"/admin/busstops/service/{self.service_timetable.pk}/mass-edit-timetable/",
            {"workbook": upload, "action": "commit"},
        )

        self.assertEqual(response.status_code, 200)
        stop_times = list(self.trip_timetable.stoptime_set.order_by("sequence"))
        self.assertEqual(stop_times[0].display_name, "Bus Station (stand A)")
        self.assertEqual(stop_times[0].timing_status, "PTP")
        self.assertEqual(stop_times[1].display_name, "Terminus")
        self.assertEqual(stop_times[1].timing_status, "OTH")
        public_timetable = self.client.get(self.service_timetable.get_absolute_url())
        self.assertContains(public_timetable, "Bus Station (stand A)")

    def test_service_timetable_preview_accepts_uploaded_workbook(self):
        self.client.force_login(self.staff_user)
        workbook = self.make_timetable_upload(
            [
                [
                    "import_key",
                    "trip_id",
                    "route_id",
                    "line_name",
                    "calendar_id",
                    "inbound",
                    "sequence",
                    "stop_atco_code",
                    "stop_name",
                    "arrival",
                    "departure",
                    "pick_up",
                    "set_down",
                ],
                [
                    "trip-1",
                    self.trip_timetable.pk,
                    self.route_timetable.pk,
                    "55",
                    self.calendar.pk,
                    "false",
                    1,
                    self.stop_timetable_a.atco_code,
                    "Alpha",
                    "",
                    "08:05",
                    "true",
                    "true",
                ],
                [
                    "trip-1",
                    self.trip_timetable.pk,
                    self.route_timetable.pk,
                    "55",
                    self.calendar.pk,
                    "false",
                    2,
                    self.stop_timetable_b.atco_code,
                    "Bravo",
                    "08:15",
                    "",
                    "true",
                    "true",
                ],
            ]
        )
        response = self.client.post(
            f"/admin/busstops/service/{self.service_timetable.pk}/mass-edit-timetable/",
            {"rows_text": "", "action": "preview", "workbook": workbook},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Preview generated")
        self.assertContains(response, str(self.trip_timetable.pk))

    def test_service_timetable_commit_updates_existing_trip_and_creates_new_trip(self):
        self.client.force_login(self.staff_user)
        response = self.client.post(
            f"/admin/busstops/service/{self.service_timetable.pk}/mass-edit-timetable/",
            {
                "rows_text": "\n".join(
                    [
                        "import_key,trip_id,route_id,line_name,calendar_id,inbound,sequence,stop_atco_code,arrival,departure,pick_up,set_down,headsign",
                        f"trip-{self.trip_timetable.pk},{self.trip_timetable.pk},{self.route_timetable.pk},55,{self.calendar.pk},false,1,{self.stop_timetable_a.atco_code},,08:05,true,true,Updated Headsign",
                        f"trip-{self.trip_timetable.pk},{self.trip_timetable.pk},{self.route_timetable.pk},55,{self.calendar.pk},false,2,{self.stop_timetable_b.atco_code},08:15,,true,true,Updated Headsign",
                        f"new-1,,{self.route_timetable.pk},55,{self.calendar.pk},true,1,{self.stop_timetable_b.atco_code},,09:00,true,true,Inbound Service",
                        f"new-1,,{self.route_timetable.pk},55,{self.calendar.pk},true,2,{self.stop_timetable_a.atco_code},09:10,,true,true,Inbound Service",
                    ]
                ),
                "action": "commit",
            },
        )
        self.assertEqual(response.status_code, 200)

        self.trip_timetable.refresh_from_db()
        self.assertEqual(self.trip_timetable.headsign, "Updated Headsign")
        self.assertEqual(str(self.trip_timetable.start), "08:05")
        self.assertEqual(str(self.trip_timetable.end), "08:15")
        updated_times = list(self.trip_timetable.stoptime_set.order_by("sequence"))
        self.assertEqual(str(updated_times[0].departure), "08:05")
        self.assertEqual(str(updated_times[1].arrival), "08:15")

        new_trip = Trip.objects.exclude(pk=self.trip_timetable.pk).get(route=self.route_timetable)
        self.assertTrue(new_trip.inbound)
        self.assertEqual(new_trip.headsign, "Inbound Service")
        self.assertEqual(str(new_trip.start), "09:00")
        self.assertEqual(str(new_trip.end), "09:10")
        self.assertEqual(new_trip.stoptime_set.count(), 2)

    def test_service_bulk_assign_operator_action(self):
        self.client.force_login(self.staff_user)
        operator = Operator.objects.create(noc="FHAM", name="First Hampshire")
        other_operator = Operator.objects.create(noc="SCBL", name="Stagecoach Blue")
        self.service_a.operator.add(other_operator)
        self.service_b.operator.add(other_operator)

        response = self.client.post(
            "/admin/busstops/service/",
            {
                "action": "assign_to_operator",
                "_selected_action": [self.service_a.id, self.service_b.id],
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Selected services")

        self.client.post(
            "/admin/busstops/service/",
            {
                "action": "assign_to_operator",
                "_selected_action": [self.service_a.id, self.service_b.id],
                "apply": "1",
                "operator": operator.pk,
            },
        )

        self.service_a.refresh_from_db()
        self.service_b.refresh_from_db()
        self.assertEqual(
            list(self.service_a.operator.order_by("noc").values_list("noc", flat=True)),
            ["FHAM"],
        )
        self.assertEqual(
            list(self.service_b.operator.order_by("noc").values_list("noc", flat=True)),
            ["FHAM"],
        )

    def test_vehicle_type_bulk_assign_manufacturer_action(self):
        self.client.force_login(self.staff_user)
        manufacturer = Manufacturer.objects.create(name="Volvo", slug="volvo")
        vt1 = VehicleType.objects.create(name="B5TL")
        vt2 = VehicleType.objects.create(name="7900 Electric")

        response = self.client.post(
            "/admin/vehicles/vehicletype/",
            {
                "action": "assign_to_manufacturer",
                "_selected_action": [vt1.id, vt2.id],
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Selected vehicle types")

        self.client.post(
            "/admin/vehicles/vehicletype/",
            {
                "action": "assign_to_manufacturer",
                "_selected_action": [vt1.id, vt2.id],
                "apply": "1",
                "manufacturer": manufacturer.id,
            },
        )

        vt1.refresh_from_db()
        vt2.refresh_from_db()
        self.assertEqual(vt1.manufacturer, manufacturer)
        self.assertEqual(vt2.manufacturer, manufacturer)

    def test_vehicle_type_bulk_assign_vehicle_group_action(self):
        self.client.force_login(self.staff_user)
        manufacturer = Manufacturer.objects.create(name="Alexander Dennis", slug="alexander-dennis")
        vehicle_group = VehicleTypeGroup.objects.create(
            manufacturer=manufacturer,
            name="Enviro200",
        )
        vt1 = VehicleType.objects.create(name="Enviro200MMC")
        vt2 = VehicleType.objects.create(name="Enviro200EV")

        response = self.client.post(
            "/admin/vehicles/vehicletype/",
            {
                "action": "assign_to_vehicle_group",
                "_selected_action": [vt1.id, vt2.id],
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Selected vehicle types")

        self.client.post(
            "/admin/vehicles/vehicletype/",
            {
                "action": "assign_to_vehicle_group",
                "_selected_action": [vt1.id, vt2.id],
                "apply": "1",
                "vehicle_group": vehicle_group.id,
            },
        )

        vt1.refresh_from_db()
        vt2.refresh_from_db()
        self.assertEqual(vt1.vehicle_group, vehicle_group)
        self.assertEqual(vt2.vehicle_group, vehicle_group)

    def test_vehicle_type_bulk_remove_manufacturer_action(self):
        self.client.force_login(self.staff_user)
        manufacturer = Manufacturer.objects.create(name="Volvo", slug="volvo")
        vt1 = VehicleType.objects.create(name="B5TL", manufacturer=manufacturer)
        vt2 = VehicleType.objects.create(name="7900 Electric", manufacturer=manufacturer)

        response = self.client.post(
            "/admin/vehicles/vehicletype/",
            {
                "action": "remove_from_manufacturer",
                "_selected_action": [vt1.id, vt2.id],
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        vt1.refresh_from_db()
        vt2.refresh_from_db()
        self.assertIsNone(vt1.manufacturer)
        self.assertIsNone(vt2.manufacturer)

    def test_vehicle_type_bulk_remove_vehicle_group_action(self):
        self.client.force_login(self.staff_user)
        manufacturer = Manufacturer.objects.create(name="Alexander Dennis", slug="alexander-dennis")
        vehicle_group = VehicleTypeGroup.objects.create(
            manufacturer=manufacturer,
            name="Enviro200",
        )
        vt1 = VehicleType.objects.create(name="Enviro200MMC", vehicle_group=vehicle_group)
        vt2 = VehicleType.objects.create(name="Enviro200EV", vehicle_group=vehicle_group)

        response = self.client.post(
            "/admin/vehicles/vehicletype/",
            {
                "action": "remove_from_vehicle_group",
                "_selected_action": [vt1.id, vt2.id],
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        vt1.refresh_from_db()
        vt2.refresh_from_db()
        self.assertIsNone(vt1.vehicle_group)
        self.assertIsNone(vt2.vehicle_group)
