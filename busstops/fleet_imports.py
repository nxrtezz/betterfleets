from __future__ import annotations

import csv
from io import StringIO

from django.db import connection, transaction
from django.db.models import Q
from django.utils import timezone
from django.utils.text import slugify
from openpyxl import load_workbook

from bustimes.models import Garage
from fleet.parsers.pdf_fleet_parser import TARGET_COLUMNS, parse_pdf
from vehicles.models import Livery, Vehicle, VehicleFeature, VehicleType, vehicle_slug

from .models import Operator


HEADER_ALIASES = {
    "fleet_num": "fleet_number",
    "fleet": "fleet_number",
    "registration": "reg",
    "prev_reg": "prev_registration",
    "previous_reg": "prev_registration",
    "vehicle_id": "vehicle_type",
    "vehicle_type_id": "vehicle_type",
    "livery_id": "livery",
    "color": "colours",
    "colors": "colours",
    "garage_id": "garage",
    "operator_noc": "operator_code",
    "operator": "operator_code",
    "historical_operator": "historical_fleet",
    "historical_operator_code": "historical_fleet",
    "historical_fleet_operator": "historical_fleet",
    "historical_fleet_operator_code": "historical_fleet",
    "historical_year": "year",
    "fleet_year": "year",
    # Advanced field aliases
    "engine": "advanced_engine",
    "seating_capacity": "advanced_seating_capacity",
    "seating-capacity": "advanced_seating_capacity",
    "gearbox": "advanced_gearbox",
}


def current_vehicle_filters(**filters):
    try:
        with connection.cursor() as cursor:
            columns = {
                column.name
                for column in connection.introspection.get_table_description(
                    cursor, Vehicle._meta.db_table
                )
            }
    except Exception:
        columns = {
            field.column
            for field in Vehicle._meta.fields
            if getattr(field, "column", None)
        }

    unsupported = {
        "withdrawn": "withdrawn",
        "preserved": "preserved",
        "operator": "operator_id",
        "operator_id": "operator_id",
    }
    for key, column in unsupported.items():
        if key in filters and column not in columns:
            filters.pop(key)
    if "historical_fleet_id" in columns:
        filters["historical_fleet__isnull"] = True
    return filters


def vehicle_db_columns():
    try:
        with connection.cursor() as cursor:
            return {
                column.name
                for column in connection.introspection.get_table_description(
                    cursor, Vehicle._meta.db_table
                )
            }
    except Exception:
        return {
            field.column
            for field in Vehicle._meta.fields
            if getattr(field, "column", None)
        }


def historical_code_with_year(code, historical_year):
    if not code or historical_year is None:
        return code
    return f"{code}-{historical_year}"


def rows_text_from_workbook(uploaded_file):
    if not uploaded_file:
        return ""

    filename = (uploaded_file.name or "").lower()
    if filename.endswith(".csv"):
        try:
            return uploaded_file.read().decode("utf-8-sig").strip()
        except UnicodeDecodeError as exc:
            raise ValueError("CSV upload must be UTF-8 encoded") from exc
    if not filename.endswith(".xlsx"):
        raise ValueError("Upload must be a .pdf, .xlsx, or .csv file")

    workbook = load_workbook(uploaded_file, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return ""

    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    if not any(headers):
        return ""

    # Check if this is a basic/advanced fleet export format (with operator info header)
    # Basic format has "NOC" in first cell of row 1
    if len(rows) >= 4 and headers[0].upper() == "NOC":
        # This is a basic/advanced export format
        # Skip rows 1-3 (operator info), row 4 is the actual header
        if len(rows) >= 4:
            headers = [str(value).strip() if value is not None else "" for value in rows[3]]
            data_rows = rows[4:]  # Skip the header row too
        else:
            data_rows = []
    else:
        # Standard format - first row is headers
        data_rows = rows[1:]

    output = StringIO()
    writer = csv.writer(output, delimiter="\t", lineterminator="\n")
    writer.writerow(headers)
    for row in data_rows:
        values = ["" if value is None else str(value).strip() for value in row[: len(headers)]]
        if any(values):
            writer.writerow(values)
    return output.getvalue()


def rows_text_from_pdf(default_operator: Operator, uploaded_file):
    try:
        rows = parse_pdf(
            uploaded_file,
            default_operator_code=default_operator.noc if default_operator else "",
        )
    except RuntimeError as exc:
        raise ValueError(
            "PDF import is not available on this server yet because the PDF extraction dependency is missing."
        ) from exc
    output = StringIO()
    writer = csv.writer(output, delimiter="\t", lineterminator="\n")
    writer.writerow(list(TARGET_COLUMNS) + ["features"])
    for row in rows:
        writer.writerow(
            [
                row.operator_code or default_operator.noc,
                row.external_id,
                row.code,
                row.fleet_number,
                row.fleet_code,
                row.registration,
                row.prev_registration,
                row.vehicle_type,
                row.livery,
                row.colours,
                row.garage,
                row.name,
                row.branding,
                row.notes,
                "true" if row.withdrawn else "false",
                "true" if row.preserved else "false",
                "true" if row.fleet_support_vehicle else "false",
                "true" if row.vor else "false",
                "true" if row.awaiting_delivery else "false",
                "true" if row.trainer_vehicle else "false",
                "true" if row.demonstrator else "false",
                "",
            ]
        )
    return output.getvalue()


def rows_text_from_upload(default_operator: Operator, uploaded_file):
    if not uploaded_file:
        return ""
    filename = (uploaded_file.name or "").lower()
    if filename.endswith(".pdf"):
        return rows_text_from_pdf(default_operator, uploaded_file)
    return rows_text_from_workbook(uploaded_file)


def normalise_header(header):
    key = header.strip().lower().replace(" ", "_")
    return HEADER_ALIASES.get(key, key)


def coerce_bool(value):
    value = value.strip().lower()
    if value in {"", "none", "null"}:
        return None
    if value in {"1", "true", "yes", "y", "on"}:
        return True
    if value in {"0", "false", "no", "n", "off"}:
        return False
    raise ValueError(f"Invalid boolean value '{value}'")


def resolve_reference(model, value, label):
    if value == "":
        return None
    if value.isdigit():
        try:
            return model.objects.get(pk=int(value))
        except model.DoesNotExist as exc:
            raise ValueError(f"Unknown {label} id '{value}'") from exc

    try:
        return model.objects.get(external_id=value)
    except (model.DoesNotExist, model.MultipleObjectsReturned):
        pass

    if hasattr(model, "name"):
        item = model.objects.filter(name__iexact=value).first()
        if item:
            return item

    if hasattr(model, "code"):
        item = model.objects.filter(code__iexact=value).first()
        if item:
            return item

    raise ValueError(f"Unknown {label} '{value}'")


def resolve_operator_preview(default_operator, value, *, allow_create):
    if not value:
        if default_operator:
            return default_operator, "match", f"Match {default_operator.noc} - {default_operator.name}"
        return None, "", "Leave operator unchanged"
    text = str(value).strip()
    if not text:
        if default_operator:
            return default_operator, "match", f"Match {default_operator.noc} - {default_operator.name}"
        return None, "", "Leave operator unchanged"

    operator = Operator.objects.filter(
        Q(noc__iexact=text) | Q(slug__iexact=text) | Q(operatorcode__code__iexact=text)
    ).first()
    if operator:
        return operator, "match", f"Match {operator.noc} - {operator.name}"
    if allow_create:
        return None, "create", f"Create operator {text}"
    raise ValueError(f"Unknown operator_code '{text}'")


def resolve_existing_operator(value, label):
    text = str(value or "").strip()
    if not text:
        return None
    operator = Operator.objects.filter(
        Q(noc__iexact=text) | Q(slug__iexact=text) | Q(operatorcode__code__iexact=text)
    ).first()
    if operator:
        return operator
    raise ValueError(f"Unknown {label} '{text}'")


def resolve_garage_preview(operator, value, *, allow_create):
    if not value:
        return None, "", ""
    text = str(value).strip()
    if not text:
        return None, "", ""
    trimmed = text[4:].strip() if text.upper().startswith("GSC ") else text
    global_matches = Garage.objects.filter(
        Q(name__iexact=text)
        | Q(code__iexact=text)
        | Q(name__iexact=trimmed)
        | Q(code__iexact=trimmed)
    )
    if operator:
        garage = global_matches.filter(operator=operator).first()
        if garage:
            label = garage.name or garage.code or str(garage.pk)
            operator_label = garage.operator_id or "unknown operator"
            return garage, "match", f"Match depot {label} for {operator_label}"
    else:
        garage = global_matches.first()
        if garage and global_matches.count() == 1:
            label = garage.name or garage.code or str(garage.pk)
            operator_label = garage.operator_id or "unknown operator"
            return garage, "match", f"Match depot {label} for {operator_label}"
    if not operator and global_matches.count() == 1:
        garage = global_matches.first()
        label = garage.name or garage.code or str(garage.pk)
        operator_label = garage.operator_id or "unknown operator"
        return garage, "match", f"Match depot {label} for {operator_label}"
    if allow_create:
        operator_label = operator.noc if operator else "new operator"
        return None, "create", f"Create depot {text} for {operator_label}"
    raise ValueError(f"Unknown garage '{text}'")


def _vehicle_match_queryset(operator, historical_fleet=None, historical_year=None):
    if not operator:
        return Vehicle.objects.none()

    if historical_year is not None:
        filters = {"historical_fleet": historical_fleet or operator}
        columns = vehicle_db_columns()
        if "historical_fleet_year" in columns:
            filters["historical_fleet_year"] = historical_year
        return operator.vehicle_set.filter(**filters)

    if operator.preserved or historical_fleet:
        filters = {}
        columns = vehicle_db_columns()
        if "preserved" in columns:
            filters["preserved"] = True
        if historical_fleet and "historical_fleet_id" in columns:
            filters["historical_fleet"] = historical_fleet
        return operator.vehicle_set.filter(**filters)

    return operator.vehicle_set.filter(**current_vehicle_filters(preserved=False))


def parse_mass_rows(
    default_operator,
    rows_text,
    *,
    allow_create=True,
    default_historical_fleet=None,
    default_historical_year=None,
):
    rows = []
    if not rows_text.strip():
        return rows

    delimiter = "\t" if "\t" in rows_text.splitlines()[0] else ","
    reader = csv.DictReader(StringIO(rows_text), delimiter=delimiter)
    if not reader.fieldnames:
        return rows

    for index, original_row in enumerate(reader, start=2):
        mapped = {}
        for key, value in original_row.items():
            if not key:
                continue
            mapped[normalise_header(key)] = (value or "").strip()
        if not any(mapped.values()):
            continue

        row = {
            "row_number": index,
            "raw": mapped,
            "errors": [],
            "action": "skip",
            "matched_by": "",
            "operator": default_operator,
            "operator_preview_action": "match" if default_operator else "",
            "operator_preview_label": (
                f"Match {default_operator.noc} - {default_operator.name}"
                if default_operator
                else "Leave operator unchanged"
            ),
            "pending_operator_code": "",
            "vehicle": None,
            "values": {},
            "features": None,
            "has_features": False,
            "garage_preview_action": "",
            "garage_preview_label": "",
            "pending_garage_name": "",
            "raw_livery_name": mapped.get("livery", ""),
            "resolved_livery": None,
            "unresolved_livery": "",
            "operator_was_explicit": bool(mapped.get("operator_code", "").strip()),
            "historical_fleet": default_historical_fleet,
            "historical_fleet_preview_label": (
                f"Attach to {default_historical_fleet.noc} - {default_historical_fleet.name}"
                if default_historical_fleet
                else ""
            ),
            "historical_year": default_historical_year,
            "final_code_preview": "",
        }

        if "operator_code" in mapped:
            try:
                (
                    row["operator"],
                    row["operator_preview_action"],
                    row["operator_preview_label"],
                ) = resolve_operator_preview(
                    default_operator,
                    mapped.get("operator_code"),
                    allow_create=allow_create,
                )
                if row["operator_preview_action"] == "create":
                    row["pending_operator_code"] = mapped.get("operator_code", "").strip()
            except ValueError as exc:
                row["errors"].append(str(exc))

        if "historical_fleet" in mapped and mapped.get("historical_fleet", "").strip():
            try:
                row["historical_fleet"] = resolve_existing_operator(
                    mapped.get("historical_fleet", ""), "historical_fleet"
                )
                row["historical_fleet_preview_label"] = (
                    f"Attach to {row['historical_fleet'].noc} - {row['historical_fleet'].name}"
                )
            except ValueError as exc:
                row["errors"].append(str(exc))

        if "year" in mapped and mapped.get("year", "").strip():
            try:
                row["historical_year"] = int(mapped.get("year", "").strip())
            except ValueError:
                row["errors"].append("year must be an integer")

        provided_code = mapped.get("code", "")
        provided_fleet_code = mapped.get("fleet_code", "")
        code = provided_code or provided_fleet_code
        fleet_number = None
        if mapped.get("fleet_number"):
            try:
                fleet_number = int(mapped["fleet_number"])
            except ValueError:
                row["errors"].append("fleet_number must be an integer")
        elif mapped.get("fleet_number") == "":
            fleet_number = None

        reg = mapped.get("reg", "").upper().replace(" ", "")
        prev_registration = mapped.get("prev_registration", "").upper().replace(" ", "")
        if not code:
            if fleet_number is not None:
                code = str(fleet_number)
            elif reg:
                code = reg

        external_id = mapped.get("external_id") or None
        vehicle = None
        if external_id:
            vehicle = Vehicle.objects.filter(external_id=external_id).first()
            if (
                vehicle
                and row.get("historical_year") is not None
                and (
                    vehicle.operator_id != getattr(row.get("operator"), "pk", None)
                    or vehicle.historical_fleet_id
                    != getattr(row.get("historical_fleet") or row.get("operator"), "pk", None)
                    or getattr(vehicle, "historical_fleet_year", None) != row["historical_year"]
                )
            ):
                max_len = Vehicle._meta.get_field("external_id").max_length
                suffix = f"-{row['historical_year']}"
                base = str(external_id)
                if len(base) + len(suffix) > max_len:
                    base = base[: max_len - len(suffix)]
                external_id = f"{base}{suffix}"
                vehicle = Vehicle.objects.filter(external_id=external_id).first()
            if vehicle:
                row["matched_by"] = "external_id"
        if not vehicle and code:
            queryset = _vehicle_match_queryset(
                row["operator"], row.get("historical_fleet"), row.get("historical_year")
            )
            if row.get("historical_year") is not None:
                vehicle = queryset.filter(
                    Q(code__iexact=code)
                    | Q(code__iexact=historical_code_with_year(code, row["historical_year"]))
                    | Q(fleet_code__iexact=provided_fleet_code or code)
                ).first()
            else:
                vehicle = queryset.filter(code__iexact=code).first()
            if vehicle:
                row["matched_by"] = "code"
        if not vehicle and reg:
            vehicle = _vehicle_match_queryset(
                row["operator"], row.get("historical_fleet"), row.get("historical_year")
            ).filter(reg__iexact=reg).first()
            if vehicle:
                row["matched_by"] = "registration"

        if not vehicle and not code:
            row["errors"].append("Could not determine vehicle identifier (code/fleet_num/registration)")
        elif not vehicle and not allow_create:
            row["errors"].append("No existing vehicle matched this row for mass edit")

        row["vehicle"] = vehicle
        row["action"] = "update" if vehicle else "create"
        if row["action"] == "create":
            if code:
                row["values"]["code"] = code
        elif provided_code:
            row["values"]["code"] = provided_code
        if mapped.get("external_id"):
            row["values"]["external_id"] = external_id
        if mapped.get("fleet_number"):
            row["values"]["fleet_number"] = fleet_number
        if provided_fleet_code:
            row["values"]["fleet_code"] = provided_fleet_code
        if reg:
            row["values"]["reg"] = reg
        if prev_registration:
            row["values"]["prev_registration"] = prev_registration
        if mapped.get("name"):
            row["values"]["name"] = mapped["name"]
        if mapped.get("notes"):
            row["values"]["notes"] = mapped["notes"]
        if mapped.get("branding"):
            row["values"]["branding"] = mapped["branding"]
        if mapped.get("colours"):
            row["values"]["colours"] = mapped["colours"]
        if mapped.get("slug"):
            row["values"]["slug"] = mapped["slug"]

        if row.get("historical_fleet"):
            row["values"]["historical_fleet"] = row["historical_fleet"]
            row["values"]["preserved"] = True
            if row.get("historical_year") is not None:
                if code and not provided_fleet_code:
                    row["values"]["fleet_code"] = code
                row["values"]["historical_fleet_year"] = row["historical_year"]
                row["final_code_preview"] = historical_code_with_year(
                    code, row["historical_year"]
                )
            else:
                row["final_code_preview"] = code
        elif row.get("historical_year") is not None:
            row["values"]["historical_fleet"] = row.get("operator")
            row["values"]["historical_fleet_year"] = row["historical_year"]
            row["values"]["preserved"] = True
            if code and not provided_fleet_code:
                row["values"]["fleet_code"] = code
            row["final_code_preview"] = historical_code_with_year(
                code, row["historical_year"]
            )
        elif row.get("operator") and row["operator"].preserved and "preserved" not in row["values"]:
            row["values"]["preserved"] = True
            row["final_code_preview"] = code
        else:
            row["final_code_preview"] = code

        for field in (
            "withdrawn",
            "preserved",
            "fleet_support_vehicle",
            "vor",
            "awaiting_delivery",
            "trainer_vehicle",
            "demonstrator",
        ):
            if not mapped.get(field):
                continue
            try:
                parsed_bool = coerce_bool(mapped[field])
            except ValueError as exc:
                row["errors"].append(str(exc))
            else:
                if parsed_bool is not None:
                    row["values"][field] = parsed_bool

        if mapped.get("vehicle_type"):
            try:
                row["values"]["vehicle_type"] = resolve_reference(
                    VehicleType, mapped.get("vehicle_type", ""), "vehicle_type"
                )
            except ValueError:
                if allow_create:
                    row["values"]["vehicle_type"] = mapped.get("vehicle_type", "")
                else:
                    row["errors"].append(f"Unknown vehicle_type '{mapped.get('vehicle_type', '')}'")

        if mapped.get("livery"):
            try:
                resolved_livery = resolve_reference(
                    Livery, mapped.get("livery", ""), "livery"
                )
                row["values"]["livery"] = resolved_livery
                row["resolved_livery"] = resolved_livery
            except ValueError:
                row["unresolved_livery"] = mapped.get("livery", "")

        if mapped.get("garage"):
            try:
                garage, row["garage_preview_action"], row["garage_preview_label"] = resolve_garage_preview(
                    row["operator"], mapped.get("garage", ""), allow_create=allow_create
                )
                if garage:
                    row["values"]["garage"] = garage
                    if garage.operator_id:
                        if row["operator"] is None:
                            row["operator"] = garage.operator
                            row["operator_preview_action"] = "match"
                            row["operator_preview_label"] = (
                                f"Match {garage.operator.noc} - {garage.operator.name} via depot"
                            )
                        elif (
                            not row["operator_was_explicit"]
                            and row["operator"].pk != garage.operator_id
                        ):
                            row["operator"] = garage.operator
                            row["operator_preview_action"] = "match"
                            row["operator_preview_label"] = (
                                f"Match {garage.operator.noc} - {garage.operator.name} via depot"
                            )
                        elif row["operator_was_explicit"] and row["operator"].pk != garage.operator_id:
                            row["errors"].append(
                                f"Depot '{mapped.get('garage', '')}' belongs to {garage.operator.noc}, but operator import data points to {row['operator'].noc}."
                            )
                elif row["garage_preview_action"] == "create":
                    row["pending_garage_name"] = mapped.get("garage", "").strip()
            except ValueError as exc:
                row["errors"].append(str(exc))
        
        # Handle advanced fields (engine, seating-capacity, gearbox)
        advanced_fields = {}
        if mapped.get("advanced_engine"):
            advanced_fields["engine"] = mapped.get("advanced_engine")
        if mapped.get("advanced_seating_capacity"):
            advanced_fields["seating-capacity"] = mapped.get("advanced_seating_capacity")
        if mapped.get("advanced_gearbox"):
            advanced_fields["gearbox"] = mapped.get("advanced_gearbox")
        
        if advanced_fields:
            # Merge with existing advanced data if any
            existing_advanced = row["values"].get("advanced", {})
            if isinstance(existing_advanced, dict):
                existing_advanced.update(advanced_fields)
                row["values"]["advanced"] = existing_advanced
            else:
                row["values"]["advanced"] = advanced_fields

        if mapped.get("features"):
            row["has_features"] = True
            row["features"] = [
                item.strip()
                for item in mapped.get("features", "").replace(";", ",").split(",")
                if item.strip()
            ]

        rows.append(row)

    return rows


def build_livery_mapping_rows(rows, *, manual_livery_selection):
    names = []
    seen = set()
    for row in rows:
        raw_livery = str(row.get("raw_livery_name") or "").strip()
        if not raw_livery:
            continue
        include = manual_livery_selection or bool(row.get("unresolved_livery"))
        if include and raw_livery not in seen:
            seen.add(raw_livery)
            names.append(
                {
                    "raw_name": raw_livery,
                    "selected_livery_id": (
                        str(row["resolved_livery"].pk) if row.get("resolved_livery") else ""
                    ),
                    "needs_attention": bool(row.get("unresolved_livery")),
                }
            )
    return names


def collect_livery_mappings(post_data, mapping_rows):
    mappings = {}
    for index, item in enumerate(mapping_rows):
        mappings[item["raw_name"]] = post_data.get(f"livery_map_{index}", "").strip()
    return mappings


def commit_mass_rows(
    default_operator,
    rows,
    *,
    allow_create=True,
    livery_mappings=None,
    default_historical_fleet=None,
    default_historical_year=None,
):
    created = 0
    updated = 0
    errors = 0
    livery_mappings = livery_mappings or {}

    for row in rows:
        if row["errors"]:
            errors += 1
            continue
        if row["action"] == "create" and not allow_create:
            row["errors"].append("Mass edit only supports existing vehicles")
            errors += 1
            continue

        try:
            with transaction.atomic():
                row_operator = row.get("operator") or default_operator
                row_historical_fleet = (
                    row.get("historical_fleet") or default_historical_fleet
                )
                row_historical_year = row.get("historical_year")
                if row_historical_year is None:
                    row_historical_year = default_historical_year
                target_vehicle = row["vehicle"]
                if not row.get("operator") and row.get("pending_operator_code"):
                    row_operator = create_operator_from_code(row["pending_operator_code"])
                    row["operator"] = row_operator
                garage_operator = row_operator or (target_vehicle.operator if target_vehicle else None)
                if row.get("pending_garage_name") and not row["values"].get("garage"):
                    if not garage_operator:
                        raise ValueError(
                            "Cannot create a depot unless the row resolves to an operator."
                        )
                    row["values"]["garage"] = create_garage_for_operator(
                        garage_operator, row["pending_garage_name"]
                    )

                if row.get("raw_livery_name"):
                    selected_livery_id = livery_mappings.get(row["raw_livery_name"], "")
                    if selected_livery_id:
                        row["values"]["livery"] = Livery.objects.get(pk=int(selected_livery_id))
                    elif row.get("resolved_livery"):
                        row["values"]["livery"] = row["resolved_livery"]
                    else:
                        notes = row["values"].get("notes", "")
                        row["values"]["notes"] = "\n".join(
                            part
                            for part in (notes, f"Imported livery text: {row['raw_livery_name']}")
                            if part
                        )
                    
                    # Clear "Imported livery text:" note if livery is now resolved
                    if row["values"].get("livery") and target_vehicle:
                        existing_notes = target_vehicle.notes or ""
                        if "Imported livery text:" in existing_notes:
                            cleaned_notes = "\n".join(
                                line for line in existing_notes.split("\n")
                                if not line.startswith("Imported livery text:")
                            ).strip()
                            row["values"]["notes"] = cleaned_notes

                if not target_vehicle and not row_operator:
                    raise ValueError(
                        "Cannot create a vehicle unless the row provides or resolves to an operator."
                    )

                vehicle = target_vehicle or Vehicle(operator=row_operator)
                if row_operator:
                    vehicle.operator = row_operator
                if row_historical_fleet:
                    vehicle.historical_fleet = row_historical_fleet
                    if row_historical_year is not None:
                        vehicle.historical_fleet_year = row_historical_year
                        incoming_code = row["values"].get("code") or vehicle.code
                        if incoming_code:
                            desired_code = historical_code_with_year(
                                incoming_code,
                                row_historical_year,
                            )
                            if row["action"] == "create" or vehicle.code != desired_code:
                                row["values"]["code"] = desired_code
                            vehicle.slug = slugify(vehicle_slug(vehicle))
                    if "preserved" not in row["values"]:
                        row["values"]["preserved"] = True
                elif row_historical_year is not None:
                    vehicle.historical_fleet = row_operator
                    vehicle.historical_fleet_year = row_historical_year
                    if "preserved" not in row["values"]:
                        row["values"]["preserved"] = True
                    incoming_code = row["values"].get("code") or vehicle.code
                    if incoming_code:
                        desired_code = historical_code_with_year(
                            incoming_code,
                            row_historical_year,
                        )
                        if row["action"] == "create" or vehicle.code != desired_code:
                            row["values"]["code"] = desired_code
                        vehicle.slug = slugify(vehicle_slug(vehicle))

                # If an imported external identifier lands on an existing vehicle via registration,
                # persist that external id so future imports can link directly without relying on
                # registration-only matching.
                if (
                    row.get("matched_by") == "registration"
                    and row["values"].get("external_id")
                    and not vehicle.external_id
                ):
                    vehicle.external_id = row["values"]["external_id"]

                for field, value in row["values"].items():
                    if field == "vehicle_type" and isinstance(value, str):
                        value = get_or_create_vehicle_type(value)
                    setattr(vehicle, field, value)

                vehicle.is_manual = True
                vehicle.manual_updated_at = timezone.now()
                vehicle.save()

                if row["has_features"]:
                    feature_objs = []
                    for feature in row["features"] or []:
                        if feature.isdigit():
                            obj = VehicleFeature.objects.filter(pk=int(feature)).first()
                            if not obj:
                                raise ValueError(f"Unknown feature id '{feature}'")
                        else:
                            obj, _ = VehicleFeature.objects.get_or_create(name=feature)
                        feature_objs.append(obj)
                    vehicle.features.set(feature_objs)

                if row["action"] == "create":
                    created += 1
                else:
                    updated += 1
        except Exception as exc:
            row["errors"].append(str(exc))
            errors += 1

    return created, updated, errors


def create_operator_from_code(code):
    code = str(code or "").strip().upper()
    operator = Operator.objects.filter(noc__iexact=code).first()
    if operator:
        return operator
    slug_base = slugify(code) or "operator"
    slug = slug_base
    suffix = 2
    while Operator.objects.filter(slug=slug).exists():
        slug = f"{slug_base}-{suffix}"
        suffix += 1
    return Operator.objects.create(
        noc=code,
        name=code,
        slug=slug,
        is_manual=True,
        manual_updated_at=timezone.now(),
    )


def create_garage_for_operator(operator, garage_name):
    garage_name = str(garage_name or "").strip()
    trimmed = garage_name[4:].strip() if garage_name.upper().startswith("GSC ") else garage_name
    garage = Garage.objects.filter(operators=operator).filter(
        Q(name__iexact=trimmed) | Q(code__iexact=trimmed)
    ).first()
    if garage:
        return garage
    garage = Garage.objects.create(
        name=trimmed,
        code=trimmed,
        is_manual=True,
        manual_updated_at=timezone.now(),
    )
    garage.operators.add(operator)
    return garage


def get_or_create_vehicle_type(value):
    text = str(value or "").strip()
    if not text:
        return None
    vehicle_type = VehicleType.objects.filter(name__iexact=text).first()
    if vehicle_type:
        return vehicle_type
    return VehicleType.objects.create(
        name=text,
        is_manual=True,
        manual_updated_at=timezone.now(),
    )


def _match_existing_vehicle(operator, field_name, value):
    if operator:
        return (
            operator.vehicle_set.filter(**current_vehicle_filters(preserved=False))
            .filter(**{f"{field_name}__iexact": value})
            .first()
        )

    matches = list(
        Vehicle.objects.filter(**current_vehicle_filters(preserved=False))
        .filter(**{f"{field_name}__iexact": value})[:2]
    )
    if len(matches) == 1:
        return matches[0]
    return None
