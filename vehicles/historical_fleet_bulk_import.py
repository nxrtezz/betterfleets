"""Spreadsheet helpers for historical and live fleet imports."""

from __future__ import annotations

import csv
import datetime
from dataclasses import dataclass
from io import StringIO
from urllib.parse import urlencode

from django.db import transaction
from django.db.models import Q
from django.urls import reverse

from bustimes.models import Garage
from vehicles.models import HistoricalVehicle, Livery, Vehicle, VehicleType

COLUMN_KEYS = (
    "noc",
    "garage_name",
    "fleet_number",
    "reg",
    "prev_reg",
    "vehicle_type_name",
    "livery_name",
    "branding",
    "rear_ad",
    "joined_fleet_date",
    "left_fleet_date",
    "preserved",
    "fsv",
    "trainer",
    "demo",
    "notes",
    "slug",
)


@dataclass
class ImportIssue:
    message: str
    line_number: int | None = None
    missing_kind: str | None = None
    value: str | None = None
    create_url: str | None = None


def _normalize_cells(row: list[str], n: int = 15) -> list[str]:
    row = [c.strip() for c in row]
    while len(row) < n:
        row.append("")
    return row[:n]


def _parse_fleet_number(raw: str) -> tuple[int | None, str | None]:
    s = (raw or "").strip()
    if not s:
        return None, None
    if s.isdigit():
        return int(s), None
    return None, f"fleet_number must be numeric or empty, got {raw!r}"


def _should_skip_header_row(first_row: list[str]) -> bool:
    if not first_row or not first_row[0].strip():
        return False
    key = first_row[0].strip().lower().replace(" ", "_")
    return key in (
        "noc",
        "fleet_number",
        "fleet_num",
        "reg",
        "#",
    )


def parse_pasted_rows(text: str) -> tuple[list[list[str]], str | None]:
    """Return list of cell rows and optional error message."""
    lines: list[str] = []
    for raw in text.splitlines():
        s = raw.strip()
        if not s or s.startswith("#"):
            continue
        lines.append(s)
    if not lines:
        return [], "No data rows found."
    delim = "\t" if "\t" in lines[0] else ","
    parsed: list[list[str]] = []
    for line in lines:
        try:
            row = next(csv.reader([line], delimiter=delim))
        except csv.Error as e:
            return [], f"CSV parse error: {e}"
        parsed.append(_normalize_cells(row))
    if parsed and _should_skip_header_row(parsed[0]):
        parsed = parsed[1:]
    return parsed, None


def _resolve_livery_id(
    operator_id: str,
    raw_name: str,
    cache: dict[str, int | None],
) -> int | None:
    """Match a published Livery by numeric id or exact name; prefer liveries used by this operator."""
    s = (raw_name or "").strip()
    if not s:
        return None
    key = s.lower()
    if key in cache:
        return cache[key]
    if s.isdigit():
        liv = Livery.objects.filter(pk=int(s), published=True).first()
        cache[key] = liv.pk if liv else None
        return cache[key]
    candidates = list(Livery.objects.filter(name__iexact=s, published=True).order_by("id"))
    if not candidates:
        cache[key] = None
        return None
    if len(candidates) == 1:
        cache[key] = candidates[0].pk
        return cache[key]
    for liv in candidates:
        if liv.vehicle_set.filter(operator_id=operator_id, historical_fleet__isnull=True).exists():
            cache[key] = liv.pk
            return cache[key]
    cache[key] = candidates[0].pk
    return cache[key]


def _resolve_vehicle_type_id(raw_name: str, cache: dict[str, int | None]) -> int | None:
    s = (raw_name or "").strip()
    if not s:
        return None
    key = s.lower()
    if key in cache:
        return cache[key]
    vehicle_type = VehicleType.objects.filter(name__iexact=s).first()
    cache[key] = vehicle_type.pk if vehicle_type else None
    return cache[key]


def _resolve_garage_id(
    operator_id: str,
    raw_name: str,
    cache: dict[tuple[str, str], int | None],
) -> int | None:
    s = (raw_name or "").strip()
    if not s:
        return None
    key = (operator_id, s.lower())
    if key in cache:
        return cache[key]
    garage = (
        Garage.objects.filter(operators=operator_id)
        .filter(Q(name__iexact=s) | Q(code__iexact=s))
        .first()
    )
    cache[key] = garage.pk if garage else None
    return cache[key]


def _create_missing_issue(kind: str, value: str, line_number: int, operator_id: str | None = None):
    label_map = {
        "livery": "Livery",
        "vehicle_type": "Vehicle type",
        "garage": "Garage",
    }
    url_name_map = {
        "livery": "admin:vehicles_livery_add",
        "vehicle_type": "admin:vehicles_vehicletype_add",
        "garage": "admin:bustimes_garage_add",
    }
    params = {}
    if kind == "garage":
        params["name"] = value
        params["code"] = value
        if operator_id:
            params["operator"] = operator_id
    else:
        params["name"] = value

    create_url = reverse(url_name_map[kind])
    if params:
        create_url = f"{create_url}?{urlencode(params)}"

    return ImportIssue(
        line_number=line_number,
        missing_kind=kind,
        value=value,
        create_url=create_url,
        message=f"Line {line_number}: {label_map[kind]} '{value}' not found.",
    )


def build_historical_vehicles(
    operator_id: str,
    rows: list[list[str]],
) -> tuple[list[HistoricalVehicle], list[str]]:
    instances: list[HistoricalVehicle] = []
    errors: list[str] = []
    livery_cache: dict[str, int | None] = {}
    for line_no, cells in enumerate(rows, start=1):
        if not any((c or "").strip() for c in cells):
            continue
        fn, err = _parse_fleet_number(cells[0])
        if err:
            errors.append(f"Line {line_no}: {err}")
            continue
        d = dict(zip(COLUMN_KEYS, cells, strict=True))
        livery_id = _resolve_livery_id(operator_id, d["livery_name"], livery_cache)
        code = (
            (d["code"] or "").strip()
            or (d["fleet_code"] or "").strip()
            or (str(fn) if fn is not None else "")
            or (d["reg"] or "").strip()
        )
        if not code:
            code = f"row-{line_no}"

        # Parse dates
        joined_date = None
        left_date = None
        if d["joined_fleet_date"]:
            try:
                joined_date = datetime.datetime.strptime(d["joined_fleet_date"], "%d-%m-%Y").date()
            except ValueError:
                errors.append(f"Line {line_no}: Invalid joined_fleet_date format (use dd-mm-yyyy)")
                continue
        if d["left_fleet_date"]:
            try:
                left_date = datetime.datetime.strptime(d["left_fleet_date"], "%d-%m-%Y").date()
            except ValueError:
                errors.append(f"Line {line_no}: Invalid left_fleet_date format (use dd-mm-yyyy)")
                continue

        v = HistoricalVehicle(
            operator_id=operator_id,
            fleet_number=fn,
            fleet_code=(d["fleet_code"] or "")[:24],
            reg=(d["reg"] or "")[:24],
            prev_registration=(d["prev_reg"] or "")[:24],
            code=code[:255],
            branding=(d["branding"] or "")[:255],
            rear_advert=(d["rear_ad"] or "")[:255],
            notes=(d["notes"] or "")[:255],
            livery_id=livery_id,
            colours=(d["colours"] or "")[:255],
            joined_fleet_date=joined_date,
            left_fleet_date=left_date,
            fleet_support_vehicle=(d["fsv"] or "").lower() in ("yes", "true", "1"),
            trainer_vehicle=(d["trainer"] or "").lower() in ("yes", "true", "1"),
            slug=(d["slug"] or "")[:255],
        )
        vtn = (d["vehicle_type_name"] or "").strip()
        if vtn:
            vt = VehicleType.objects.filter(name__iexact=vtn).first()
            if vt:
                v.vehicle_type_id = vt.pk
        gn = (d["garage_name"] or "").strip()
        if gn:
            g = (
                Garage.objects.filter(operators=operator_id)
                .filter(Q(name__iexact=gn) | Q(code__iexact=gn))
                .first()
            )
            if g:
                v.garage_id = g.id
        instances.append(v)
    return instances, errors


# NOTE: The following functions are not currently used in the codebase
# and were written for a HistoricalFleet feature that appears to have been removed.
# They are kept here for reference but commented out to avoid import errors.

# def build_vehicles(
#     fleet_id: int,
#     operator_id: str,
#     rows: list[list[str]],
# ) -> tuple[list[Vehicle], list[str]]:
#     instances: list[Vehicle] = []
#     errors: list[str] = []
#     livery_cache: dict[str, int | None] = {}
#     op_id = operator_id
#     for line_no, cells in enumerate(rows, start=1):
#         if not any((c or "").strip() for c in cells):
#             continue
#         fn, err = _parse_fleet_number(cells[0])
#         if err:
#             errors.append(f"Line {line_no}: {err}")
#             continue
#         d = dict(zip(COLUMN_KEYS, cells, strict=True))
#         livery_id = _resolve_livery_id(op_id, d["livery_name"], livery_cache)
#         code = (
#             (d["code"] or "").strip()
#             or (d["fleet_code"] or "").strip()
#             or (str(fn) if fn is not None else "")
#             or (d["reg"] or "").strip()
#         )
#         if not code:
#             code = f"row-{line_no}"

#         v = Vehicle(
#             operator_id=op_id,
#             historical_fleet_id=fleet_id,
#             fleet_number=fn,
#             fleet_code=(d["fleet_code"] or "")[:24],
#             reg=(d["reg"] or "")[:24],
#             code=code[:255],
#             branding=(d["branding"] or "")[:255],
#             name=(d["name"] or "")[:255],
#             notes=(d["notes"] or "")[:255],
#             livery_id=livery_id,
#             colours=(d["colours"] or "")[:255],
#             slug=(d["slug"] or "")[:255],
#         )
#         vtn = (d["vehicle_type_name"] or "").strip()
#         if vtn:
#             vt = VehicleType.objects.filter(name__iexact=vtn).first()
#             if vt:
#                 v.vehicle_type_id = vt.pk
#         gn = (d["garage_name"] or "").strip()
#         if gn:
#             g = (
#                 Garage.objects.filter(operators=op_id)
#                 .filter(Q(name__iexact=gn) | Q(code__iexact=gn))
#                 .first()
#             )
#             if g:
#                 v.garage_id = g.id
#         instances.append(v)
#     return instances, errors


def build_live_vehicles(operator, rows: list[list[str]]):
    instances: list[Vehicle] = []
    errors: list[str] = []
    issues: list[ImportIssue] = []
    livery_cache: dict[str, int | None] = {}
    vehicle_type_cache: dict[str, int | None] = {}
    garage_cache: dict[tuple[str, str], int | None] = {}
    seen_codes: set[str] = set()

    for line_no, cells in enumerate(rows, start=1):
        if not any((c or "").strip() for c in cells):
            continue
        fn, err = _parse_fleet_number(cells[0])
        if err:
            errors.append(f"Line {line_no}: {err}")
            continue
        d = dict(zip(COLUMN_KEYS, cells, strict=True))
        code = (
            (d["code"] or "").strip()
            or (d["fleet_code"] or "").strip()
            or (str(fn) if fn is not None else "")
            or (d["reg"] or "").strip()
        )
        if not code:
            code = f"row-{line_no}"

        normalized_code = code.lower()
        if normalized_code in seen_codes:
            errors.append(f"Line {line_no}: Duplicate code '{code}' in this import.")
            continue
        seen_codes.add(normalized_code)

        if Vehicle.objects.filter(
            operator=operator,
            historical_fleet__isnull=True,
            code__iexact=code,
        ).exists():
            errors.append(
                f"Line {line_no}: {operator} already has a live vehicle with code '{code}'."
            )
            continue

        missing_reference = False

        livery_name = (d["livery_name"] or "").strip()
        livery_id = _resolve_livery_id(operator.pk, livery_name, livery_cache)
        if livery_name and not livery_id:
            issues.append(_create_missing_issue("livery", livery_name, line_no))
            missing_reference = True

        vehicle_type_name = (d["vehicle_type_name"] or "").strip()
        vehicle_type_id = _resolve_vehicle_type_id(vehicle_type_name, vehicle_type_cache)
        if vehicle_type_name and not vehicle_type_id:
            issues.append(_create_missing_issue("vehicle_type", vehicle_type_name, line_no))
            missing_reference = True

        garage_name = (d["garage_name"] or "").strip()
        garage_id = _resolve_garage_id(operator.pk, garage_name, garage_cache)
        if garage_name and not garage_id:
            issues.append(_create_missing_issue("garage", garage_name, line_no, operator.pk))
            missing_reference = True

        if missing_reference:
            continue

        instances.append(
            Vehicle(
                operator=operator,
                fleet_number=fn,
                fleet_code=(d["fleet_code"] or "")[:24],
                reg=(d["reg"] or "")[:24],
                code=code[:255],
                branding=(d["branding"] or "")[:255],
                name=(d["name"] or "")[:255],
                notes=(d["notes"] or "")[:255],
                livery_id=livery_id,
                colours=(d["colours"] or "")[:255],
                vehicle_type_id=vehicle_type_id,
                garage_id=garage_id,
                slug=(d["slug"] or "")[:255],
                is_manual=True,
            )
        )

    return instances, errors, issues


def bulk_import_historical_vehicles(operator_id: str, text: str) -> tuple[int, list[str]]:
    """Returns (created_count, error_messages)."""
    rows, parse_err = parse_pasted_rows(text)
    if parse_err:
        return 0, [parse_err]
    instances, row_errors = build_historical_vehicles(operator_id, rows)
    if not instances and not row_errors:
        return 0, ["No valid vehicle rows to import."]
    if row_errors and not instances:
        return 0, row_errors
    with transaction.atomic():
        for v in instances:
            v.save()
    return len(instances), row_errors


# NOTE: The following functions are not currently used in the codebase
# and were written for a HistoricalFleet feature that appears to have been removed.
# They are kept here for reference but commented out to avoid import errors.

# def bulk_import_vehicles(fleet_id: int, operator_id: str, text: str) -> tuple[int, list[str]]:
#     """Returns (created_count, error_messages)."""
#     rows, parse_err = parse_pasted_rows(text)
#     if parse_err:
#         return 0, [parse_err]
#     instances, row_errors = build_vehicles(fleet_id, operator_id, rows)
#     if not instances and not row_errors:
#         return 0, ["No valid vehicle rows to import."]
#     if row_errors and not instances:
#         return 0, row_errors
#     with transaction.atomic():
#         for v in instances:
#             v.save()
#     return len(instances), row_errors


def bulk_import_live_vehicles(operator, text: str):
    rows, parse_err = parse_pasted_rows(text)
    if parse_err:
        return 0, [parse_err], []
    instances, row_errors, issues = build_live_vehicles(operator, rows)
    if not instances and not row_errors and not issues:
        return 0, ["No valid vehicle rows to import."], []
    if instances:
        with transaction.atomic():
            for vehicle in instances:
                vehicle.save()
    return len(instances), row_errors, issues


def rows_text_from_uploaded_workbook(uploaded_file) -> str:
    """Convert an uploaded .xlsx (first sheet) or UTF-8 .csv to TSV for :func:`parse_pasted_rows`."""
    from openpyxl import load_workbook

    if not uploaded_file:
        return ""

    filename = (uploaded_file.name or "").lower()
    if filename.endswith(".csv"):
        try:
            content = uploaded_file.read().decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise ValueError("CSV upload must be UTF-8 encoded") from exc
        return content.strip()

    if not filename.endswith(".xlsx"):
        raise ValueError("Upload must be a .xlsx or .csv file")

    workbook = load_workbook(uploaded_file, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return ""

    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    if not any(headers):
        return ""

    output = StringIO()
    writer = csv.writer(output, delimiter="\t", lineterminator="\n")
    writer.writerow(headers)
    for row in rows[1:]:
        values = ["" if value is None else str(value).strip() for value in row[: len(headers)]]
        if any(values):
            writer.writerow(values)
    return output.getvalue()


# NOTE: The following function is not currently used in the codebase
# and was written for a HistoricalFleet feature that appears to have been removed.
# It is kept here for reference but commented out to avoid import errors.

# def export_fleet_rows(fleet_id: int) -> list[tuple]:
#     """Rows matching :data:`COLUMN_KEYS` for spreadsheet export."""
#     rows = []
#     qs = Vehicle.objects.filter(historical_fleet_id=fleet_id).select_related("vehicle_type", "garage", "livery").order_by(
#         "fleet_number", "fleet_code", "reg", "id"
#     )
#     for v in qs:
#         vt_name = v.vehicle_type.name if v.vehicle_type_id else ""
#         livery_label = v.livery.name if v.livery_id else ""
#         garage_label = ""
#         if v.garage_id:
#             garage_label = (v.garage.name or v.garage.code or "").strip()
#         rows.append(
#             (
#                 v.fleet_number if v.fleet_number is not None else "",
#                 v.fleet_code or "",
#                 v.reg or "",
#                 v.code or "",
#                 v.branding or "",
#                 v.name or "",
#                 v.notes or "",
#                 vt_name,
#                 livery_label,
#                 v.colours or "",
#                 garage_label,
#                 v.slug or "",
#             )
#         )
#     return rows


def export_operator_fleet_rows(operator) -> list[tuple]:
    rows = []
    qs = (
        Vehicle.objects.filter(operator=operator, historical_fleet__isnull=True)
        .select_related("vehicle_type", "garage", "livery")
        .order_by("fleet_number", "fleet_code", "reg", "id")
    )
    for v in qs:
        vt_name = v.vehicle_type.name if v.vehicle_type_id else ""
        livery_label = v.livery.name if v.livery_id else ""
        garage_label = ""
        if v.garage_id:
            garage_label = (v.garage.name or v.garage.code or "").strip()
        rows.append(
            (
                v.fleet_number if v.fleet_number is not None else "",
                v.fleet_code or "",
                v.reg or "",
                v.code or "",
                v.branding or "",
                v.name or "",
                v.notes or "",
                vt_name,
                livery_label,
                v.colours or "",
                garage_label,
                v.slug or "",
            )
        )
    return rows


def build_template_workbook(data_rows=None):
    """Blank template or pre-filled data (e.g. current fleet)."""
    from openpyxl import Workbook

    workbook = Workbook()
    ws = workbook.active
    ws.title = "Vehicles"
    ws.append(list(COLUMN_KEYS))
    for row in data_rows or ():
        ws.append(row)
    ws.freeze_panes = "A2"

    instructions = workbook.create_sheet("Instructions")
    instructions.append(["Column", "Notes"])
    notes = {
        "noc": "Operator NOC code",
        "garage_name": "Garage name or code for this operator",
        "fleet_number": "Integer, optional",
        "reg": "Registration, no spaces",
        "prev_reg": "Previous registration",
        "vehicle_type_name": "Matches VehicleType name when possible",
        "livery_name": "Published livery name (case-insensitive), or numeric livery id",
        "branding": "Branding text",
        "rear_ad": "Rear advertisement text",
        "joined_fleet_date": "Date vehicle joined fleet (dd-mm-yyyy format)",
        "left_fleet_date": "Date vehicle left fleet (dd-mm-yyyy format)",
        "preserved": "Yes if preserved",
        "fsv": "Yes if fleet support vehicle",
        "trainer": "Yes if trainer vehicle",
        "demo": "Yes if demonstrator",
        "notes": "Short notes",
        "slug": "URL slug for the vehicle",
    }
    for key in COLUMN_KEYS:
        instructions.append([key, notes.get(key, "")])
    return workbook
